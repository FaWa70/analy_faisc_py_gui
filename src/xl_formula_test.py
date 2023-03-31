import openpyxl as xl  # for writing excel files with multi line header
import numpy as np
from scipy.stats import sem, t

truc = sorted(xl.utils.FORMULAE) # type is frozenset
print("AVERAGE", "AVERAGE" in truc)

print("NB", "NB" in truc)
print("NUMBER", "NUMBER" in truc)

print("STDEV STDEVA", "STDEV STDEVA" in truc)
print("TINV", "TINV" in truc)

print("truc: ")
print("len(truc): ", len(truc))
for count in range(len(truc)//4):
    print(truc[4 * count:4 * count + 4])

"""
'STDEV STDEVA', 'STDEVP', 'STDEVPA STEYX', LEN = NBCAR
"""

def col_2_str(col_no):  # mais ca existe aussi comme xl.utils.get_column_letter(num)
    out = ""
    while col_no > 0:
        col_no, remainder = divmod(col_no-1, 26)
        out += chr(65 + remainder)
    return out  # string of type "AC"


mywb = xl.Workbook()  # workbook
ws1 = mywb.active

ws1['A2'] = 2
ws1['A3'].value = 4
ws1['A4'] = 4
ws1['A5'].value = 2
ws1['C2'] = 2 + 2
ws1['C3'].value = 4 + 2
ws1['C4'] = 4 + 2
ws1['C5'].value = 2 + 2

ws1['A7'] = '=AVERAGE(A2:A5)'
confidence = 0.6827
ws1['B7'] = confidence
dataa = np.array([2, 4, 4, 2])
abs_std_err_of_mean = sem(dataa) * t.ppf((1 + confidence) / 2, df=len(dataa)-1)
print("abs_std_err_of_mean: ", abs_std_err_of_mean)
ws1['A8'] = '=COUNT(A2:A5)'
ws1['A9'] = '=SQRT(COUNT(A2:A5))'
ws1['A10'] = '=STDEVA(A2:A5)'
ws1['A12'] = '=TINV( 1-B7 , COUNT(A2:A5)-1)'  # 2-sided
ws1['A13'] = t.ppf((1 + confidence) / 2, df=len(dataa)-1)  # 1-sided

ws1.cell(row=8, column=3).value = '=MIN(A2:A5)'
ws1.cell(row=9, column=3).value = ('=MAX(' +
                                   xl.utils.get_column_letter(1) + str(2) + ':' +
                                   xl.utils.get_column_letter(3) + str(5) + ')')
ws1.cell(row=10, column=3).value = ('=MAX(' +
                                    col_2_str(1) + str(2) + ':' +
                                    col_2_str(3) + str(5) + ')')

mywb.save("formula_test.xlsx")
