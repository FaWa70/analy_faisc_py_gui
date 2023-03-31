"""
@author: Wagner
"""

import glob
import sys  # still useful ? yes if not it would be grey
import os
import struct
import re

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget,
                             QGridLayout, QVBoxLayout, QHBoxLayout, QSplitter,
                             QSlider, QTabWidget, QButtonGroup, QRadioButton,
                             QPushButton, QCheckBox, QLabel, QFileDialog, QLineEdit,
                             QSpinBox, QProgressBar, QComboBox, QTableWidget, QTableWidgetItem,
                             QSizePolicy, QScrollArea)
from PyQt5.QtCore import Qt, QRectF  # for the focus from keyboard
from PyQt5.QtGui import QFont, QFontMetricsF, QIcon

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure  # better do not use pyplot (but it's possible)
import matplotlib.pyplot as plt
from scipy import ndimage as ndi
from scipy.optimize import minimize, leastsq   # use minimize_skalar ?
from scipy.stats import sem, t  # t is the student t distribution (all sorts in the object)
from skimage.measure import block_reduce as SkiMeasBR

from src.fitting_module_v2 import gauss2D_cst_offs, sot_r_gau, gaussEll2D_cst_offs

import numpy as np
import datetime as dt
import time as ti
import openpyxl as xl  # for writing excel files with multi line header

# define global variables at module level (if useful for shorter names compared to definition in __init__)
rel_sat_limit = 0.95  # 1.00 is really saturated
confidence = 0.6827
crop_ulx = 0
crop_uly = 0
auto_crop_ulx = 0
auto_crop_uly = 0
frame_width = 0
ti0 = 0  # time in secs used for analyzing one image
good_idx = []  # list of indices of the images to analyze in the stack
roiS = np.ones((3, 4))*np.nan  # each line is: xfrom, xto, yfrom, yto (include all indices)
history_length = 10  # history_length : keep history_length sets of fit parameters (initialisations and results)
b_fits = history_length * [[None, None, None]]  # (history_length, 3) list of lists with elements dict (later)
results = {}
mywb = xl.Workbook()  # workbook
f_name = ""

def col_2_str(col_no):  ## there is also an official version in openpyxl.utils
    print("col_2_str called")

    out = ""
    while col_no > 0:
        col_no, remainder = divmod(col_no - 1, 26)
        out += chr(65 + remainder)
    return out  # string of type "AC"


class PathLabel(QLabel):
    """Use setLongText instead of setText for a usual label"""
    def __init__(self, parent=None):
        super(PathLabel, self).__init__(parent)
        self.long_text = ""

    def make_short_text(self):
        print("make_short_text called")

        """works fine but is not perfect for fnames as the middle is hidden.
        -> better hide the middle of the path but not long filenames"""
        font_m = QFontMetricsF(self.font())  # F or not ?
        avail_width = self.width() - 3  # - 3 px for a little space at the end
        short_text = font_m.elidedText(self.long_text, Qt.ElideMiddle, avail_width)
        return short_text

    def setLongText(self, text_in):
        print("setLongText called")

        """Use this instead of setText for a usual label"""
        self.long_text = text_in
        self.setToolTip(text_in)  # tooltip shows the full text
        short_text = self.make_short_text()
        self.setText(short_text)

    def resizeEvent(self, *args, **kwargs):
        print("resizeEvent called")

        short_text = self.make_short_text()
        self.setText(short_text)


class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)

        # Get screen size.
        scr_si = QApplication.desktop().size()
        scr_si_x = scr_si.width()
        scr_si_y = scr_si.height()
        print("scr_si_x = ", scr_si_x)
        print("scr_si_y = ", scr_si_y)

        # Define window size, position and title.
        self.left = 5
        self.top = 40
        self.width = scr_si_x - 50
        self.height = scr_si_y - self.top - 50
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.title = "V5 - This analyzes peaked beam profiles (64-image wcf file, or similar images in folder)"
        self.setWindowTitle(self.title)
        # # Put an icon on the window: use https://icoconvert.com/ to create an .ico from .png 256x256 px
        # script_dir = os.path.dirname(os.path.realpath(__file__))
        # self.setWindowIcon(QIcon(script_dir + os.path.sep + 'help.ico'))

        # Instantiate the tab component and set as Central.
        self.tw = MyTableWidget(self)
        self.setCentralWidget(self.tw)

        self.statBar = self.statusBar()
        self.progBar = QProgressBar()
        self.statBar.addPermanentWidget(self.progBar)  # Layout is automatic
        self.statBar.showMessage('Ready')
        self.progBar.setMinimum(1)  # if it's not done here, it has to be done at different places later on
        self.progBar.setMaximum(64)

        self.show()


########################    ############################
class MyTableWidget(QWidget):
    def __init__(self, parent):
        print('__init__ called')
        super().__init__(parent)

        global tab_idx_curr, tab_idx_old
        global backg
        global b_max  # index 0: x-position, 1: y-position, 2: z-value
        b_max = -100 * np.ones((3,))
        backg = -100  # error value
        tab_idx_curr = 0
        tab_idx_old = 0

        # define special fonts
        section_h_font = QFont("Times", 10, QFont.Bold)

        # Initialize tab screen
        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self.onTabChange)

        # Add pages (tab1, tab2...) to tabWidget (tabs)
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, "Load and pre-process images")  # includes noise subtraction
        self.tab2 = QWidget()
        self.tabs.addTab(self.tab2, 'Full analysis of one image')
        self.tab3 = QWidget()
        self.tabs.addTab(self.tab3, "Stack statistics and export")  # includes histograms
        self.tab4 = QWidget()
        self.tabs.addTab(self.tab4, "Z-evolution of a parameter")  # includes histograms
        self.tab5 = QWidget()
        self.tabs.addTab(self.tab5, "How to use this program")  # short instructions

        def add_widgets_to_tab1():  # Create widgets of 1st tab : inspect
            print('add_widgets_to_tab1 called')
            # For defining the wcf file
            self.BtLoadFile1 = QPushButton('Load file')
            # self.BtLoadFile1.setFixedSize(80, 30)
            self.BtLoadFile1.clicked.connect(self.Load_File1)
            self.LbFolderName = PathLabel(
               '  Select a wcf-file with 64 images or an image file in a folder of similar images.  ')  # label (actif)
            # self.LbFolderName = QLabel(
            #    '  Select a wcf-file with 64 images or an image file in a folder of similar images.  ')  # label (actif)
            self.LbImageSize = QLabel('Image resolution before crop')  # label (actif) - for image size information

            # For defining the pixel size (replace by combo box?)
            self.LbPiSi = QLabel("Pixel size:")  # label
            self.EdPiSi = QLineEdit('0.65')
            self.EdPiSi.setFixedSize(40, 20)
            self.LbPiSiUnit = QLabel(" µm ")  # label
            self.LbPiSiUnit.setFixedSize(40, 20)
            # self.LbPiSiUnit.setHorizontalSpacing(self, '2')

            # For selecting a file in the stack
            self.BtNumDwn = QPushButton('<-')
            self.BtNumDwn.clicked.connect(self.display_down)
            self.BtNumDwn.setFixedSize(100, 30)
            self.SlNum = QSlider(Qt.Horizontal, self)
            self.SlNum.valueChanged.connect(self.display)
            self.BtNumUp = QPushButton(' -> ')
            self.BtNumUp.clicked.connect(self.display_up)
            self.BtNumUp.setFixedSize(100, 30)
            self.LbImNo = QLabel(" N.A.")

            # saturation
            self.LbPixSat = QLabel("Saturated pixels count")

            # Radiobutton group for choosing the view ("frame by frame" or "mosaic overview")
            rdbtViewLayout = QHBoxLayout()
            self.rdbtwidgetView = QWidget(self)
            self.rdbtwidgetView.setLayout(rdbtViewLayout)
            self.rdbtGroupView = QButtonGroup(self.rdbtwidgetView)
            self.rdbtFbF = QRadioButton('Frame by Frame')
            self.rdbtGroupView.addButton(self.rdbtFbF, 0)
            rdbtViewLayout.addWidget(self.rdbtFbF)
            self.rdbtFbF.setChecked(True)
            self.rdbtOv = QRadioButton('Overview')
            self.rdbtGroupView.addButton(self.rdbtOv, 1)
            rdbtViewLayout.addWidget(self.rdbtOv)
            self.rdbtGroupView.buttonClicked[int].connect(self.view_changed)
            self.LbView = QLabel('View : ')

            # For plotting the image
            self.figureIM1 = Figure()
            self.canvasIM1 = FigureCanvas(self.figureIM1)
            self.toolbarIM1 = NavigationToolbar(self.canvasIM1, self)  # Canvas widget and a parent
            # Connect a function that gets the key that is pressed
            self.canvasIM1.mpl_connect('key_press_event', self.GetKeyIm1)
            # Configure the canvas widget to process keyboard events
            self.canvasIM1.setFocusPolicy(Qt.StrongFocus)  # listen to keyboard too!
            # Connect a function that gets the mouse position on every move
            self.canvasIM1.mpl_connect('motion_notify_event', self.GetMouseIm1)

            # For cropping
            self.CbCrop = QCheckBox("apply crop margins")
            self.CbCrop.stateChanged.connect(self.crop_or_uncrop)
            self.LbUpper = QLabel("pixels")
            self.SbUpper = QSpinBox()
            self.SbLeft = QSpinBox()
            self.SbRight = QSpinBox()
            self.SbLower = QSpinBox()
            self.SbUpper.valueChanged.connect(self.set_crop_margins)
            self.SbLeft.valueChanged.connect(self.set_crop_margins)
            self.SbRight.valueChanged.connect(self.set_crop_margins)
            self.SbLower.valueChanged.connect(self.set_crop_margins)

            # Radiobutton group for median filter
            rdBtLayoutMedFi = QHBoxLayout()  # layout for the radio button widget
            self.rdBtwidgetMedFi = QWidget(self)  # radio button widget :: Place this in the main Layout
            self.rdBtwidgetMedFi.setLayout(rdBtLayoutMedFi)
            self.rdbtgroupMedFi = QButtonGroup(self.rdBtwidgetMedFi)
            #self.rdBtwidgetMedFi.setStyleSheet("background-color: 255; border: 1px solid black; border-radius: 20px;")
            rdbtMedFiNo = QRadioButton("No Median filter")
            self.rdbtgroupMedFi.addButton(rdbtMedFiNo, 0)  # 0 is the id in the group
            rdBtLayoutMedFi.addWidget(rdbtMedFiNo)
            rdbtMedFiNo.setChecked(True)  # check one button on creation
            rdbtMedFiPrev = QRadioButton("Preview (1 image)")
            self.rdbtgroupMedFi.addButton(rdbtMedFiPrev, 1)
            rdBtLayoutMedFi.addWidget(rdbtMedFiPrev)
            rdbtMedFiApply = QRadioButton("Apply (stack)")
            self.rdbtgroupMedFi.addButton(rdbtMedFiApply, 2)
            rdBtLayoutMedFi.addWidget(rdbtMedFiApply)
            self.rdbtgroupMedFi.buttonClicked[int].connect(self.medFi_state_changed)
            self.LbMedFi = QLabel("pixels")
            self.SbMedFi = QSpinBox()
            self.SbMedFi.setRange(1, 15)
            self.SbMedFi.setValue(3)
            self.SbMedFi.setSingleStep(1)  # even numbers create a displacement of the image, but no problem here
            self.SbMedFi.valueChanged.connect(self.medFi_sb_changed)

            ### Set fixed color scale
            self.CbUseVfixed = QCheckBox('Freeze color scale')
            self.CbUseVfixed.stateChanged.connect(self.view_changed)
            self.LbVmax = QLabel('Choose color map ->')
            self.CoboCmap = QComboBox()
            self.CoboCmap.addItems(["gist_ncar", "jet", "gray", "hot", "hsv", "inferno"])
            # common color maps:     jet,   gray,   hot,   hsv,    inferno,  gist_ncar
            self.CoboCmap.currentIndexChanged.connect(self.view_changed)
            self.LbVmin = QLabel('Set min, max:')
            self.SbVmin = QSpinBox()
            self.SbVmax = QSpinBox()
            self.SbVmin.valueChanged.connect(self.view_changed)
            self.SbVmax.valueChanged.connect(self.view_changed)

            # For ROI 1
            self.lbROI1 = QLabel("ROI 1 (background)")
            self.lbROI1ul = QLabel("tbd")
            self.lbROI1lr = QLabel("tbd")
            self.cbROI1redefine = QCheckBox('redefine')
            self.cbROI1redefine.clicked.connect(self.roi_def)

            # For choose range of good indices
            self.LbImIndic = QLabel("Choose range of good images :")
            self.LbImFrom = QLabel('Image from :')
            self.SbImFrom = QSpinBox()
            self.SbImFrom.valueChanged.connect(self.set_good_idx)
            self.LbImTo = QLabel('                to : ')
            self.LbImTo.setFixedSize(100, 30)
            self.SbImTo = QSpinBox()
            self.SbImTo.valueChanged.connect(self.set_good_idx)

            # dark images
            self.lbDarkIm = QLabel("Dark images ")

            # For mask definition
            self.lbMaskTit = QLabel("Modify mask: ")
            self.CoboMaskMode = QComboBox()
            self.CoboMaskMode.addItems(["Add pixels", "Remove pixels"])
            self.cbMaskActivateIt = QCheckBox('Do it now')
            self.cbMaskActivateIt.clicked.connect(self.roi_def)
            self.cbMaskShow = QCheckBox('Show mask')
            self.cbMaskShow.clicked.connect(self.display)
            self.lbMaskInfo = QLabel("0 pixels masked")

            self.SlNum.setMinimum(1)
            self.SlNum.setMaximum(64)

        def set_layout_of_tab1():
            print('set_layout_of_tab1 called')

            # setting the layout of the 1st tab
            grid1 = QGridLayout(self)
            grid1.setSpacing(1)  # defines the spacing between widgets
            # horizontally: end of label to edit, and vertically: edit to edit

            # file loading
            grid1.addWidget(self.BtLoadFile1, 0, 0)  # (object, row, column, rowSpan, colSpan)
            grid1.addWidget(self.LbFolderName, 0, 1, 1, 5)
            grid1.addWidget(self.LbImageSize, 0, 6, 1, 3)  # For image size information

            # For defining the pixel size (replace by combo box?)
            grid1.addWidget(self.LbPiSi, 2, 0)
            grid1.addWidget(self.EdPiSi, 2, 1)
            grid1.addWidget(self.LbPiSiUnit, 2, 2)

            # Navigate in stack
            grid1.addWidget(self.BtNumDwn, 3, 1)
            grid1.addWidget(self.SlNum, 3, 2, 1, 1)
            grid1.addWidget(self.BtNumUp, 3, 3)
            grid1.addWidget(self.LbImNo, 3, 0, 1, 1)

            # saturation
            grid1.addWidget(self.LbPixSat, 4, 3, 1, 2)

            # choose the view
            grid1.addWidget(self.rdbtwidgetView, 4, 1, 1, 2)
            grid1.addWidget(self.LbView, 4, 0)

            # Showing the present image
            grid1.addWidget(self.canvasIM1, 5, 0, 15, 5)
            grid1.addWidget(self.toolbarIM1, 21, 0, 1, 5)

            # Define crop margins (in pixels)
            grid1.addWidget(self.CbCrop, 2, 6)
            grid1.addWidget(self.SbUpper, 2, 7)
            grid1.addWidget(self.LbUpper, 2, 8)
            grid1.addWidget(self.SbLeft, 3, 6)
            grid1.addWidget(self.SbRight, 3, 8)
            grid1.addWidget(self.SbLower, 4, 7)

            ### Activate Median Filter
            grid1.addWidget(self.rdBtwidgetMedFi, 5, 6, 1, 4)
            grid1.addWidget(self.SbMedFi, 6, 7)
            grid1.addWidget(self.LbMedFi, 6, 8)

            ### Set fixed color scale
            grid1.addWidget(self.CbUseVfixed, 8, 6)
            grid1.addWidget(self.LbVmax, 8, 7)
            grid1.addWidget(self.CoboCmap, 8, 8)
            grid1.addWidget(self.LbVmin, 9, 6)
            grid1.addWidget(self.SbVmin, 9, 7)
            grid1.addWidget(self.SbVmax, 9, 8)

            # For ROI 1
            grid1.addWidget(self.lbROI1, 11, 6)
            grid1.addWidget(self.lbROI1ul, 11, 7)
            grid1.addWidget(self.lbROI1lr, 12, 7)
            grid1.addWidget(self.cbROI1redefine, 12, 6)

            # For choose range of good indices
            grid1.addWidget(self.LbImIndic, 14, 6, 1, 2)
            grid1.addWidget(self.LbImFrom, 15, 6)
            grid1.addWidget(self.SbImFrom, 15, 7)
            grid1.addWidget(self.LbImTo, 16, 6)
            grid1.addWidget(self.SbImTo, 16, 7)

            # For dark images
            grid1.addWidget(self.lbDarkIm, 18, 6, 1, 3)

            # For mask definition and display
            grid1.addWidget(self.lbMaskTit, 20, 6)
            grid1.addWidget(self.CoboMaskMode, 20, 7)
            grid1.addWidget(self.cbMaskActivateIt, 20, 8)
            grid1.addWidget(self.cbMaskShow, 21, 6)
            grid1.addWidget(self.lbMaskInfo, 21, 7)

            self.tab1.setLayout(grid1)
            # grid1.setRowStretch(2, 1)

        add_widgets_to_tab1()
        set_layout_of_tab1()

        def design_tab2():
            print("design_tab2 called")

            # main design is a VBox
            tab2_layout = QVBoxLayout(self)

            # The upper element is a grid with the labels and controls for background subtraction,
            # the beam center choice, feedback on the beam energy, and choice of image
            tab2_upper_grid2 = QGridLayout()
            # tab2_upper_grid2.setSpacing(1)

            # For background subtraction
            self.LbTitleSub = QLabel('Background Subtraction')
            self.LbTitleSub.setFont(section_h_font)
            tab2_upper_grid2.addWidget(self.LbTitleSub, 0, 0, 1, 4)

            self.cob_dark_im = QComboBox()
            self.cob_dark_im.addItems(["Don't use dark image", 'Use a dark image'])
            tab2_upper_grid2.addWidget(self.cob_dark_im, 1, 0, 1, 1)
            self.cob_dark_im.currentIndexChanged.connect(self.onImageChange)

            self.LbDark = QLabel('Number of dark Im. ')
            tab2_upper_grid2.addWidget(self.LbDark, 1, 1, 1, 1)

            self.SbDark = QSpinBox()
            tab2_upper_grid2.addWidget(self.SbDark, 1, 2, 1, 1)
            # self.SbDark.valueChanged.connect(self.onImageChange)
            # Makes a problem on load because of auto-fill before tab1 is finished

            self.cob_use_mean = QComboBox()
            self.cob_use_mean.addItems(["Use mean from ROI", "Use A<sub>eff</sub>-curve fit"])
            tab2_upper_grid2.addWidget(self.cob_use_mean, 2, 0, 1, 1)
            self.cob_use_mean.currentIndexChanged.connect(self.onImageChange)

            self.cob_roi = QComboBox()
            self.cob_roi.addItems(["Outside ROI", "Inside ROI"])
            tab2_upper_grid2.addWidget(self.cob_roi, 2, 1, 1, 1)
            self.cob_roi.currentIndexChanged.connect(self.on_inside_chg)

            self.cb_no_auto_crop = QCheckBox("No autocrop")
            tab2_upper_grid2.addWidget(self.cb_no_auto_crop, 2, 2, 1, 1)
            self.cb_no_auto_crop.stateChanged.connect(self.onImageChange)

            # Aeff-curve parameters
            self.LbTitleSub = QLabel('For A<sub>eff</sub>-curve and horizontality fit')
            tab2_upper_grid2.addWidget(self.LbTitleSub, 0, 4, 1, 2)

            self.LbStepDarkFrame = QLabel('Steps in dark frame:')
            tab2_upper_grid2.addWidget(self.LbStepDarkFrame, 1, 4, 1, 1)
            self.SbSteps = QSpinBox()
            self.SbSteps.setValue(10)
            tab2_upper_grid2.addWidget(self.SbSteps, 1, 5, 1, 1)
            self.SbSteps.valueChanged.connect(self.onImageChange)

            self.lbNbPoint = QLabel('Nb of line-fit points: ')
            tab2_upper_grid2.addWidget(self.lbNbPoint, 2, 4, 1, 1)
            self.sbNbPoint = QSpinBox()
            self.sbNbPoint.setValue(5)
            tab2_upper_grid2.addWidget(self.sbNbPoint, 2, 5, 1, 1)
            self.sbNbPoint.valueChanged.connect(self.onImageChange)

            # Beam maximum
            self.LbTitleMax = QLabel('Maximum smoothing')
            self.LbTitleMax.setFont(section_h_font)
            tab2_upper_grid2.addWidget(self.LbTitleMax, 0, 10, 1, 2)

            self.CbShowBMax = QCheckBox("Show maximum")
            tab2_upper_grid2.addWidget(self.CbShowBMax, 1, 10, 1, 1)
            self.CbShowBMax.stateChanged.connect(self.display2)

            self.cob_max = QComboBox()
            self.cob_max.addItems(['Max pixel', 'Max pixel (3x3 mean)', 'Max pixel (5x5 mean)',
                                   'Cap fit 95%', 'Cap fit 90%', 'Cap fit 80%', 'Cap fit 70%',
                                   'Centroid', 'Centroid (3x3 mean)', 'Centroid (5x5 mean)'])
            tab2_upper_grid2.addWidget(self.cob_max, 1, 11)
            self.cob_max.currentIndexChanged.connect(self.onImageChange)

            # info on found values (energy, backg, max)
            self.LbOffSet = QLabel('Found offset: ')
            tab2_upper_grid2.addWidget(self.LbOffSet, 1, 7, 1, 2)

            self.LbEng = QLabel('Found energy: ')
            tab2_upper_grid2.addWidget(self.LbEng, 2, 7, 1, 2)

            self.LbMax = QLabel('Found max. value: ')
            tab2_upper_grid2.addWidget(self.LbMax, 3, 7, 1, 2)

            self.LbMaxPos = QLabel('Found max. position: ')
            tab2_upper_grid2.addWidget(self.LbMaxPos, 3, 10, 1, 2)

            self.LbTi2 = QLabel('Used time: XX ms')
            tab2_upper_grid2.addWidget(self.LbTi2, 4, 10, 1, 2)

            # Navigate in good images
            self.LbImInfo = QLabel('Present image NA')  # change '...' when function is OK
            tab2_upper_grid2.addWidget(self.LbImInfo, 4, 0, 1, 1)

            self.BtNumDwn2 = QPushButton('<-')
            self.BtNumDwn2.clicked.connect(self.display_down2)
            tab2_upper_grid2.addWidget(self.BtNumDwn2, 4, 1)

            self.SlNum2 = QSlider(Qt.Horizontal, self)
            self.SlNum2.setMinimum(1)
            self.SlNum2.setMaximum(64)
            self.SlNum2.valueChanged.connect(self.onImageChange)
            tab2_upper_grid2.addWidget(self.SlNum2, 4, 2, 1, 6)

            self.BtNumUp2 = QPushButton(' -> ')
            self.BtNumUp2.clicked.connect(self.display_up2)
            tab2_upper_grid2.addWidget(self.BtNumUp2, 4, 8, 1, 1)

            # finish upper element
            tab2_upper_widget = QWidget()
            tab2_upper_widget.setLayout(tab2_upper_grid2)
            tab2_layout.addWidget(tab2_upper_widget, stretch=5)  # peut-être on aurait pu faire aussi .addLayout
            # au lieu de .addWidget qui a eu le Layout avec .setLayout

            # The lower element is vertical splitter with the left side used for the plots (in a HBox)
            # and the right side for the tables (in a VBox)
            tab2_lower_spli_widg = QSplitter(Qt.Horizontal)  # Horizontally moving

            # 2D-display
            left_widg = QWidget()
            # this will be a VBox layout
            left_wi_lay = QVBoxLayout()

            self.CoboCmap2 = QComboBox()
            self.CoboCmap2.addItems(["gist_ncar", "jet", "gray", "hot", "hsv", "inferno"])
            # common color maps:     jet,   gray,   hot,   hsv,    inferno,  gist_ncar
            self.CoboCmap2.setCurrentText("jet")
            self.CoboCmap2.currentIndexChanged.connect(self.display2)
            left_wi_lay.addWidget(self.CoboCmap2)

            self.Cob2Ddisp = QComboBox()
            self.Cob2Ddisp.addItems(["Measurement", "Round G. model", "Round G. residual (measurement - model)",
                                     "Elliptic G. model", "Elliptic G. residual (measurement - model)"])
            self.Cob2Ddisp.currentIndexChanged.connect(self.display2)
            left_wi_lay.addWidget(self.Cob2Ddisp)

            self.figureIM2 = Figure(figsize=(1.2, 1))  # by default make a small figure
            self.canvasIM2 = FigureCanvas(self.figureIM2)
            self.canvasIM2.setSizePolicy(QSizePolicy.Expanding,  # but stretch is if there is space
                                         QSizePolicy.Expanding)
            self.canvasIM2.updateGeometry()
            left_wi_lay.addWidget(self.canvasIM2)

            self.toolbarIM2 = NavigationToolbar(self.canvasIM2, self)
            left_wi_lay.addWidget(self.toolbarIM2)

            # left_wi_lay.addStretch(1)  # keeps the image vertically small
            left_widg.setLayout(left_wi_lay)  # add VBox layout to widget
            tab2_lower_spli_widg.addWidget(left_widg)  # add widget to splitter

            # 1D-display
            middle_widg = QWidget()
            # this will be a VBox layout
            middle_wi_lay = QVBoxLayout()
            self.Cob1Ddisp = QComboBox()
            self.Cob1Ddisp.addItems(["A<sub>eff</sub> curves", "Cross sections", "Surface over threshold (SOT)"])
            self.Cob1Ddisp.currentIndexChanged.connect(self.plot2_1D)
            middle_wi_lay.addWidget(self.Cob1Ddisp)

            self.figurePlt2 = Figure(figsize=(1.2, 1))  # by default make a small figure
            self.canvasPlt2 = FigureCanvas(self.figurePlt2)
            self.canvasPlt2.setSizePolicy(QSizePolicy.Expanding,  # but stretch is if there is space
                                          QSizePolicy.Expanding)
            self.canvasPlt2.updateGeometry()
            middle_wi_lay.addWidget(self.canvasPlt2)

            self.toolbarPlt2 = NavigationToolbar(self.canvasPlt2, self)
            middle_wi_lay.addWidget(self.toolbarPlt2)

            # middle_wi_lay.addStretch(1)  # keeps the image vertically small
            middle_widg.setLayout(middle_wi_lay)  # add VBox layout to widget
            tab2_lower_spli_widg.addWidget(middle_widg)  # add widget to splitter

            # Tables for fits and Aeff values
            right_widg = QWidget()
            # this will be a VBox layout
            right_wi_lay = QVBoxLayout()

            # Add a bit of space before the Aeff grid
            # right_wi_lay.addStretch(1)

            # Add the Aeff grid
            # Title
            aeff_grid_lay = QGridLayout()
            self.LbAeffTitle = QLabel('Effective beam surface (A<sub>eff</sub>):')
            self.LbAeffTitle.setFont(section_h_font)
            aeff_grid_lay.addWidget(self.LbAeffTitle, 0, 0, 1, 2)

            # use maximum from model beam fit below
            self.cbUseModMax = QCheckBox('Use maximum from model beam fit below.')
            self.cbUseModMax.stateChanged.connect(self.onImageChange)
            aeff_grid_lay.addWidget(self.cbUseModMax, 1, 0, 1, 2)

            # Aeff display
            self.LbAeffg = QLabel('A<sub>eff</sub> : XX px')
            aeff_grid_lay.addWidget(self.LbAeffg, 2, 0)
            # add uncertainty in string? see get_aeffg

            # The w of the corresponding Gaussian beam
            self.LbAeffg_to_gau_w = QLabel('-> w_1 of G.: XX px.')
            aeff_grid_lay.addWidget(self.LbAeffg_to_gau_w, 2, 1)

            # Quality check parameters
            self.LbSlope = QLabel('Slope : XX')
            aeff_grid_lay.addWidget(self.LbSlope, 3, 0)
            self.LbStraightness = QLabel('Straightness : XX')
            aeff_grid_lay.addWidget(self.LbStraightness, 3, 1)

            right_wi_lay.addLayout(aeff_grid_lay)

            # Add a bit of space
            # right_wi_lay.addStretch(1)

            # Add the Fit table
            self.LbFitTitle = QLabel("Analytical beam models: ")
            self.LbFitTitle.setFont(section_h_font)
            right_wi_lay.addWidget(self.LbFitTitle)

            self.TbFit2 = QTableWidget(11, 3)
            self.TbFit2.setHorizontalHeaderLabels(["Round G.", "Same SOT G.", "Ell. G."])
            self.TbFit2.setVerticalHeaderLabels(['Make fit?', "fit limit (rel.)",
                                                 'Use for A<sub>eff</sub> \n(if activated above)',
                                                 'status', 'Which GOF?', 'w_1 (px)', 'Max val (rel.)',
                                                 'x_0 (px)', 'y_0 (px)', 'w_2 (px)', 'Angle (°)'])

            def fill_fit_tb():
                self.rdbtgroup_for_aeff = QButtonGroup(self.rdBtwidgetMedFi)

                # col 0 - 2
                for col_idx in range(3):  # the names can be reused but new instances need to be created
                    cob_YesNo = QComboBox()
                    cob_YesNo.addItems(['Yes', 'No'])
                    cob_YesNo.setCurrentText("No")
                    self.TbFit2.setCellWidget(0, col_idx, cob_YesNo)

                    ed_f_lim = QLineEdit('0.5')
                    self.TbFit2.setCellWidget(1, col_idx, ed_f_lim)

                    rdbt_for_aeff = QRadioButton()
                    rdbt_for_aeff.setStyleSheet('QRadioButton::indicator { width: 100px;};')
                    # Didn't succeed to center the indicator (standard position is left top)
                    self.TbFit2.setCellWidget(2, col_idx, rdbt_for_aeff)
                    self.rdbtgroup_for_aeff.addButton(self.TbFit2.cellWidget(2, col_idx), col_idx)

                    lb_f_status = QLabel('NA')
                    self.TbFit2.setCellWidget(3, col_idx, lb_f_status)

                    lb_GOF = QLabel('NA')
                    self.TbFit2.setCellWidget(4, col_idx, lb_GOF)

                    ed_f_w1 = QLineEdit('30')
                    self.TbFit2.setCellWidget(5, col_idx, ed_f_w1)

                    ed_f_max = QLineEdit('1')
                    self.TbFit2.setCellWidget(6, col_idx, ed_f_max)

                # add widgets for round and ell. Gaussian
                for col_idx in [0, 2]:  # the names can be reused but new instances need to be created
                    ed_x_0 = QLineEdit('100')
                    self.TbFit2.setCellWidget(7, col_idx, ed_x_0)

                    ed_y_0 = QLineEdit('100')
                    self.TbFit2.setCellWidget(8, col_idx, ed_y_0)

                # add additional widgets for col 3
                ed_f_w2 = QLineEdit('30')
                self.TbFit2.setCellWidget(9, 2, ed_f_w2)

                ed_teta = QLineEdit('0')
                self.TbFit2.setCellWidget(10, 2, ed_teta)

                self.TbFit2.resizeRowsToContents()


            fill_fit_tb()
            self.TbFit2.resizeColumnsToContents()
            right_wi_lay.addWidget(self.TbFit2, stretch=3)

            # Things below the fit table
            low_line_widg = QWidget()
            # this will be a VBox layout
            low_line_lay = QHBoxLayout()

            lb_hist_tit = QLabel("Scroll parameter history (recent=0):")
            low_line_lay.addWidget(lb_hist_tit)

            self.sb_hist = QSpinBox()
            low_line_lay.addWidget(self.sb_hist)
            self.sb_hist.setRange(0, history_length - 1)
            self.SbMedFi.setValue(0)
            self.sb_hist.valueChanged.connect(self.scroll_para_history)

            self.bt_fit_upd = QPushButton("Update fits")
            low_line_lay.addWidget(self.bt_fit_upd)
            self.bt_fit_upd.clicked.connect(self.onImageChange)

            low_line_widg.setLayout(low_line_lay)  # add HBox layout to widget

            right_wi_lay.addWidget(low_line_widg)
            right_widg.setLayout(right_wi_lay)  # add VBox layout to widget
            tab2_lower_spli_widg.addWidget(right_widg)  # add widget to splitter

            # Set the layout of the splitter widget
            tab2_lower_spli_widg.setStretchFactor(0, 1)  # index, factor
            tab2_lower_spli_widg.setStretchFactor(1, 1)  # index, factor
            tab2_lower_spli_widg.setStretchFactor(2, 1)  # index, factor
            tab2_lower_spli_widg.setLayout(QHBoxLayout())  # so the splitter takes the full width
            # tab2_lower_spli_widg.setStyleSheet("border: 3px solid black;")

            # add the splitter to the VBox layout of tab2
            tab2_layout.addWidget(tab2_lower_spli_widg, stretch=20)

            # attach the whole thing (VBox) to tab2
            self.tab2.setLayout(tab2_layout)

        design_tab2()

        def design_tab3():
            print("design_tab3 called")
            # Try to avoid making widgets at each hierarchy level, add layouts to other layouts

            # layouts
            tab3_VB_lay = QVBoxLayout()  # the largest one

            tab3_upper_HB_lay = QHBoxLayout()  # The upper part

            tab3_ul_VB_lay = QVBoxLayout()  # the upper left part

            # title
            self.LbTitle4 = QLabel('Check stack statistics and export')  # label
            # self.LbTitle4.setStyleSheet(""" QLabel {background-color: rgb(230, 230, 255);} """)
            # font properties cannot be changed with the StyleSheet. Use this:
            # section_h_font = QFont("Times", 8, QFont.Bold)  # copied above for use elsewhere
            self.LbTitle4.setFont(section_h_font)
            tab3_ul_VB_lay.addWidget(self.LbTitle4)

            # time consumption warning
            self.LbTiWarn = QLabel(
                "This will take about {} s.".format(len(good_idx)*ti0))  # label
            tab3_ul_VB_lay.addWidget(self.LbTiWarn)

            # filter images to use
            self.LbFiTitle = QLabel("Do not use these images: ")  # label
            tab3_ul_VB_lay.addWidget(self.LbFiTitle)
            self.LbFiDark = QLabel("  Dark images: ")  # label
            tab3_ul_VB_lay.addWidget(self.LbFiDark)
            self.EdFiDark = QLineEdit()  # label
            tab3_ul_VB_lay.addWidget(self.EdFiDark)
            self.LbFiFilt = QLabel("  Filtered images: ")  # label
            tab3_ul_VB_lay.addWidget(self.LbFiFilt)
            self.EdFiFilt = QLineEdit()  # label
            tab3_ul_VB_lay.addWidget(self.EdFiFilt)

            # Button 'analyze'
            self.btPlot4 = QPushButton('Analyze stack now')
            self.btPlot4.clicked.connect(self.make_result_dict)
            tab3_ul_VB_lay.addWidget(self.btPlot4)

            tab3_ul_VB_lay.addStretch(1)

            # COBO title
            self.LbChosePlots = QLabel("Choose histograms to plot")  # label
            tab3_ul_VB_lay.addWidget(self.LbChosePlots)

            # choose variables to plot: Cobos in a grid
            # These are populated later depending on chosen fits
            tab3_cobo_grid_wid = QWidget()
            tab3_cobo_grid_lay = QGridLayout(tab3_cobo_grid_wid)

            self.CobPl31 = QComboBox()
            tab3_cobo_grid_lay.addWidget(self.CobPl31, 0, 0)
            self.CobPl31.activated.connect(self.plot3_1D)

            self.CobPl32 = QComboBox()
            tab3_cobo_grid_lay.addWidget(self.CobPl32, 0, 1)
            self.CobPl32.activated.connect(self.plot3_1D)

            self.CobPl33 = QComboBox()
            tab3_cobo_grid_lay.addWidget(self.CobPl33, 1, 0)
            self.CobPl33.activated.connect(self.plot3_1D)

            self.CobPl34 = QComboBox()
            tab3_cobo_grid_lay.addWidget(self.CobPl34, 1, 1)
            self.CobPl34.activated.connect(self.plot3_1D)

            tab3_ul_VB_lay.addWidget(tab3_cobo_grid_wid)

            tab3_ul_VB_lay.addStretch(1)

            tab3_upper_HB_lay.addLayout(tab3_ul_VB_lay, stretch=1)

            tab3_ur_VB_lay = QVBoxLayout()  # the upper right part
            # output (4 graphs)
            self.figurePlt3 = Figure()
            self.canvasPlt3 = FigureCanvas(self.figurePlt3)
            tab3_ur_VB_lay.addWidget(self.canvasPlt3)
            self.toolbarPlt3 = NavigationToolbar(self.canvasPlt3, self)
            tab3_ur_VB_lay.addWidget(self.toolbarPlt3)
            tab3_upper_HB_lay.addLayout(tab3_ur_VB_lay, stretch=3)
            # Connect a function that gets the mouse position on every click
            self.canvasPlt3.mpl_connect('button_release_event', self.GetMousePlot3)

            tab3_VB_lay.addLayout(tab3_upper_HB_lay, stretch=3)

            tab3_lower_HB_lay = QHBoxLayout()

            # preview and select export quantities
            self.tb_res3 = QTableWidget(7, 0)
            tab3_lower_HB_lay.addWidget(self.tb_res3)
            self.tb_res3.setVerticalHeaderLabels(
                ['Export?', 'Mean', 'Rel err.', 'Std err.', 'Min.', 'Max.', 'Median'])

            # Export
            self.btExport4 = QPushButton('Export')
            tab3_lower_HB_lay.addWidget(self.btExport4)
            self.btExport4.clicked.connect(self.export)

            tab3_VB_lay.addLayout(tab3_lower_HB_lay, stretch=1)

            self.tab3.setLayout(tab3_VB_lay)

        design_tab3()

        def design_tab4():
            print("design_tab4 called")
            # Try to avoid making widgets at each hierarchy level, add layouts to other layouts

            # layouts
            tab4_VB_lay = QVBoxLayout()  # the largest one

            tab4_upper_HB_lay = QHBoxLayout()

            tab4_ul_VB_lay = QVBoxLayout()

            #self.LbTitle1 = QLabel('Check stack statistics and export')  # label
            #self.LbTitle4.setFont(section_h_font)
            #tab4_ul_VB_lay.addWidget(self.LbTitle1)
            #elf.LbDiscp1 = QLabel("Do not use these images: ")  # label
            #tab4_ul_VB_lay.addWidget(self.LbDiscp1)
            self.LbChosePlots = QLabel("Choose histograms to plot")  # label
            tab4_ul_VB_lay.addWidget(self.LbChosePlots)

            tab4_cobo_grid_wid = QWidget()
            tab4_cobo_grid_lay = QGridLayout(tab4_cobo_grid_wid)

            self.CobPl1 = QComboBox()
            tab4_cobo_grid_lay.addWidget(self.CobPl1, 0, 0)
            self.CobPl1.activated.connect(self.set_checkb_state4)
            self.CobPl2 = QComboBox()
            tab4_cobo_grid_lay.addWidget(self.CobPl2, 0, 1)
            self.CobPl2.activated.connect(self.set_checkb_state4)
            self.Cb4Pl1 = QCheckBox("Use eff. radius?")
            # self.Cb4Pl1.setEnabled(False)
            tab4_cobo_grid_lay.addWidget(self.Cb4Pl1, 1, 0)
            self.Cb4Pl2 = QCheckBox("Use eff. radius?")
            # self.Cb4Pl2.setEnabled(False)
            tab4_cobo_grid_lay.addWidget(self.Cb4Pl2, 1, 1)

            tab4_ul_VB_lay.addWidget(tab4_cobo_grid_wid)
            tab4_ul_VB_lay.addStretch(1)
            bt4_launch_z = QPushButton("Launch processing (attention slow)")
            tab4_ul_VB_lay.addWidget(bt4_launch_z)
            bt4_launch_z.clicked.connect(self.make_z_curves)
            tab4_ul_VB_lay.addStretch(1)

            tab4_upper_HB_lay.addLayout(tab4_ul_VB_lay, stretch=1)

            tab4_ur_VB_lay = QVBoxLayout()  # the upper right part

            self.figurePlt1 = Figure()
            self.canvasPlt1 = FigureCanvas(self.figurePlt1)
            tab4_ur_VB_lay.addWidget(self.canvasPlt1)
            self.toolbarPlt1 = NavigationToolbar(self.canvasPlt1, self)
            tab4_ur_VB_lay.addWidget(self.toolbarPlt1)
            tab4_upper_HB_lay.addLayout(tab4_ur_VB_lay, stretch=3)

            self.canvasPlt1.mpl_connect('button_release_event', self.GetMousePlot3)

            tab4_VB_lay.addLayout(tab4_upper_HB_lay, stretch=3)

            self.tab4.setLayout(tab4_VB_lay)

        design_tab4()

        def design_tab5():
            print("design_tab5 called")
            # Try to avoid making widgets at each hierarchy level, add layouts to other layouts

            # layouts
            scroll_VB_lay = QVBoxLayout()  # the largest one

            # title general usage
            self.LbDescTi0 = QLabel('Aim of this program')  # label
            self.LbDescTi0.setFont(section_h_font)
            scroll_VB_lay.addWidget(self.LbDescTi0)
            # First paragraph for usage of tab1
            self.Lb0desc = QLabel(
                "This program aims to describe (peaked) spatial beam profiles of laser beams. " +
                "The first 3 tabs are used to analyze a stack of images that were all taken " +
                "at the same position with the same laser and setup settings. " +
                "(If all instruments were perfect, all images in a stack would be identical.) " +
                "In order to get the characteristic numbers of the beam profile and their " +
                "uncertainties we go through 3 steps, one per tab. \n" +
                "TAB 1:  Load the images of the same stack; crop them to a size that leaves " +
                "enough surface around the beam to determine the background level; select the " +
                "region of interest (ROI) that separates the beam zone from the background zone; " +
                "possibly smooth out isolated noise pixels " +
                "by median-filtering the images; possibly select the part of the stack that " +
                "contains valid images (exclude zero-images (compulsory) and background images " +
                "(optional)) if they are contained in a contiguous region of the stack. \n" +
                "TAB 2:  Set the way how the good images of the stack should be analyzed. " +
                "Several options for determination of the background grey-level (GL) and the " +
                "maximum GL are available. You can also fit 3 versions of Gaussian beams to " +
                "your beam profiles. \n" +
                "TAB 3:  Launch the stack analysis; inspect the histograms; possibly discard " +
                "some images based on the histograms (for example images with strong background); " +
                "export a MS-Excel xlsx-file with the data. "
            )
            self.Lb0desc.setWordWrap(True)
            scroll_VB_lay.addWidget(self.Lb0desc)

            # title usage of tab1
            self.LbDescTi1 = QLabel('Tab1: Loading images')  # label
            # self.LbTitle4.setStyleSheet(""" QLabel {background-color: rgb(230, 230, 255);} """)
            # font properties cannot be changed with the StyleSheet. Use this:
            # section_h_font = QFont("Times", 8, QFont.Bold)  # copied above for use elsewhere
            self.LbDescTi1.setFont(section_h_font)
            scroll_VB_lay.addWidget(self.LbDescTi1)
            # First paragraph for usage of tab1
            self.Lb1desc = QLabel(
                "The upper left button 'Load file' opens a file dialog. The program can read a " +
                "stack of images provided in two ways: \n" +
                "1. The contents of a 64-image wcf-file which was written by the DataRay " +
                "software of the WinCam beam profiling cameras. \n" +
                "2. All image files (*.tif, *.jpg, *.png) that have the same size (pixels) " +
                "and that are stored in the same directory. \n" +
                "After loading, information on the images is displayed in the first line " +
                "(path/filename; image size and format). \n" +
                "The slider and the arrow buttons allow to look at all the images one by one" +
                "(frame-by-frame mode). The color map and its limits can be chosen on the right " +
                "side. A tiled overview-mode of the stack is also available. (Radio buttons " +
                "below the slider.) \n\n" +
                "Up to now the Pixel size input under the 'Load file' button is not used. " +
                "All lateral output is in pixels (px). " +
                "All vertical output is in grey levels (GL). \n\n" +
                "First the full images are displayed with the crop frame in a dashed white " +
                "line. The crop frame can be modified using the 4 spinboxes (smaller increments " +
                "than the default are possible by writing in the boxes). Care should be taken " +
                "to keep enough space around or besides the beam, where the background is visible. " +
                "Best keep a dark frame with a width of a beam radius around outer edge of the " +
                "beam. For better visibility of the outer beam edge you can switch the color map if " +
                "necessary. Finally activate the cropping by using the checkbox ('apply crop margins') " +
                "near the cropping spinboxes. (The crop margins are maintained when loading a new image" +
                "stack.) \n"
                "After cropping the ROI has to be selected. The rectangular ROI will delimit the region " +
                "where the background signal is extracted from. Best use it to delimit a black frame " +
                "around the beam (with some safety margin). To select the ROI, check the checkbox " +
                "'redefine' below the label 'ROI I' and click on the image. The display of a red " +
                "dashed cross will help you to chose the first edge of the rectangle. So place the " +
                "mouse at the right position and push the key <h> on your keyboard to confirm. Then " +
                "choose the second edge and confirm with <n>. \n" +
                "The last compulsory action on this tab is to check that the range of good images does " +
                "not contain any images that are completely at zero. (These would crash the calculations " +
                "on tab 2.) Best, define the range of good images such that only valid images of the same " +
                "type are contained. In this version of the program only one region of the stack can be " +
                "excluded in its center (from > to) or first and last images can be excluded (from < to). " +
                "More flexibility may be added later. \n" +
                "Before going to the second tab, check that no saturated images are detected by reading " +
                "the label below the right arrow button above the image. The relative saturation limit is " +
                "0.95 by default. A possibility to deal with isolated saturated pixels (broken pixels or " +
                "noise) is to use median filtering of the images. If this does not help, it is useless to " +
                "go further. Acquire new images."
            )  # label
            self.Lb1desc.setWordWrap(True)
            scroll_VB_lay.addWidget(self.Lb1desc)

            # title usage of tab1
            self.LbDescTi2 = QLabel('Tab2: Define the analyses to run')  # label
            # self.LbTitle4.setStyleSheet(""" QLabel {background-color: rgb(230, 230, 255);} """)
            # font properties cannot be changed with the StyleSheet. Use this:
            # section_h_font = QFont("Times", 8, QFont.Bold)  # copied above for use elsewhere
            self.LbDescTi2.setFont(section_h_font)
            scroll_VB_lay.addWidget(self.LbDescTi2)
            # First paragraph for usage of tab1
            self.Lb2desc = QLabel(
                "Once the pretreatment is finished and the zero images are excluded, the images can be " +
                "analyzed using the controls in tab2. Some controls do not trigger an update of the tab. " +
                "To force an update, change image by using the slider or the arrow buttons. \n" +
                "First the background needs to be determined and subtracted. If the background is not " +
                "homogeneous and you have a background image use the controls in the first line: " +
                "Switch to 'Use a dark image' and give the number of the dark image (look it up in tab 1). \n" +
                "In the second line you choose the way how the homogeneous background correction is " +
                "calculated. The simplest and fastest way is to use the average value inside or outside the " +
                "ROI defined in tab 1. (If the red frame is around the beam, use 'outside', if the red frame " +
                "is besides the beam use 'inside'.) The 'autocrop' which is used to generate a black frame " +
                "of same width on all sides can be disabled using the checkbox. Disabling autocrop is " +
                "necessary if the ROI is besides the beam. \n" +
                "The second method 'Use A<sub>eff</sub>-curve fit' needs some more explanation: The effective surface" +
                "area, A<sub>eff</sub>, of a laser beam is defined as the proportionality factor between pulse energy " +
                "(Ep) and peak fluence (F): F = Ep/A<sub>eff</sub> (Or power (P) and peak irradiance (I): I = P/A<sub>eff</sub>) " +
                "From a beam image without background and without noise A<sub>eff</sub> can be obtained by A<sub>eff</sub> = Ep/F " +
                "= sum(pixels) / max(pixels). If the image is background corrected, adding more pixels of the" +
                "background in the calculation does not change the value of A<sub>eff</sub>. Thus the A<sub>eff</sub>-curve: " +
                "A<sub>eff</sub>(image_size) is flat. This fact can be used to retrieve the value of the homogeneous " +
                "background correction: (i) take a guess for the background, (ii) subtract it from the image, " +
                "(iii) make the A<sub>eff</sub>-curve by cropping the borders of the image and (iv) evaluate the slope " +
                "of the line through the last points of the A<sub>eff</sub>-curve. If the 'A<sub>eff</sub>-curve fit' is used to " +
                "determine the background level, the inside/outside control is not taken into account. \n" +
                "In any case, it's a good idea to take a look at the A<sub>eff</sub>-curve (shown by default on the right " +
                "graph). If the last points of the A<sub>eff</sub>-curve " +
                "are not on a straight line, this is a sign of either inhomogeneous background or too close " +
                "cropping (the edge of the image contains energy from the pulses low foothills). \n" +
                "After defining the background definition method, the method for the definition of the maximum " +
                "has to be chosen from the upper right combobox. The choice should consider the number of pixels " +
                "that are close to the maximum of the beam. If there are many pixels close to the maximum, the " +
                "pixel with the highest gray level (Max Pixel) is likely to be a noise pixel and some smoothing " +
                "method should be used ('...mean...' or 'cap fit ...') to avoid over-estimation of the the " +
                "maximal GL. If the pixels are large with respect to the" +
                "beam size, smoothing needs to be avoided to avoid under-estimation of the maximal GL."
            )  # label
            self.Lb2desc.setWordWrap(True)
            scroll_VB_lay.addWidget(self.Lb2desc)
            wrapper_widg = QWidget()
            wrapper_widg.setLayout(scroll_VB_lay)

            self.scroll_widg = QScrollArea()
            self.scroll_widg.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.scroll_widg.setWidgetResizable(True)
            self.scroll_widg.setWidget(wrapper_widg)

            tab5_VB_lay = QVBoxLayout()  # the largest one
            tab5_VB_lay.addWidget(self.scroll_widg)
            self.tab5.setLayout(tab5_VB_lay)


        design_tab5()

        # Add tabs-widget to centralwidget
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.tabs)

        self.firstLoad = True

    def range_str_conv(self, inp):
        print("range_str_conv called")

        if type(inp) is str:
            # print("string input - list output")
            outp = []
            try:
                inp = inp.split(",")
                for idx in range(len(inp)):
                    lims = inp[idx].split("-")
                    if len(lims) == 1:
                        outp = outp + [int(lims[0])]
                    else:
                        outp = outp + list(range(int(lims[0]), int(lims[1]) + 1))
            except ValueError:
                outp = []
        else:
            # print("list input - string output")
            if len(inp) == 0:
                return ""
            inp = np.array(inp).astype(int)
            outp = ""
            first = inp[0]
            last = first
            for idx in range(1, len(inp)):
                if inp[idx] == last + 1:
                    last += 1
                else:
                    if last > first:
                        outp = outp + str(first) + "-" + str(last) + ", "
                    else:
                        outp = outp + str(first) + ", "
                    first = inp[idx]
                    last = first
            if last > first:
                outp = outp + str(first) + "-" + str(last)
            else:
                outp = outp + str(first)
        return outp

    def display_down(self):
        print('display_down called')

        if self.SlNum.value() > 1:
            self.SlNum.setValue(self.SlNum.value() - 1)
            self.rdbtFbF.setChecked(True)  # to refresh the display

    def display_up(self):
        print('display_up called')

        if self.SlNum.value() < imStack.shape[2]:
            self.SlNum.setValue(self.SlNum.value() + 1)
            self.rdbtFbF.setChecked(True)  # to refresh the display

    def Load_File1(self):
        """Is called by
        self.BtLoadFile1.clicked.connect

        Get the filename and split it
        """
        print('Load_File1 called')
        global dirname, baseName, f_name, file_type

        # choose an image of the stack using a dialog
        f_name = QFileDialog.getOpenFileName(
            self,
            "Open file",
            "",  # default folder of wcf files. "" for local folder
            "All files (*.*);; Dataray files (*.wcf);; image files (*.tif *.tiff *.png *.jpg) ")

        if not f_name[0]:
            return  # function continues if a file was chosen
        print('File name: ', f_name[0])
        f_name = f_name[0]

        # display the filename in the label
        self.LbFolderName.setLongText(f_name)
        self.LbFolderName.setMinimumSize(1, 1)  # enables to reduce the window even if the path is long
        # publish the filename for use in other functions and tabs
        baseName = f_name.split('/')[-1]  # just the file name not the path
        dirname = '/'.join(f_name.split('/')[:-1])  # just the path
        file_type = baseName.split('.')[-1]
        baseName = '.'.join(baseName.split('.')[:-1])  # just the filename that may contain dots
        print('Basename: ' + baseName)
        self.load_stack()  # so no button click is necessary


    def load_stack(self):
        """ Called by:
        Load_File, crop_or_uncrop

        Load all images to memory
        :return:
        """
        print('load_stack called')
        global imStack
        global roiS
        global masked, mask_during_modif
        global im1_width, im1_height  # size of the original images
        global sat_value  # the maximum integer value in the images

        main.statBar.showMessage('Reading images... please wait')
        file_res_info = "Image size (horiz, vert) in pixels: "

        if file_type.upper() == "WCF":  # read wcf file that contains 64 images (taken in identic cond.)
            # get image resolution from filesize
            fi_size = os.path.getsize(f_name)
            print("fi_size = ", fi_size, " Bytes")
            pix_num_est = int(np.round((fi_size - 934 - 5592) / 64 / 2 / 1000))  # kpixels
            #  934 was found by trial and error in comparison to Dataray software
            #  5592 can be read in the first header
            #  64 images
            #  2 bytes per pixel
            print("pix_num_est = ", pix_num_est, " kpixels")
            std_image_reslS = {  # key in in kpixels, tuple in pixels
                int(np.round(64 * 64 / 1000)): (64, 64),
                int(np.round(128 * 128 / 1000)): (128, 128),
                int(np.round(256 * 256 / 1000)): (256, 256),
                int(np.round(348 * 348 / 1000)): (348, 348),
                int(np.round(512 * 512 / 1000)): (512, 512),
                int(np.round(752 * 752 / 1000)): (752, 752),
                int(np.round(1024 * 1024 / 1000)): (1024, 1024),
                int(np.round(1200 * 1024 / 1000)): (1200, 1024)
            }
            try:
                im1_height = std_image_reslS[pix_num_est][1]
                im1_width = std_image_reslS[pix_num_est][0]
                file_res_info = file_res_info + "{:d}, {:d}; uint16".format(im1_width, im1_height)
                self.LbImageSize.setText(file_res_info)
                sat_value = 2**16 - 1

            except KeyError:
                im1_width = np.nan
                im1_height = np.nan
                file_res_info = file_res_info + "NAN, NAN"
                self.LbImageSize.setText(file_res_info)
                sat_value = 10
                return   # exit this function, if image size estimation not successful

            # start loading
            im_size_in_bytes = im1_width * im1_height * 2  # 2 bytes per pixel (uint16)
            main.progBar.setMinimum(0)
            main.progBar.setMaximum(64)
            cm = 1  # pixel (crop margin, because the nominal image size contains a saturated pixel)
            imStack = np.ndarray((im1_height - 2 * cm, im1_width - 2 * cm, 64))  # reserve memory for stack
            with open(f_name, mode="rb") as f:
                f.seek(934 + 5592)
                #  934 was found by trial and error in comparison to Dataray software
                #  5592 can be read in the first header
                im_data = f.read(im_size_in_bytes)
                main.progBar.setValue(1)
                QApplication.processEvents()
                for im_num in range(64):  # 64 images
                    # extract the image from the array of bytes and store it (except the outer frame)
                    im_data = struct.unpack("H" * ((len(im_data)) // 2), im_data)  # "H" is uint16
                    imStack[:, :, im_num] = np.reshape(im_data, (im1_height, im1_width))[cm:-cm, cm:-cm]
                    #imStack[:, :, im_num] = np.reshape(im_data, (im1_height, im1_width))

                    main.progBar.setValue(im_num + 1)
                    QApplication.processEvents()

                    # read the next array of bytes
                    im_data = f.read(im_size_in_bytes)
            # f.close() is not needed after with block
            im1_height = im1_height - 2 * cm  # for the global variable to reflect the size on imstack
            im1_width = im1_width - 2 * cm
            main.progBar.setValue(0)

        else:  # Is this a folder of valid files?  globals:  dirname, baseName, f_name
            imStack = []  # start as list, will be transformed to
            try:  # check if the selected image works
                im = plt.imread(f_name)  # read selected file first ATTENTION MAY BE skimage reader is better
                # there are cases where the plt reader does not keep the right data format
                imSi = im.shape
                imTy = str(im.dtype)
                sat_value = np.iinfo(im.dtype).max  # give feed back on image dtype
                imStack.append(im)
            except:  # the selected file is not a readable image
                main.statBar.showMessage('The selected file is not a readable image', 2000)
                self.LbFolderName.setText('The selected file is not a readable image or wcf-file')
                return
            # get the list of filenames
            filenames = next(os.walk(dirname))[2]
            main.progBar.setMinimum(0)
            main.progBar.setMaximum(len(filenames))
            for im_num, this_file in enumerate(filenames):
                this_file = '/'.join([dirname, this_file])

                main.progBar.setValue(im_num + 1)
                QApplication.processEvents()

                if this_file == f_name:
                    continue
                try:
                    im = plt.imread(this_file)
                    if (imSi == im.shape) and (str(im.dtype) == imTy):
                        imStack.append(im)
                except:
                    pass
            imStack = np.dstack(imStack)  # convert to numpy array
            imStack = imStack.astype(np.float64)  # convert to float type
            im1_height = imSi[0]  # for the global variable to reflect the size on imstack
            im1_width = imSi[1]
            file_res_info = file_res_info + "{:d}, {:d}; ".format(im1_width, im1_height) + imTy
            self.LbImageSize.setText(file_res_info)

        self.SlNum.setMaximum(imStack.shape[2])
        main.statBar.showMessage('Finished loading images', 2000)  # the message shows 2000 ms and disappears
        main.progBar.setValue(0)

        # for setting the maximum value of the vertical scale spin boxes
        vMaxMax = 0  # for setting the maximum value of the vertical scale spin boxes
        for im_num in range(imStack.shape[2]):
            vMaxNow = imStack[:, :, im_num].max()
            if vMaxNow > vMaxMax:
                vMaxMax = vMaxNow

        # initialize the spinboxes for cropping
        self.SbUpper.setRange(0, im1_height - 1)
        self.SbLower.setRange(0, im1_height - 1)
        self.SbLeft.setRange(0, im1_width - 1)
        self.SbRight.setRange(0, im1_width - 1)
        self.SbUpper.setSingleStep(im1_height//100)
        self.SbLeft.setSingleStep(im1_width//100)
        self.SbRight.setSingleStep(im1_width//100)
        self.SbLower.setSingleStep(im1_height//100)

        # Execute this only one time, if not, a user setting will be lost on reloading the stack.
        if self.firstLoad:
            # set initial crop values
            self.SbUpper.setValue(im1_height//5)
            self.SbLeft.setValue(im1_width//5)
            self.SbRight.setValue(im1_width//5)
            self.SbLower.setValue(im1_height//5)

            # setting the maximum value of the vertical scale spin boxes
            self.SbVmin.setRange(0, vMaxMax)
            self.SbVmax.setRange(0, vMaxMax)
            self.SbVmin.setSingleStep(vMaxMax // 100)
            self.SbVmax.setSingleStep(vMaxMax // 100)
            self.SbVmax.setValue(vMaxMax)
            self.firstLoad = False

            # Spinbox initialisation for image choice
            self.SbImFrom.setValue(1)
            self.SbImTo.setValue(imStack.shape[2])

            roiS = np.ones((3, 4))*np.nan  # xfrom, xto, yfrom, yto (include all indices)
            # has to be float if not, there is no "nan"
            masked = []
            mask_during_modif = False

        if self.CbCrop.isChecked():  # this case is for using more than one wcf file without closing the program
            # keep the cropping, if it was on
            imStack = imStack[self.SbUpper.value():(im1_height - 1 - self.SbLower.value()),
                              self.SbLeft.value():(im1_width - 1 - self.SbRight.value()), :]
        self.get_sat_pixels()
        self.get_dark_images()
        self.showRoiS()
        self.display()


    def get_sat_pixels(self):
        print("get_sat_pixels called")

        """
        Get saturated pixels and display result
        uses imStack
        :return: self.LbPixSat.setText
        """
        global imStack
        abs_sat_limit = sat_value * rel_sat_limit

        sat_pix_count = 0
        sat_im_count = 0
        for im_nb in range(imStack.shape[2]):
            new_sat_pix_count = np.count_nonzero(imStack[:, :, im_nb] > abs_sat_limit)
            sat_pix_count += new_sat_pix_count
            if new_sat_pix_count > 0:
                sat_im_count += 1

        if sat_pix_count > 0:
            self.LbPixSat.setText("ATTENTION {:d} saturated pixels found in {:d} images".
                                  format(sat_pix_count, sat_im_count))
        else:
            self.LbPixSat.setText("All images are ok (no saturation found)")

    def get_dark_images(self):
        print("get_dark_images called")

        global imStack
        # global dark_im_list

        # find dark limit in imStack
        im_max_arr = np.ndarray((imStack.shape[2]))
        for im_idx in range(imStack.shape[2]):
            im_max_arr[im_idx] = imStack[:, :, im_idx].max()

        dark_limit = np.median(im_max_arr) - 3 * (im_max_arr.max() - np.median(im_max_arr))

        # find dark images number
        dark_im_list = []
        mess = ""
        for im_idx in range(imStack.shape[2]):
            if imStack[:, :, im_idx].max() < dark_limit:
                dark_im_list.append(im_idx + 1)
                mess += "{:d}, ".format(im_idx + 1)
        if len(mess) == 0:
            mess = "No dark image found"
            self.SbDark.setValue(1)
            self.SbImFrom.setValue(1)
            self.SbImTo.setValue(imStack.shape[2])
        else:
            mess = "Number of possible dark images : " + mess[:-2]
            self.SbDark.setValue(dark_im_list[0])
            self.SbImFrom.setValue(dark_im_list[0] + 1)
            self.SbImTo.setValue(dark_im_list[0] - 1)
        self.lbDarkIm.setText(mess)
        print(mess)

        self.EdFiDark.setText(self.range_str_conv(dark_im_list))

    def set_good_idx(self):
        print("set_good_idx called")

        global good_idx
        im_from = self.SbImFrom.value()
        im_to = self.SbImTo.value()
        if im_from <= im_to:
            good_idx = list(range(im_from - 1, im_to))
        elif im_from > im_to:
            good_idx = list(range(im_to)) + list(range(im_from - 1, imStack.shape[2]))
        # print("good_id : ", good_idx)

    def view_changed(self):
        print('view_changed called')
        global axIM1
        global imStack
        ViewBtIdx = self.rdbtGroupView.checkedId()
        if ViewBtIdx == 0:  # display normal frame by frame view
            self.display()
        elif ViewBtIdx == 1:  # Display grid of all (downscaled) images
            ti1 = ti.time()
            st_copy = imStack.copy()
            max_im_size = 2048  # 4096 -> 1.1 sec # pixels in both directions

            # Downscaling loop (if necessary)
            x_num = max_im_size // st_copy.shape[1]  # how many images would fit into max_im_size
            y_num = max_im_size // st_copy.shape[0]
            while x_num * y_num < st_copy.shape[2]:
                # it does not fit -> downscale by a factor of 2
                st_copy = SkiMeasBR(st_copy, block_size=(2, 2, 1), func=np.max)  # block_reduce
                x_num = max_im_size // st_copy.shape[1]  # how many images would fit into max_im_size
                y_num = max_im_size // st_copy.shape[0]

            # make a close-to-square tile image
            im1_height = st_copy.shape[0]
            im1_width = st_copy.shape[1]
            x_num = int(np.sqrt(st_copy.shape[2] * im1_height / im1_width))
            y_num = int(np.ceil(st_copy.shape[2] / x_num))
            im_to_show = -st_copy.max()/10 * np.ones((im1_height*y_num, im1_width*x_num))  # adaptive nonsense
            for row in range(y_num):
                for col in range(x_num):
                    if col + x_num * row < st_copy.shape[2]:
                        im_to_show[row*im1_height:(row+1)*im1_height,
                                   col*im1_width:(col+1)*im1_width] = st_copy[:, :, col + x_num * row]
            # Clear the figure from earlier uses
            self.figureIM1.clear()
            # prepare the axis
            axIM1 = self.figureIM1.add_subplot(111)
            axIM1.axis("off")
            if self.CbUseVfixed.isChecked():
                vmi1 = self.SbVmin.value()
                vma1 = self.SbVmax.value()
            else:
                vmi1 = im_to_show.min()
                vma1 = im_to_show.max()

            cax = axIM1.imshow(im_to_show, cmap=self.CoboCmap.currentText(), vmin=vmi1, vmax=vma1)
            # common color maps: jet, gray, hot, hsv, inferno, gist_ncar
            self.figureIM1.colorbar(cax, orientation='vertical')
            self.figureIM1.tight_layout()

            # refresh canvas
            self.canvasIM1.draw()
            print("showing overview took {:f} seconds".format(ti.time()-ti1))
        # print('view_changed finished')


    def showRoiS(self):
        print("showRoiS called")

        # roi1
        self.lbROI1ul.setText("x: {:.0f}; y: {:.0f}".format(roiS[0, 0] - crop_ulx, roiS[0, 2] - crop_ulx))
        self.lbROI1lr.setText("x: {:.0f}; y: {:.0f}".format(roiS[0, 1] - crop_uly, roiS[0, 3] - crop_uly))


    def set_crop_margins(self):
        """Is called by
        self.SbLeft.valueChanged.connect(self.set_crop_margins)
        and the other crop margin spin boxes
        """
        print('set_crop_margins called')

        # This makes sure that inversions are impossible upper/lower left/right
        if self.sender() is self.SbUpper:
            self.SbLower.setRange(0, im1_height - 1 - self.sender().value())
        elif self.sender() is self.SbLower:
            self.SbUpper.setRange(0, im1_height - 1 - self.sender().value())
        elif self.sender() is self.SbLeft:
            self.SbRight.setRange(0, im1_width - 1 - self.sender().value())
        elif self.sender() is self.SbRight:
            self.SbLeft.setRange(0, im1_width - 1 - self.sender().value())
        self.display()

    def display(self):
        """Is called by
        self.SlNum.valueChanged.connect(self.display)
        and many others to display the present image of the stack
        """
        print('display called')
        try:
            if self.rdbtgroupMedFi.checkedId() == 1:  # Median filter preview is active
                self.display_MedFi_preview()
            else:  # just show the stack
                number = self.SlNum.value()  # read present slider value
                self.LbImNo.setText("Image " + "\n" + "    " + str(number))
                ImDisp1 = imStack[:, :, number - 1]  # make a view of the right image
                self.showIm1(ImDisp1)  # show the image
                self.rdbtFbF.setChecked(True) # for refresh display
        except NameError:  # necessary because the method is executed on initialisation
            return

    def showIm1(self, im_to_show):
        print('showIm1 called')
        global axIM1

        # TODO: Accelerate this method
        # Clear the figure from earlier uses
        self.figureIM1.clear()
        # prepare the axis
        axIM1 = self.figureIM1.add_subplot(111)
        axIM1.axis("off")
        if self.CbUseVfixed.isChecked():
            vmi1 = self.SbVmin.value()
            vma1 = self.SbVmax.value()
        else:
            vmi1 = im_to_show.min()
            vma1 = im_to_show.max()

        # mark masked pixels if wished
        im_modified = im_to_show.copy()  # modify a copy of im_to_show
        if self.cbMaskShow.isChecked():
            # calculate the values of the color switch limits and the
            # replacement values (for the masked pixels to become visible)
            MiniVal = np.min(im_to_show)
            MaxiVal = np.max(im_to_show)
            infernoLimits = [0.35, 0.82]
            infernoColVals = [0.66, 1]
            # change the values (colors) of the pixels in the list
            for count in range(len(masked)):
                try:
                    if (masked[count][1] - crop_uly >= 0) and (masked[count][0] - crop_ulx >= 0):  # no neg indices
                        IsVal = im_to_show[masked[count][1] - crop_uly, masked[count][0] - crop_ulx]
                        if infernoLimits[0] < IsVal / (MaxiVal - MiniVal) < infernoLimits[1]:  # light color
                            putVal = infernoColVals[1] * (MaxiVal - MiniVal) + MiniVal
                        else:
                            putVal = infernoColVals[0] * (MaxiVal - MiniVal) + MiniVal
                        im_modified[masked[count][1] - crop_uly, masked[count][0] - crop_ulx] = putVal
                except IndexError:
                    pass  # does not crash if pixels outside the present crop are part of the mask

        cax = axIM1.imshow(im_modified, cmap=self.CoboCmap.currentText(), vmin=vmi1, vmax=vma1)
        # common color maps: jet, gray, hot, hsv, inferno, gist_ncar
        self.figureIM1.colorbar(cax, orientation='vertical')
        self.figureIM1.tight_layout()

        self.UpdateLinesIm1()

        # refresh canvas
        self.canvasIM1.draw()


    def UpdateLinesIm1(self, xM =-10, yM = -10):
        global axIM1
        global roiS
        print("UpdateLinesIm1 called")

        # delete lines on axis (if there were any)
        try:
            axIM1.lines = []
        except NameError:  # just a guess (that seems to work)
            pass

        # Draw white lines of crop limits if crop inactive
        if not self.CbCrop.isChecked():
            linew = 1
            col = 'w'
            ulx = self.SbLeft.value()
            lrx = im1_width - 1 - self.SbRight.value()
            uly = self.SbUpper.value()
            lry = im1_height - 1 - self.SbLower.value()
            axIM1.plot([ulx, lrx, lrx, ulx, ulx], [uly, uly, lry, lry, uly],
                       color=col, linestyle='--', linewidth=linew)

        maxX = imStack[:, :, 0].shape[1]-1
        maxY = imStack[:, :, 0].shape[0]-1

        if xM == -10 and yM == -10:  # was not called by mouse event
            if (not(np.any(np.isnan(roiS[0, :])))
                and((0 <= roiS[0, 0] - crop_ulx <= maxX) or (0 <= roiS[0, 1] - crop_ulx <= maxX)
                    or (0 <= roiS[0, 2] - crop_uly <= maxY) or (0 <= roiS[0, 3] - crop_uly <= maxY))):
                col = 'r'
                ulx = max([roiS[0, 0] - crop_ulx, -1])
                lrx = min([roiS[0, 1] - crop_ulx, maxX + 1])
                uly = max([roiS[0, 2] - crop_uly, -1])
                lry = min([roiS[0, 3] - crop_uly, maxY + 1])
                axIM1.plot([ulx, lrx, lrx, ulx, ulx], [uly, uly, lry, lry, uly],
                           color=col, linestyle='--', linewidth=1)
            return  # do not draw the lines depending on the mouse position

        # was called by mouse event: draw ROI
        if xM is not None and 0 <= xM <= maxX and yM is not None and 0 <= yM <= maxY:  # mouse in crop limits
            # print('** xM = ', xM, ', yM = ', yM)
            linew = 1
            col = 'r'

            if np.isnan(roiS[roiNo, 0]):
                # add the cross
                ulx = xM
                uly = yM
                axIM1.plot([ulx, ulx], [0, maxY], color=col, linestyle='--', linewidth=linew)
                axIM1.plot([0, maxX], [uly, uly], color=col, linestyle='--', linewidth=linew)
            else:  # ul edge is already chosen
                # add the frame
                # use (ulx, uly) and (lrx, lry)
                if xM >= roiS[roiNo, 0] - crop_ulx:
                    ulx = roiS[roiNo, 0] - crop_ulx
                    lrx = xM
                else:
                    ulx = xM
                    lrx = roiS[roiNo, 0] - crop_ulx
                if yM >= roiS[roiNo, 2] - crop_uly:
                    uly = roiS[roiNo, 2] - crop_uly
                    lry = yM
                else:
                    uly = yM
                    lry = roiS[roiNo, 2] - crop_uly
                axIM1.plot([ulx, lrx, lrx, ulx, ulx],
                           [uly, uly, lry, lry, uly], color=col, linestyle='--', linewidth=1)
                # ca m'a l'air plus joli, mais c'est seulement légèrement plus rapide


    def modify_mask(self):
        print("modify_mask called")

        """ Adds or removes the third line of 'roiS' to 'masked'
        """
        global masked  # list of tuples

        # Generate list of selected points
        points_new = []  # list of tuples
        ulx = int(np.round(roiS[roiNo, 0]))
        lrx = int(np.round(roiS[roiNo, 1]))
        uly = int(np.round(roiS[roiNo, 2]))
        lry = int(np.round(roiS[roiNo, 3]))
        # fill columnwise
        for xCoo in range(ulx, lrx + 1):  # from left to right
            # column loop
            for yCoo in range(uly, lry + 1):  # from upper to lower
                # line loop
                points_new.append((xCoo, yCoo))  # (x, y) coordinates

        # add or remove points_new to masked
        if self.CoboMaskMode.currentText() == "Add pixels":
            #             # add really new points to MaskList
            masked = list(set(masked + points_new))  # sorts at the same time
        elif self.CoboMaskMode.currentText() == "Remove pixels":
            # remove entries of newPoints that existed in MaskList
            masked = list(set(masked) - set(points_new))  # sorts at the same time
        # TODO: move to absolute coords (is it simply a + as for a numpy array?)

        print('len masked = ', len(masked))
        self.lbMaskInfo.setText("{:.0f} pixels masked".format(len(masked)))

    def scroll_para_history(self):
        print("scroll_para_history called")

        hist_idx = self.sb_hist.value()
        try:
            for col_idx in range(self.TbFit2.columnCount()):
                self.TbFit2.cellWidget(3, col_idx).setText(b_fits[hist_idx][col_idx]["status"])
                self.TbFit2.cellWidget(4, col_idx).setText("{:.2f}".format(b_fits[hist_idx][col_idx]["GOF"]))
                self.TbFit2.cellWidget(5, col_idx).setText("{:.3f}".format(b_fits[hist_idx][col_idx]["w1"]))
                self.TbFit2.cellWidget(6, col_idx).setText("{:.2f}".format(b_fits[hist_idx][col_idx]["max"]))

                if col_idx in [0, 2]:
                    self.TbFit2.cellWidget(7, col_idx).setText("{:.2f}".format(b_fits[hist_idx][col_idx]["x_0"]))
                    self.TbFit2.cellWidget(8, col_idx).setText("{:.2f}".format(b_fits[hist_idx][col_idx]["y_0"]))

                if col_idx == 2:
                    self.TbFit2.cellWidget(9, col_idx).setText("{:.2f}".format(b_fits[hist_idx][col_idx]["w2"]))
                    self.TbFit2.cellWidget(10, col_idx).setText("{:.2f}".format(b_fits[hist_idx][col_idx]["angle"]))
        except (TypeError, AttributeError):
            pass

    def read_to_b_fits(self):
        print("read_to_b_fits called")

        global b_fits  # (history_length, 3) array with elements dict (later)

        #try:
        # save old values
        b_fits[1:] = b_fits[:-1]  # drop the last entry. Fist entry is now double

        for col_idx in range(self.TbFit2.columnCount()):
            if col_idx in [0, 2]:
                b_fits[0][col_idx] = {
                    "name": self.TbFit2.horizontalHeaderItem(col_idx).text(),
                    "lim": float(self.TbFit2.cellWidget(1, col_idx).text()),
                    "status": "not launched",  # only output
                    "GOF": np.nan,  # only output
                    "w1": float(self.TbFit2.cellWidget(5, col_idx).text()),
                    "max": float(self.TbFit2.cellWidget(6, col_idx).text()),
                    "x_0": float(self.TbFit2.cellWidget(7, col_idx).text()),
                    "y_0": float(self.TbFit2.cellWidget(8, col_idx).text())
                }  # dict["b"] - dict["a"]

                if col_idx == 2:
                    b_fits[0][col_idx]["w2"] = float(self.TbFit2.cellWidget(9, col_idx).text())
                    b_fits[0][col_idx]["angle"] = float(self.TbFit2.cellWidget(10, col_idx).text())

            elif col_idx == 1:
                b_fits[0][col_idx] = {
                    "name": self.TbFit2.horizontalHeaderItem(col_idx).text(),
                    "lim": float(self.TbFit2.cellWidget(1, col_idx).text()),
                    "status": "not launched",  # only output
                    "GOF": np.nan,  # only output
                    "w1": float(self.TbFit2.cellWidget(5, col_idx).text()),
                    "max": float(self.TbFit2.cellWidget(6, col_idx).text())
                }
        # except:
        #    return


    def make_fits(self):
        print("make_fits called")

        """
        Makes the fits
        :return:
        """
        global b_fits  # (history_length, 3) array with elements dict (later)

        self.read_to_b_fits()
        # print(b_fits)
        main.statBar.showMessage('Fitting in progress (slow for elliptic Gaussian)')

        for col_idx in range(self.TbFit2.columnCount()):
            if self.TbFit2.cellWidget(0, col_idx).currentText() == "No":
                continue
            print("yes: " + self.TbFit2.horizontalHeaderItem(col_idx).text())
            # write initialization using .guess method of fit model (see fitting module) ?
            # no not here at least. In onTabChange?

            if col_idx == 0:  # col_idx in [0, 2]  # round Gaussian: gauss2D_cst_offs
                p_ini = [b_fits[0][col_idx]["max"] * b_max[2],  # because the max is relative
                         b_fits[0][col_idx]["x_0"],
                         b_fits[0][col_idx]["y_0"],
                         b_fits[0][col_idx]["w1"]]
                # As the table displays the fit result this initializes with the last result (if not edited)

                XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
                # resid_2D(self, p, x, y, z, **kwargs):
                p_fit, success = leastsq(
                    gauss2D_cst_offs.resid_2D,
                    p_ini,
                    args=(XX[(im_to_show2 > b_fits[0][col_idx]["lim"] * b_max[2])
                             & (im_to_show2 < rel_sat_limit * sat_value - backg)],
                          YY[(im_to_show2 > b_fits[0][col_idx]["lim"] * b_max[2])
                             & (im_to_show2 < rel_sat_limit * sat_value - backg)],
                          im_to_show2[(im_to_show2 > b_fits[0][col_idx]["lim"] * b_max[2])
                                      & (im_to_show2 < rel_sat_limit * sat_value - backg)]),
                    full_output=False)
                if success in [1, 2, 3, 4]:  # If success is equal to 1, 2, 3 or 4, the solution was found.
                    # save old values
                    b_fits[1:] = b_fits[:-1]  # drop the last entry. Fist entry is now double
                    b_fits[0][col_idx]["status"] = "OK: {:d}".format(success)
                    b_fits[0][col_idx]["max"] = p_fit[0] / b_max[2]  # TODO: HERE it modifies b_fits[0, ..][] AND b_fits[1, ..][]
                    b_fits[0][col_idx]["x_0"] = p_fit[1]
                    b_fits[0][col_idx]["y_0"] = p_fit[2]
                    b_fits[0][col_idx]["w1"] = np.abs(p_fit[3])
                    self.sb_hist.setValue(0)
                    self.TbFit2.cellWidget(3, col_idx).setText(b_fits[0][col_idx]["status"])
                    self.TbFit2.cellWidget(4, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["GOF"]))
                    self.TbFit2.cellWidget(5, col_idx).setText("{:.3f}".format(b_fits[0][col_idx]["w1"]))
                    self.TbFit2.cellWidget(6, col_idx).setText("{:.4f}".format(b_fits[0][col_idx]["max"]))
                    self.TbFit2.cellWidget(7, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["x_0"]))
                    self.TbFit2.cellWidget(8, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["y_0"]))
                else:
                    b_fits[0][col_idx]["status"] = "Problem: {:d}".format(success)
                    print("Round Gaussian: Fitting problem encountered")

            elif col_idx == 2:  # elliptic gau: gaussEll2D_cst_offs
                # paras = [vert_factor, cent_x, cent_y, waist_rad_a, waist_rad_b, theta]
                p_ini = [b_fits[0][col_idx]["max"] * b_max[2],  # because the max is relative
                         b_fits[0][col_idx]["x_0"],
                         b_fits[0][col_idx]["y_0"],
                         b_fits[0][col_idx]["w1"],
                         b_fits[0][col_idx]["w2"],
                         b_fits[0][col_idx]["angle"]]
                # As the table displays the fit result this initializes with the last result (if not edited)

                XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
                # resid_2D(self, p, x, y, z, **kwargs):
                p_fit, success = leastsq(
                    gaussEll2D_cst_offs.resid_2D,  # gaussEll2D_cst_offs
                    p_ini,
                    args=(XX[(im_to_show2 > b_fits[0][col_idx]["lim"] * b_max[2])
                             & (im_to_show2 < rel_sat_limit * sat_value - backg)],
                          YY[(im_to_show2 > b_fits[0][col_idx]["lim"] * b_max[2])
                             & (im_to_show2 < rel_sat_limit * sat_value - backg)],
                          im_to_show2[(im_to_show2 > b_fits[0][col_idx]["lim"] * b_max[2])
                                      & (im_to_show2 < rel_sat_limit * sat_value - backg)]),
                    full_output=False)
                if success in [1, 2, 3, 4]:  # If success is equal to 1, 2, 3 or 4, the solution was found.
                    # save old values
                    b_fits[1:] = b_fits[:-1]  # drop the last entry. First entry is now double
                    b_fits[0][col_idx]["status"] = "OK: {:d}".format(success)
                    b_fits[0][col_idx]["max"] = p_fit[0] / b_max[
                        2]  # TODO: HERE it modifies b_fits[0, ..][] AND b_fits[1, ..][]
                    b_fits[0][col_idx]["x_0"] = p_fit[1]
                    b_fits[0][col_idx]["y_0"] = p_fit[2]
                    # Reduce the angle to (-90°, 90°] and set w1 as the long half axis.
                    red_angle = p_fit[5] % 180
                    if red_angle > 90:
                        red_angle = -(180 - red_angle)
                    w1 = np.abs(p_fit[3])
                    w2 = np.abs(p_fit[4])
                    if np.abs(p_fit[3]) < np.abs(p_fit[4]):  # if w1 < w2: exchange and rotate angle by 90°
                        buffer = w1
                        w1 = w2
                        w2 = buffer
                        if red_angle < 0:
                            red_angle += 90
                        else:
                            red_angle -= 90
                    b_fits[0][col_idx]["w1"] = w1
                    b_fits[0][col_idx]["w2"] = w2
                    b_fits[0][col_idx]["angle"] = red_angle
                    self.sb_hist.setValue(0)
                    self.TbFit2.cellWidget(3, col_idx).setText(b_fits[0][col_idx]["status"])
                    self.TbFit2.cellWidget(4, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["GOF"]))
                    self.TbFit2.cellWidget(5, col_idx).setText("{:.3f}".format(b_fits[0][col_idx]["w1"]))
                    self.TbFit2.cellWidget(6, col_idx).setText("{:.4f}".format(b_fits[0][col_idx]["max"]))
                    self.TbFit2.cellWidget(7, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["x_0"]))
                    self.TbFit2.cellWidget(8, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["y_0"]))
                    self.TbFit2.cellWidget(9, col_idx).setText("{:.3f}".format(b_fits[0][col_idx]["w2"]))
                    self.TbFit2.cellWidget(10, col_idx).setText("{:.1f}".format(b_fits[0][col_idx]["angle"]))
                else:
                    b_fits[0][col_idx]["status"] = "Problem: {:d}".format(success)
                    print("Elliptic Gaussian: Fitting problem encountered")

            elif col_idx == 1:  # sot_r_gau
                p_ini = [b_fits[0][col_idx]["max"], b_fits[0][col_idx]["w1"]]
                # As the table displays the fit result this initializes with the last result (if not edited)

                # residuals(self, p, x, y, z, **kwargs):
                p_fit, success = leastsq(sot_r_gau.residuals, p_ini,
                                         args=(sot_x[(sot_x > b_fits[0][col_idx]["lim"])
                                                     & (sot_x < (rel_sat_limit * sat_value - backg)/b_max[2])],
                                               sot_y[(sot_x > b_fits[0][col_idx]["lim"])
                                                     & (sot_x < (rel_sat_limit * sat_value - backg)/b_max[2])]),
                    full_output=False)
                if success in [1, 2, 3, 4]:  # If success is equal to 1, 2, 3 or 4, the solution was found.
                    # save old values
                    b_fits[1:] = b_fits[:-1]  # drop the last entry. Fist entry is now double
                    b_fits[0][col_idx]["status"] = "OK: {:d}".format(success)
                    b_fits[0][col_idx]["max"] = p_fit[0]  # HERE it modifies b_fits[0][..][] AND b_fits[1][..][]
                    b_fits[0][col_idx]["w1"] = np.abs(p_fit[1])
                    self.sb_hist.setValue(0)
                    self.TbFit2.cellWidget(3, col_idx).setText(b_fits[0][col_idx]["status"])
                    self.TbFit2.cellWidget(4, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["GOF"]))
                    self.TbFit2.cellWidget(5, col_idx).setText("{:.3f}".format(b_fits[0][col_idx]["w1"]))
                    self.TbFit2.cellWidget(6, col_idx).setText("{:.2f}".format(b_fits[0][col_idx]["max"]))
                else:
                    b_fits[0][col_idx]["status"] = "Problem: {:d}".format(success)
                    print("SOT curve: Fitting problem encountered")
        main.statBar.showMessage('Fitting finished', 5000)
        self.TbFit2.resizeColumnsToContents()

    def roi_def(self):  # called by checkbox
        print("roi1_def called")
        global roiS, roiNo
        if self.cbROI1redefine.isChecked():
            roiNo = 0  # first line in roi array
        elif self.cbMaskActivateIt.isChecked():
            roiNo = 2  # third line in roi array
        # if one cb was checked set the corresponding line to nan
        roiS[roiNo, :] = np.ones(4)*np.nan
        if self.rdbtGroupView.checkedId() == 1:  # switch to single frame view if tile view was active
            self.rdbtFbF.setChecked(True)
            self.display()
        self.showRoiS()

    def make_comment_sheet_for_export(self):
        print("make_comment_sheet_for_export called")

        global mywb  # was generated at the very beginning

        # check if the sheet "comments" exists.
        if "comments" in mywb.sheetnames:
            # If yes, delete and redefine (empty)
            ws1 = mywb["comments"]
            mywb.remove(ws1)
            ws1 = mywb.create_sheet("comments")
        else:
            # if not, create from active sheet (sheet1)
            ws1 = mywb.active  # worksheet
            ws1.title = "comments"

        # write the contents
        ws1.cell(row=1, column=1).value = \
            "Describes how the numbers in the data sheet were obtained."

        ws1.cell(row=3, column=1).value = "File information: "
        ws1.cell(row=4, column=2).value = "Location of treated file: "
        ws1.cell(row=4, column=3).value = dirname
        ws1.cell(row=5, column=2).value = "File type: "
        ws1.cell(row=5, column=3).value = file_type
        ws1.cell(row=6, column=2).value = "File name: "
        ws1.cell(row=6, column=3).value = baseName

        ws1.cell(row=8, column=1).value = "Background correction method used: "
        ws1.cell(row=9, column=2).value = "Dark image number: "
        if self.cob_dark_im.currentText() == "Use a dark image":
            ws1.cell(row=9, column=3).value = str(self.SbDark.value())
        else:
            ws1.cell(row=9, column=3).value = "No dark image used"
        ws1.cell(row=10, column=2).value = "Size of analyzed image (x, y in px): "
        # ws1.cell(row=10, column=3).value = size
        if self.cb_no_auto_crop.isChecked():
            ws1.cell(row=11, column=2).value = "Symmetric auto-crop was disabled."
        else:
            ws1.cell(row=11, column=2).value = "Symmetric auto-crop was enabled."
        if self.cob_use_mean.currentText() == "Use mean from ROI":
            if self.cob_roi.currentText() == "Outside ROI":
                mess = "Mean outside ROI"
            else:
                mess = "Mean inside ROI"
        else:
            mess = "Horizontal Aeff-curve fit"  # using ..pts, .. blue points
        ws1.cell(row=12, column=2).value = mess
        ws1.cell(row=12, column=2).font = xl.styles.Font(bold=True)

        ws1.cell(row=14, column=1).value = "Maximum smoothing method used for Aeff: "
        if (self.cbUseModMax.isChecked() and
                self.rdbtgroup_for_aeff.checkedId() >= 0 and
                self.TbFit2.cellWidget(0, self.rdbtgroup_for_aeff.checkedId()).currentText() == "Yes"):
            # take the header of the checked rdbt
            mess = "Used global fit: "
            mess += self.TbFit2.horizontalHeaderItem(self.rdbtgroup_for_aeff.checkedId()).text()
            ws1.cell(row=15, column=2).value = mess
        else:  # take the text of the combo_box
            ws1.cell(row=15, column=2).value = self.cob_max.currentText()
        ws1.cell(row=15, column=2).font = xl.styles.Font(bold=True)

        ws1.cell(row=17, column=1).value = "Analytical model-beams fitted to the data:"
        fitted_idx_li = []
        for col_idx in range(3):
            if self.TbFit2.cellWidget(0, col_idx).currentText() == "Yes":
                fitted_idx_li.append(col_idx)  # 0: Round Gauss; 1: SOT fit; 2: ell.G.
        if len(fitted_idx_li) == 0:
            ws1.cell(row=18, column=2).value = "No beam fits were performed."
            return

        ws1.cell(row=18, column=2).value = "Model name"
        ws1.cell(row=19, column=2).value = "Vert. fit limit (rel.)"
        for mod_idx in range(len(fitted_idx_li)):
            ws1.cell(row=18, column=3+mod_idx).value = \
                self.TbFit2.horizontalHeaderItem(fitted_idx_li[mod_idx]).text()
            ws1.cell(row=18, column=3+mod_idx).font = xl.styles.Font(bold=True)
            ws1.cell(row=19, column=3+mod_idx).value = \
                self.TbFit2.cellWidget(1, fitted_idx_li[mod_idx]).text()

        ws1.append([""])  # continue with empty line
        ws1.append(["Relative values (rel.) in GL (grey levels)"])
        ws1.append(["are defined with respect to the maximum used for the Aeff calculation. "])
        ws1.append(["Thus the absolute vertical fit limit depends slightly on the "])
        ws1.append(["maximum smoothing method chosen for the Aeff calculation. "])
        ws1.append(["If the Aeff max is taken from the fit, the vertical fit limit depends "])
        ws1.append(["on the max obtained by the method displayed on the combobox on tab2. "])

    def export(self):
        print("export called")

        """
        "results"-sheet: ws2
        """

        # check if the sheet "results" exists.
        if "results" in mywb.sheetnames:
            # If yes, delete
            ws2 = mywb["results"]
            mywb.remove(ws2)
        ws2 = mywb.create_sheet("results", 0)  # empty results-sheet goes first

        # write reference of analysis: file name and time of analysis
        ws2.cell(row=1, column=1).value = "File: "
        ws2.cell(row=1, column=2).value = os.path.normpath(dirname+"/"+baseName+"."+file_type)
        ws2.cell(row=2, column=1).value = "Analyzed on: "
        tod = dt.datetime.today()
        tod_str = "{:0d}-{:0d}-{:0d} {:0d}h{:0d}m{:0d}s".format(
            tod.year, tod.month, tod.day, tod.hour, tod.minute, tod.second)
        ws2.cell(row=2, column=2).value = tod_str
        ws2.cell(row=2, column=4).value = "Used no of imges: "
        ws2.cell(row=2, column=5).value = len(good_filt_idx)
        ws2.cell(row=2, column=7).value = "Err. confidence"
        ws2.cell(row=2, column=8).value = confidence  # H2

        # see what should be exported. Cols that are not shown in the table are always exported.
        key_li = list(results.keys())  # list of all possible columns
        # print("key_li: ", key_li)
        # print("shown_li: ", shown_li)
        export_idx = []  # get idx of columns to export
        for col_idx in range(len(key_li)):
            if col_idx in shown_li:
                # check if export wanted
                if self.tb_res3.cellWidget(0, shown_li.index(col_idx)).currentText() == "Yes":
                    export_idx.append(col_idx)
            else:  # Cols that are not shown in the table are always exported.
                export_idx.append(col_idx)
        # print("export_idx: ", export_idx)

        li_list = []
        for exp_col_idx in range(len(export_idx)):
            li_list.append(key_li[export_idx[exp_col_idx]])
        ws2.append([""])  # continue with empty line
        ws2.append([""] + li_list[1:])  # write headers (except "image numbers") on top of stat lines

        # write first column indicating the meaning of the statistics lines
        str_li = ['Mean', 'Rel err.', 'Std err.', 'Min.', 'Max.', 'Median']
        for idx in range(len(str_li)):
            ws2.cell(row=5 + idx, column=1).value = str_li[idx]

        ws2.append([""])  # continue with empty line and then the header in row 11
        ws2.append(li_list)  # repeat headers on top of data lines
        # Data in row 13 and below
        # (row by row so later it may be changed to faster version)
        for li_idx in range(len(good_filt_idx)):  # good_filt_idx, good_idx
            li_list = []
            for exp_col_idx in range(len(export_idx)):
                li_list.append(results[key_li[export_idx[exp_col_idx]]][good_filt_idx[li_idx]])
            ws2.append(li_list)  # appends a row the data of all columns

        # Add statistical formulas in the first lines
        for col_idx in range(1, len(export_idx)):
            range_str = (col_2_str(col_idx + 1) + str(13) + ':' +
                         col_2_str(col_idx + 1) + str(13+len(good_idx)))
            ws2.cell(row=5, column=col_idx + 1).value = '=AVERAGE(' + range_str + ')'
            abs_err_str = ('=STDEVA(' + range_str +
                           ')/SQRT(COUNT(' + range_str + ')) * TINV(1-H2, COUNT(' + range_str + ')-1)')
            ws2.cell(row=6, column=col_idx + 1).value = (abs_err_str + '/AVERAGE(' + range_str + ')')
            ws2.cell(row=6, column=col_idx + 1).number_format = '0.00%'
            ws2.cell(row=7, column=col_idx + 1).value = abs_err_str
            ws2.cell(row=8, column=col_idx + 1).value = '=MIN(' + range_str + ')'
            ws2.cell(row=9, column=col_idx + 1).value = '=MAX(' + range_str + ')'
            ws2.cell(row=10, column=col_idx + 1).value = '=MEDIAN(' + range_str + ')'
        """
        str_li = ['Mean', 'Rel err.', 'Std err.', 'Min.', 'Max.', 'Median']
        line no.    5        6            7         8       9        10
        13 ff are the data lines
        """

        fpath = os.path.normpath((dirname + "/" + baseName + " " + tod_str + ".xlsx"))
        mywb.save(fpath)

        main.statBar.showMessage("Finished Exporting xlsx-file", 5000)
        """ Forbidden characters in windows paths:   
        < (less than)
        > (greater than)
        : (colon - sometimes works, but is actually NTFS Alternate Data Streams)
        " (double quote)
        / (forward slash)
        \ (backslash)
        | (vertical bar or pipe)
        ? (question mark)
        * (asterisk)
        
        # open a dialog for input of the filename
        f_name = QFileDialog.getSaveFileName(self,
                                            'Save data to file',
                                            '../output/',
                                            "All files (*.*) ;; Excel files (*.xlsx *.xls)")
        # save the workbook
        if f_name:  # a file was chosen
            mywb.save(f_name)
        """

    def GetMousePlot3(self, event):
        # Manages the drawing of the lines depending on the mouse position
        print("GetMousePlot3 called")
        # useless events
        # print("event.inaxes", event.inaxes)  # is an axes of the figure check with "is"
        print("event.xdata", event.xdata)  # can be used to find the which bin was clicked
        # print("event.button", event.button)  # to see if it was left or right clicked:
        # MouseButton.LEFT = 1; middle = 2; right = 3

        if event.button != 3:
            print("This was not the right mouse button")
            return

        if event.inaxes is None:
            return
        ax_idx = axPlt_li.index(event.inaxes)
        print("Axes index: ", ax_idx)

        if "POS" in CobPl_li[ax_idx].currentText().upper():
            print("this is no histogram")
            return

        # redo histogram and get elements of bin
        if not type(results[CobPl_li[ax_idx].currentText()]) is np.ndarray:
            print("this is no numerical data")
            return  # do not do what follows

        if np.any(np.isnan(results[CobPl_li[ax_idx].currentText()])):
            print("there are nans in the data")
            return  # do not do what follows

        _, bins = np.histogram(results[CobPl_li[ax_idx].currentText()], bins="auto")
        # input bins="auto", bins=15, np.histogram is used by pyplot in plot3_1D
        idx = 1
        while (event.xdata > bins[idx]) and (idx < len(bins)):
            idx += 1
        lim_low = bins[idx - 1]
        lim_high = bins[idx]
        # print("bins: ", bins)
        # print("lim_low: ", lim_low)
        # print("lim_high: ", lim_high)

        bin_im_nums = []  # the numbers of the images in the clicked bin
        # print("results:", results[CobPl_li[ax_idx].currentText()])
        val_min = results[CobPl_li[ax_idx].currentText()].min()
        for idx in range(len(good_idx)):
            value = results[CobPl_li[ax_idx].currentText()][idx]
            if lim_low == val_min:  # the first bin is closed on both sides
                if (value >= lim_low) and (value <= lim_high):
                    bin_im_nums.append(good_idx[idx] + 1)
            else:  # the other bins are closed only on the right side
                if (value > lim_low) and (value <= lim_high):
                    bin_im_nums.append(good_idx[idx] + 1)
        # toggle bin_im_nums in filtered images
        before = set(self.range_str_conv(self.EdFiFilt.text()))
        new = set(bin_im_nums)
        new = list(new ^ before)  # or list(set([short list]).symmetric_difference([long list]))
        new = sorted(new)
        self.EdFiFilt.setText(self.range_str_conv(new))

        # update the plots
        self.show_res_statistics_table(ti.time())
        self.plot3_1D()

    def GetMouseIm1(self, event):
        #print("GetMouseIm1 called")

        # Manages the drawing of the lines depending on the mouse position

        # useless events
        if event.inaxes is None:
            return
        # Cases where nothing should be done = inverse of cases of action
        if not(self.cbROI1redefine.isChecked()
               or self.cbMaskActivateIt.isChecked()):
            return

        # useful events
        self.UpdateLinesIm1(int(np.round(event.xdata)),
                            int(np.round(event.ydata)))
        self.canvasIM1.draw()


    def GetKeyIm1(self, event):
        print("GetKeyIm1 called")
        # Manipulates the Mask (Pixels not to use for the fit evaluation)
        # h : upper left corner
        # n : lower right corner
        # b : single pixel in mask mode
        global roiS  # for keeping the memory between two calls:
        # roiS un champ 3x4
        # A line contains: x_min, x_max, y_min, y_max of the ROI if it is defined
        # Nan after ticking the cb
        # x_min, nan, y_min, nan during definition
        # the choice of the line is done using: if elif elif sur les cb.

        print('Im1: ', event.key)
        #print('Im1: ', event.ydata)

        # useless events
        if event.inaxes is None: return
        # Cases where nothing should be done = inverse of cases of action
        if not(self.cbROI1redefine.isChecked() or self.cbMaskActivateIt.isChecked()):
            return

        letter = str(event.key).upper()
        if letter == 'H':
            # memorize first point
            # TODO: To enable for zooming and dezooming by cropping (for precise
            # masking of dead camera pixels) I should pass the coordinates of
            # the masked pixels to absolute coordinates.
            # but is this necessary ? the lens tool does it !
            roiS[roiNo, 0] = np.round(event.xdata) + crop_ulx
            roiS[roiNo, 2] = np.round(event.ydata) + crop_uly
            # reset second point
            roiS[roiNo, 1] = np.nan
            roiS[roiNo, 3] = np.nan
            self.showRoiS()
            finished = False
        if letter == 'B':
            # memorize the single selected point
            roiS[roiNo, 0] = np.round(event.xdata) + crop_ulx
            roiS[roiNo, 2] = np.round(event.ydata) + crop_uly
            # set second point identical to first one
            roiS[roiNo, 1] = np.round(event.xdata) + crop_ulx
            roiS[roiNo, 3] = np.round(event.ydata) + crop_uly
            self.showRoiS()
            finished = True
        elif letter == 'N':
            # memorize second point
            roiS[roiNo, 1] = np.round(event.xdata) + crop_ulx
            roiS[roiNo, 3] = np.round(event.ydata) + crop_uly

            if not(np.isnan(roiS[roiNo, 0]) or np.isnan(roiS[roiNo, 2])):  # the rectangle gets finished
                # flip points if necessary
                if roiS[roiNo, 1] < roiS[roiNo, 0]:  # right x should be larger than left x
                    buffer = roiS[roiNo, 0]
                    roiS[roiNo, 0] = roiS[roiNo, 1]
                    roiS[roiNo, 1] = buffer
                if roiS[roiNo, 3] < roiS[roiNo, 2]:  # lower y should be larger than upper y
                    buffer = roiS[roiNo, 2]
                    roiS[roiNo, 2] = roiS[roiNo, 3]
                    roiS[roiNo, 3] = buffer

                self.showRoiS()
                finished = True
        # end letter 'N'

        if finished:
            # Tell other function to no longer update the rectangle drawing
            if roiNo == 0:  # first line in roi array
                self.cbROI1redefine.setChecked(False)
            elif roiNo == 2:  # third line in roi array
                self.modify_mask()  # modifies the list of masked points
                self.cbMaskActivateIt.setChecked(False)

        print(roiS[roiNo, :])

    def crop_or_uncrop(self):
        """Is called by
        self.CbCrop.stateChanged.connect(self.crop_or_uncrop)
        """
        print('crop_or_uncrop called')
        global imStack, crop_ulx, crop_uly
        global im1_height, im1_width

        if self.CbCrop.isChecked():  # take a view
            crop_ulx = self.SbLeft.value()
            crop_uly = self.SbUpper.value()
            imStack = imStack[crop_uly:(im1_height - 1 - self.SbLower.value()),
                              crop_ulx:(im1_width - 1 - self.SbRight.value()),
                              :]
            im1_height = imStack.shape[0]
            im1_width = imStack.shape[1]
            self.SbLower.setEnabled(False)
            self.SbRight.setEnabled(False)
            self.SbLeft.setEnabled(False)
            self.SbUpper.setEnabled(False)

            self.rdbtFbF.setChecked(True)
            self.display()
            self.get_sat_pixels()  # refresh saturated pixels number
        else:
            crop_ulx = 0
            crop_uly = 0
            self.SbLower.setEnabled(True)
            self.SbRight.setEnabled(True)
            self.SbLeft.setEnabled(True)
            self.SbUpper.setEnabled(True)
            self.rdbtFbF.setChecked(True)
            self.display()
            self.load_stack()  # reload the full field images

    def get_local_roi_line(self, line_idx=0):
        loc_roi_li = np.round(roiS[line_idx]).astype(int)  # extract good roi line and convert to int
        loc_roi_li[:2] -= crop_ulx  # switch to relative coords (with respect to crop)
        loc_roi_li[2:] -= crop_uly  # xfrom, xto, yfrom, yto
        loc_roi_li[:2] -= auto_crop_ulx  # switch to relative coords (with respect to crop)
        loc_roi_li[2:] -= auto_crop_uly  # xfrom, xto, yfrom, yto
        return loc_roi_li

    def medFi_state_changed(self):
        print("medFi_state_changed called")

        global oldStack, imStack
        newBtIdx = self.rdbtgroupMedFi.checkedId()
        if newBtIdx == 1:  # preview became active
            self.rdbtFbF.setChecked(True)
            self.display_MedFi_preview()
            self.get_sat_pixels()  # to refresh the nb of saturated pixels
        elif newBtIdx == 2:  # Apply filter to stack (but backup before)
            oldStack = imStack.copy()  # make a backup
            MedFiPara = self.SbMedFi.value()
            main.statBar.showMessage('Processing ...  please wait')
            for Nb in range(1, imStack.shape[2] + 1):
                imStack[:, :, Nb - 1] = ndi.median_filter(imStack[:, :, Nb - 1], MedFiPara)
                main.progBar.setValue(Nb)
                QApplication.processEvents()
            main.statBar.showMessage('Finished', 2000)
            main.progBar.setValue(0)
            # main.statBar.showMessage('Ready')
            self.rdbtFbF.setChecked(True)
            self.display()
            self.get_sat_pixels()# for refresh saturated pixels number
        elif newBtIdx == 0:  # revert to earlier stage if possible
            try:
                imStack = oldStack.copy()
                self.rdbtFbF.setChecked(True)
                self.display()
                self.get_sat_pixels() # for refresh saturated pixels number
            except NameError:  # necessary because the method is executed on initialisation
                return

    def medFi_sb_changed(self):
        print("medFi_sb_changed called")

        if self.rdbtgroupMedFi.checkedId() == 1:
            self.rdbtFbF.setChecked(True)
            self.display_MedFi_preview()

    def display_MedFi_preview(self):
        print('display_MedFi_preview called')
        try:
            number = self.SlNum.value()  # read present slider value
            self.LbImNo.setText(str(number))  # display it in the label below the image

            if self.SbMedFi.value() > 1:
                main.statBar.showMessage('Processing')
                ImDisp1 = ndi.median_filter(imStack[:, :, number - 1], self.SbMedFi.value())
                main.statBar.showMessage('Ready')
            else:
                ImDisp1 = imStack[:, :, number - 1]
            self.showIm1(ImDisp1)  # show the image
        except NameError:
            return

    def onTabChange(self, tab_num):
        print("onTabChange called")

        global tab_idx_curr, tab_idx_old
        global auto_crop_ulx, auto_crop_uly
        global ti0

        tab_idx_old = tab_idx_curr
        tab_idx_curr = tab_num
        print("tab_idx_curr", tab_idx_curr)

        if len(good_idx) == 0:
            return  # do not do the rest if tab 1 is not finished

        if (tab_idx_old == 0) and (tab_idx_curr == 1):
            try:
                tab1_im_idx = good_idx.index(self.SlNum.value() - 1)
            except ValueError:
                tab1_im_idx = 0

            self.SlNum2.setMinimum(1)
            self.SlNum2.setMaximum(len(good_idx))
            if self.SlNum2.value() == tab1_im_idx + 1:
                # self.read_to_b_fits()
                self.onImageChange()
                # # change back to the list of calls with auto-init of fits inserted after get_beam_max
                # self.get_step()
                # self.make_auto_crop()
                # self.get_backg()  # calls make_auto_crop_dark if needed
                # self.get_beam_max()
                # self.get_aeffg()
                # self.make_fits()
                # self.display2()
                # self.plot2_1D()
            else:
                self.SlNum2.setValue(tab1_im_idx + 1)
                self.LbImInfo.setText("Image " + str(tab1_im_idx))

        elif (tab_idx_old == 1) and (tab_idx_curr == 2):
            self.LbTiWarn.setText("This will take about {:1.2e} s.".format(len(good_idx) * ti0))  # label

        elif tab_idx_curr == 0:
            auto_crop_ulx = 0  # reset autocrop to 0 when going back to tab 1
            auto_crop_uly = 0

    def onImageChange(self):
        print("onImageChange called")

        global ti0
        ti0 = ti.time()
        self.get_step()
        self.make_auto_crop()  # generates im_to_show2
        self.get_backg()  # calls make_auto_crop_dark if needed; updates im_to_show2
        self.get_beam_max()
        self.get_sot_data()
        self.make_fits()
        self.display2()
        self.plot2_1D()
        self.get_aeffg()  # can use b_max[2] or b_fit value as max
        ti0 = ti.time() - ti0

    def get_beam_max(self):
        print("get_beam_max called")

        global b_max
        fit_pb = False
        """, 'Cap fit 80%', 'Cap fit 70%',"""
        if 'Cap fit' in self.cob_max.currentText():
            # Initialize the (round) Gaussian fit with the 3x3 mean around the maximal pixel
            # The cap_factor will be defined with respect to the 3x3 mean too.
            (b_max[1], b_max[0]) = np.unravel_index(im_to_show2.argmax(), im_to_show2.shape)
            (bmx, bmy) = b_max[0:2].round().astype("int")  # b_max[0] = x-position
            b_max[2] = im_to_show2[bmy - 1: bmy + 2, bmx - 1: bmx + 2].mean()
            p_ini = [b_max[2]*0.99, # max_ini = 99% of max pixel
                     b_max[0]-1.5, # x_0 = a bit away from the max pixel
                     b_max[1]-1.5, # y_0 = a bit away from the max pixel
                     np.mean(im_to_show2.shape)/10]  #  supposing there are 5 beam diameters on the image
            XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))

            # resid_2D(self, p, x, y, z, **kwargs):
            if self.cob_max.currentText() == 'Cap fit 95%':
                cap_factor = 0.95
            elif self.cob_max.currentText() == 'Cap fit 90%':
                cap_factor = 0.9
            elif self.cob_max.currentText() == 'Cap fit 80%':
                cap_factor = 0.8
            elif self.cob_max.currentText() == 'Cap fit 70%':
                cap_factor = 0.7
            p_fit, success = leastsq(
                gauss2D_cst_offs.resid_2D,
                p_ini,
                args=(XX[im_to_show2 > b_max[2] * cap_factor],
                      YY[im_to_show2 > b_max[2] * cap_factor],
                      im_to_show2[im_to_show2 > b_max[2] * cap_factor]),
                full_output=False)

            if success in [1, 2, 3, 4]:  # If success is equal to 1, 2, 3 or 4, the solution was found.
                b_max[2] = p_fit[0]
                b_max[0] = p_fit[1]
                b_max[1] = p_fit[2]
            else:
                fit_pb = True  # and use: 'Max pixel (3x3 mean)'
                (b_max[1], b_max[0]) = np.unravel_index(im_to_show2.argmax(), im_to_show2.shape)
                (bmx, bmy) = b_max[0:2].round().astype("int")  # b_max[0] = x-position
                b_max[2] = im_to_show2[bmy - 1: bmy + 2, bmx - 1: bmx + 2].mean()

        elif self.cob_max.currentText() == 'Max pixel':
            b_max[2] = im_to_show2.max()
            (b_max[1], b_max[0]) = np.unravel_index(im_to_show2.argmax(), im_to_show2.shape)
        elif self.cob_max.currentText() == 'Max pixel (3x3 mean)':
            (b_max[1], b_max[0]) = np.unravel_index(im_to_show2.argmax(), im_to_show2.shape)
            (bmx, bmy) = b_max[0:2].round().astype("int")  # b_max[0] = x-position
            b_max[2] = im_to_show2[bmy - 1: bmy + 2, bmx - 1: bmx + 2].mean()
        elif self.cob_max.currentText() == 'Max pixel (5x5 mean)':
            (b_max[1], b_max[0]) = np.unravel_index(im_to_show2.argmax(), im_to_show2.shape)
            (bmx, bmy) = b_max[0:2].round().astype("int")  # b_max[0] = x-position
            b_max[2] = im_to_show2[bmy - 2: bmy + 3, bmx - 2: bmx + 3].mean()
        elif self.cob_max.currentText() == 'Centroid':
            # TODO test : scipy.ndimage.center_of_mass
            ti1 = ti.time()
            XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
            b_max[0] = (XX * im_to_show2).sum() / im_to_show2.sum()
            b_max[1] = (YY * im_to_show2).sum() / im_to_show2.sum()
            b_max[2] = im_to_show2[int(np.round(b_max[1])), int(np.round(b_max[0]))]
            print('Manual centroid b_max = ', b_max)
            print("Centroid calc took {:f} sec.".format(ti.time()-ti1))
        elif self.cob_max.currentText() == 'Centroid (3x3 mean)':
            XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
            b_max[0] = (XX * im_to_show2).sum() / im_to_show2.sum()
            b_max[1] = (YY * im_to_show2).sum() / im_to_show2.sum()
            b_max[2] = im_to_show2[int(np.round(b_max[1])) - 1: int(np.round(b_max[1])) + 2,
                       int(np.round(b_max[0])) - 1: int(np.round(b_max[0])) + 2].mean()
        elif self.cob_max.currentText() == 'Centroid (5x5 mean)':
            XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
            b_max[0] = (XX * im_to_show2).sum() / im_to_show2.sum()
            b_max[1] = (YY * im_to_show2).sum() / im_to_show2.sum()
            b_max[2] = im_to_show2[int(np.round(b_max[1])) - 2: int(np.round(b_max[1])) + 3,
                       int(np.round(b_max[0])) - 2: int(np.round(b_max[0])) + 3].mean()

        if not fit_pb:
            self.LbMax.setText('Found max.: {:.2f} GL'.format(b_max[2]))
            self.LbMaxPos.setText('Found max. position: ({:.2f}/{:.2f}) (x/y) in px'.format(b_max[0], b_max[1]))
            print('b_max = ', b_max)
        else:
            self.LbMax.setText('Fit pb, used max.: {:.2f} GL'.format(b_max[2]))
            print('Fit pb, used b_max = ', b_max)


    def make_auto_crop(self):  # only for the stack imge
        print("make_auto_crop called")

        global im_to_show2  # to avoid multiple execution
        global auto_crop_uly, auto_crop_ulx

        im_idx = good_idx[self.SlNum2.value() - 1]  # read present slider value
        self.LbImInfo.setText("Image " + str(im_idx + 1))  # display it in the label right of the slider button

        im_to_show = imStack[:, :, im_idx].copy()  # make a view of the right image
        # auto crop to square around ROI if not disabled
        if not self.cb_no_auto_crop.isChecked() \
                and (self.cob_roi.currentText() == "Outside ROI"
                     or self.cob_use_mean.currentText() == "Use A<sub>eff</sub>-curve fit"):
            roi_li = np.round(roiS[0]).astype(int)  # extract good roi line and convert to int
            roi_li[:2] -= crop_ulx  # switch to relative coords (with respect to crop)
            roi_li[2:] -= crop_uly  # xfrom, xto, yfrom, yto
            auto_crop_uly = roi_li[2] - frame_width
            auto_crop_ulx = roi_li[0] - frame_width
            im_to_show = im_to_show[
                         auto_crop_uly:roi_li[3] + frame_width,
                         auto_crop_ulx:roi_li[1] + frame_width]
        else:
            auto_crop_uly = 0
            auto_crop_ulx = 0
        im_to_show2 = im_to_show  # .copy()  ???

    def make_auto_crop_dark(self):  # only for the stack imge
        print("make_auto_crop_dark called")

        global auto_crop_uly, auto_crop_ulx

        darkim_cropped = imStack[:, :, self.SbDark.value() - 1].copy()  # make a view of the right image
        # auto crop to square around ROI if not disabled
        if not self.cb_no_auto_crop.isChecked() \
                and (self.cob_roi.currentText() == "Outside ROI"
                     or self.cob_use_mean.currentText() == "Use A<sub>eff</sub>-curve fit"):
            roi_li = np.round(roiS[0]).astype(int)  # extract good roi line and convert to int
            roi_li[:2] -= crop_ulx  # switch to relative coords (with respect to crop)
            roi_li[2:] -= crop_uly  # xfrom, xto, yfrom, yto
            auto_crop_uly = roi_li[2] - frame_width
            auto_crop_ulx = roi_li[0] - frame_width
            darkim_cropped = darkim_cropped[
                             auto_crop_uly:roi_li[3] + frame_width,
                             auto_crop_ulx:roi_li[1] + frame_width]
        else:
            auto_crop_uly = 0
            auto_crop_ulx = 0
        return darkim_cropped  # .copy()  ???

    def cost_aeff_slope_sq(self, offset, PixMax):
        print("cost_aeff_slope_sq called")

        NbPoint = self.sbNbPoint.value()
        Aeff_vec, _ = self.AeffCurveBasic(im_to_show2 - offset, PixMax=PixMax, step=step)
        x_bidon = np.arange(len(Aeff_vec))  # use amplified slope by putting the points closer together
        p = np.polyfit(x_bidon[-NbPoint:], Aeff_vec[-NbPoint:], 1)
        # print('slope = ', p[0])
        return p[0]**2

    def get_backg(self):
        print("get_backg called")

        print('get_backg called')
        global backg
        global energy
        global im_to_show2

        if self.cob_dark_im.currentText() == 'Use a dark image':
            im_to_show2 -= self.make_auto_crop_dark()

        if np.any(np.isnan(roiS[0])):
            self.LbOffSet.setText("The ROI is not defined")
            main.statBar.showMessage("The ROI is not defined", 4000)
            return

        if self.cob_use_mean.currentText() == "Use mean from ROI":  # calculate mean of ROI
            # Create boolean mask 'ma' using 'ROI1', 'Out/In' and 'Mask' information
            ma = np.zeros(im_to_show2.shape, dtype=bool)  # use a roi
            # ma de base
            roi_li = self.get_local_roi_line()
            # count roi borders as inside
            ma[roi_li[2]:roi_li[3] + 1, roi_li[0]:roi_li[1] + 1] = np.ones(
                (roi_li[3] + 1 - roi_li[2], roi_li[1] + 1 - roi_li[0]), dtype=bool)
            # print("taille ma de base: ", np.sum(ma))
            # outside?
            if ((self.cob_roi.currentText() == "Outside ROI")
                    or (self.cob_use_mean.currentText() == "Use A<sub>eff</sub>-curve fit")):
                ma = ~ma  # or use  np.logical_not
            backg = im_to_show2[ma].mean()

        elif self.cob_use_mean.currentText() == "Use A<sub>eff</sub>-curve fit":  # get offset from Aeff-curve fit
            # aeffvec et imsize vec(données)
            result = minimize(self.cost_aeff_slope_sq, x0=backg, args=(b_max[2]))
            # init with good value: x0=backg (last image) is fast (0.55s/63i); x0=0 is slow (0.8s/63i)
            # cost_aeff_slope_sq(self, offset, PixMax):
            backg = result.x[0]  # result.x is an array-like
            print(result.message)

        else:
            self.LbOffSet.setText("PROBLEM WITH BACKGROUND CORRECTION.")
            print("PROBLEM WITH BACKGROUND CORRECTION.")
            return

        # use backg
        self.LbOffSet.setText("Found offset: {:3.2f} GL".format(backg))
        im_to_show2 -= backg

        # get energy of the corrected image
        energy = im_to_show2.sum()
        self.LbEng.setText('Found energy: {:0.4e} GL'.format(energy))
        print('energy = ', energy)


    def AeffCurveBasic(self, image, PixMax=-100, points=20, step=5):
        print("AeffCurveBasic called")

        """ Returns only the outermost points of the Aeff curve.
        :param image:
        :param PixMax: use this maximum, if not given use maximum pixel in remaining image
        :param points: length of curve, should be passed, should be the same as in calling function
        :param step: cropping step in pixels, should be passed, should be the same as in calling function
        :return: 
           AeffS - vector of length 'points' (or less)
           ImSizeS - vector of length 'points' (or less)
        """
        # print('AeffCurveBasic called')
        if PixMax == -100:
            PixMax = image.max()  # Pixel maximal

        AeffS = np.zeros(points)  # line vector
        ImSizeS = np.zeros(points)  # line vector
        Taille = image.shape
        while (Taille[0] - step > step) & (Taille[1] - step > step) & (points > 0):
            usedimage = image[step:Taille[0] - step, step:Taille[1] - step]
            # on enleve une partie de l'image

            Taille = usedimage.shape
            ImSizeS[points - 1] = np.array(Taille).prod()
            AeffS[points - 1] = np.sum(usedimage) / PixMax

            points = points - 1
        return AeffS, ImSizeS

    def get_aeffg(self):
        print("get_aeffg called")

        global Aeffg
        global slope
        # global straightness
        # Should I replace b_max by the beam fit values if those are used for Aeffg? # global b_max
        # print('get_aeffg called')

        if self.cbUseModMax.isChecked() and self.rdbtgroup_for_aeff.checkedId() >= 0:
            print(self.rdbtgroup_for_aeff.checkedId())
            rdbt_no = self.rdbtgroup_for_aeff.checkedId()
            if self.TbFit2.cellWidget(0, rdbt_no + 1).currentText() == "Yes":
                # the chosen fit is activated
                max_for_aeff = b_fits[0][rdbt_no]["max"] * b_max[2]
                # possibly modify b_max here (copy the beam fit values of the max (x, y, z))
            else:
                max_for_aeff = b_max[2]
        else:
            max_for_aeff = b_max[2]

        NbPoint = self.sbNbPoint.value()
        Aeff_vec, ImSize_vec = self.AeffCurveBasic(im_to_show2, PixMax=max_for_aeff, step=step)

        Aeffg = Aeff_vec[-NbPoint:].mean()
        print('Aeffg = ', Aeffg)
        self.LbAeffg.setText('A<sub>eff</sub> : {:.2f} px.'.format(Aeffg))

        self.LbAeffg_to_gau_w.setText('-> w_1 of G.: {:.2f} px.'.format(np.sqrt(2*Aeffg/np.pi)))

        slope = np.polyfit(ImSize_vec[-NbPoint:], Aeff_vec[-NbPoint:], 1)[0]
        self.LbSlope.setText('Slope : {:1.4e}'.format(slope))

        # self.LbStraightness.setText('Straightness : {:1.4e}'.format(straightness))

        self.LbTi2.setText('Last used time {:.3f} ms'.format((ti.time() - ti0) / 1e-3))

    def get_sot_data(self):
        print("get_sot_data called")

        global sot_x, sot_y
        # sot fit-model has parameters max_sot and w_sot
        hi_edges = np.linspace(-0.01, 1.2, num=122)  # makes steps of 1%, needs more than 1 due to smoothing
        # hi_count, _ = np.histogram(im_to_show2 / sat_value, bins=hi_edges)  # normalized to image saturation
        hi_count, _ = np.histogram(im_to_show2 / b_max[2], bins=hi_edges)  # normalized to b_max[2]
        sot_x = hi_edges[:-1] + 0.005
        sot_y = hi_count[::-1].cumsum()[::-1]

        # res = plt.hist(x, bins) -> res[0] 1D-arr of counts; res[1] 1D-arr of bin limits;
        # Data outside the bins (> last limit or < first limit) is not counted.
        # The first intervals includes the first limit but not the second, the last int includes both limits.
        # plt.hist trace immediately; np.histogram only calculates (same as above) for only one series
        # the plt version has options to control the style and can plot several series.
        # the np version always returns 2 1D-arrays -> can be called by: counts, edges = np.histogram(...
        # bins = number  and  bins = 'auto' are also possible
        # "pandas" and "seaborn" have also histogram plotting

        """
        plot(centerS./ SmPeak,surfS,'.');
        ylim([0 1e3]);
        set(figure(140),'name','Surface over threshold');
        set(figure(140),'position',[2*scrsz(3)/4 scrsz(4)/2 scrsz(3)/4 scrsz(4)/2-menB]); 
        % position vector: left bottom width heigth
        ylabel('surface over threshold in square pixels');
        xlabel('Threshold 16bit-max = 1');
        
        function GauSurf = GauSurfFct(T, F, w)
        % Gives the surface with F_local > Thr for a cricular Gaussian beam.
        % w = 1/e^2 waist radius of the beam
        GauSurf = zeros(size(T));
        GauSurf(T < F) = pi/2*w.^2 .* log(F./T(T < F));
        """
        pass

    def get_step(self, minstep=2):
        print("get_step called")

        global step
        global frame_width

        if np.any(np.isnan(roiS[0])):  # check if ROI is OK
            self.LbOffSet.setText("The ROI is not defined")
            return
        roi_li = np.round(roiS[0]).astype(int)  # extract good roi line and convert to int
        roi_li[:2] -= crop_ulx  # switch to relative coords (with respect to crop)
        roi_li[2:] -= crop_uly
        frame_width = np.min([roi_li[0], imStack.shape[1] - roi_li[1],
                              roi_li[2], imStack.shape[0] - roi_li[3]])
        if ((self.cob_use_mean.currentText() == "Use mean from ROI"  # is there a reasonable frame ?
             and self.cob_roi.currentText() == "Outside ROI") or
                (self.cob_use_mean.currentText() == "Use A<sub>eff</sub>-curve fit")):
            step = int(frame_width / (self.SbSteps.value() - 1))
            if step < minstep:  # set minimum step to 2
                step = minstep
        else:  # no reasonable frame present
            step = minstep
        return

    def plot2_1D(self):
        print("plot2_1D called")

        point_alpha = 0.4
        point_size = 10
        # print("plot2_1D called")
        # Clear the figure from earlier uses
        self.figurePlt2.clear()
        # prepare the axis
        axPlt2 = self.figurePlt2.add_subplot(111)

        # self.Cob1Ddisp.addItems(["Aeff curves", "Cross sections", "Surface over threshold (SOT)"])
        if self.Cob1Ddisp.currentText() == "Cross sections":
            # plot normalized cross sections of the corrected image
            # vertical cross section
            c_section = im_to_show2[:, int(round(b_max[0]))] / b_max[2]
            axPlt2.plot(range(im_to_show2.shape[0]), c_section, '*', label="vertical",
                        color="green", alpha=point_alpha, markeredgecolor="none", ms=point_size)
            # horizontal cross section
            c_section = im_to_show2[int(round(b_max[1])), :] / b_max[2]
            axPlt2.plot(range(im_to_show2.shape[1]), c_section, '*', label="horizontal",
                        color="red", alpha=point_alpha, markeredgecolor="none", ms=point_size)
            # horizontal line for used maximum
            axPlt2.plot([0, max(im_to_show2.shape)], [1, 1], ':', color="gray")  # this line
            # helps checking that the max smoothing worked

            # plot model data: prepare
            col_idx = 0  # round Gaussian
            paras_ro = [b_fits[0][col_idx]["max"], b_fits[0][col_idx]["x_0"],
                        b_fits[0][col_idx]["y_0"], b_fits[0][col_idx]["w1"]]
            col_idx = 2  # elliptic Gaussian
            paras_el = [b_fits[0][col_idx]["max"], b_fits[0][col_idx]["x_0"],
                        b_fits[0][col_idx]["y_0"], b_fits[0][col_idx]["w1"],
                        b_fits[0][col_idx]["w2"], b_fits[0][col_idx]["angle"]]

            # plot model data: vertical cross sections
            YY = np.array(range(im_to_show2.shape[0]))  # the line numbers
            XX = np.ones(len(YY)) * round(b_max[0])  # the corresponding col no
            axPlt2.plot(YY, gauss2D_cst_offs.f(XX, YY, paras_ro),
                        '--g', label="round G. vert.")
            axPlt2.plot(YY, gaussEll2D_cst_offs.f(XX, YY, paras_el),
                        '--', color="cyan", label="ell. G. vert.")

            # plot model data: horizontal cross sections
            XX = np.array(range(im_to_show2.shape[1]))  # the col numbers
            YY = np.ones(len(XX)) * round(b_max[1])  # the corresponding line no
            axPlt2.plot(XX, gauss2D_cst_offs.f(XX, YY, paras_ro),
                        '-.', color="red", label="round G. horiz.")
            axPlt2.plot(XX, gaussEll2D_cst_offs.f(XX, YY, paras_el),
                        '-.', color="blueviolet", label="ell. G. horiz.")

            axPlt2.set_ylabel('Normalized local energy (by max used for $A_{eff}$)')
            axPlt2.set_xlabel('Position (x and y) in pixels')

        elif self.Cob1Ddisp.currentText() == "Surface over threshold (SOT)":
            axPlt2.plot(sot_x, sot_y, '*', label="measurement",
                        color="blue", alpha=point_alpha, markeredgecolor="none", ms=point_size)
            axPlt2.set_ylabel('Surface over threshold (SOT) in pixels')
            axPlt2.set_xlabel('Normalized threshold (by max used for $A_{eff}$)')

            col_idx = 1  # SOT of round Gaussian
            # axPlt2.plot(sot_x[sot_x > b_fits[0][col_idx]["lim"]],  # to mark data used for fit
            #             sot_y[sot_x > b_fits[0][col_idx]["lim"]],
            #             'x', ms=5, mfc='none', color="blue", label="used for fit")
            paras = [b_fits[0][col_idx]["max"], b_fits[0][col_idx]["w1"]]
            mod_y = sot_r_gau.f(sot_x[sot_x > 0], paras)
            axPlt2.plot(sot_x[sot_x > 0], mod_y, '-', color="black", label="Gaussian SOT")

            first_used_sot = sot_y[sot_x > b_fits[0][col_idx]["lim"]][0]
            axPlt2.set_ylim([- 0.05 * first_used_sot, 3 * first_used_sot])
            axPlt2.legend()

        elif self.Cob1Ddisp.currentText() == "A<sub>eff</sub> curves":
            # AeffCurveBasic(self, image, PixMax=-100, points=5, step=5)
            # return AeffS, ImSizeS

            # # The curve before any correction using Pixmax - but later the plot gets scaled on the corrected curve
            # im_idx = good_idx[self.SlNum2.value() - 1]  # read present slider value
            # Aeff_vec, ImSize_vec = self.AeffCurveBasic(imStack[:, :, im_idx], step=step)
            # axPlt2.plot(ImSize_vec/1000, Aeff_vec, '--', color="grey", label="no offset")
            # # removed because autocrop was not taken into account

            # the curve after offset correction using the b_max for the max
            Aeff_vec, ImSize_vec = self.AeffCurveBasic(im_to_show2, PixMax=b_max[2], step=step)
            Aeff_corr_max = Aeff_vec.max()
            Aeff_corr_min = Aeff_vec.min()
            axPlt2.plot(ImSize_vec/1000, Aeff_vec, '.r', label="with offset")

            # vertical line for the limit of the black frame (roi)
            NbPoint = self.sbNbPoint.value()
            imSizeBf = (im_to_show2.shape[0] - 2 * frame_width) * (im_to_show2.shape[1] - 2 * frame_width)
            axPlt2.plot([imSizeBf/1000, imSizeBf/1000],
                        [0.9 * Aeff_corr_min, 1.1 * Aeff_corr_max], '-.', color='blue', label='dark frame lim')
            axPlt2.plot(ImSize_vec[-NbPoint:]/1000, Aeff_vec[-NbPoint:], '.b', label="used for fit")

            # axPlt2.set_ylim([0.9 * Aeff_corr_min, 1.8 * Aeff_corr_max])
            axPlt2.set_ylabel('$A_{eff}$ of cropped image (pixels$^2$)')
            axPlt2.set_xlabel('Pixels in cropped image (kPix)')

        axPlt2.legend()
        # refresh canvas
        self.canvasPlt2.draw()

    def on_inside_chg(self):
        print("on_inside_chg called")

        if self.cob_roi.currentText() == "Inside ROI":
            self.cb_no_auto_crop.setChecked(True)
            self.cb_no_auto_crop.setEnabled(False)
        else:
            self.cb_no_auto_crop.setEnabled(True)

    def display_down2(self):
        print('display_down2 called')
        if self.SlNum2.value() > 1:
            self.SlNum2.setValue(self.SlNum2.value() - 1)

    def display_up2(self):
        print('display_up2 called')
        if self.SlNum2.value() < imStack.shape[2]:
            self.SlNum2.setValue(self.SlNum2.value() + 1)

    def display2(self):
        """Is called by
        self.SlNum2.valueChanged.connect(self.display2)
        and many others to display the present image of the stack
        """
        print('display2 called')
        measurement = im_to_show2 / b_max[2]
        # self.Cob2Ddisp.addItems(["Measurement", "Round G. model", "Round G. residual (measurement - model)",
        #                          "Elliptic G. model", "Elliptic G. residual (measurement - model)"])
        if self.Cob2Ddisp.currentText() == "Measurement":
            self.show_image2(im=measurement)

        elif self.Cob2Ddisp.currentText() == "Round G. model":
            # calculate model data (slow version)
            XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
            try:
                col_idx = 0
                paras = [b_fits[0][col_idx]["max"], b_fits[0][col_idx]["x_0"],
                         b_fits[0][col_idx]["y_0"], b_fits[0][col_idx]["w1"]]
                mod_z = gauss2D_cst_offs.f(XX, YY, paras)
                self.show_image2(im=mod_z, vmi2=measurement.min(), vma2=measurement.max())
            except:
                pass

        elif self.Cob2Ddisp.currentText() == "Round G. residual (measurement - model)":
            # calculate model data (slow version)
            ti_mod = ti.time()
            XX, YY = np.meshgrid(range(im_to_show2.shape[1]), range(im_to_show2.shape[0]))
            # try:
            col_idx = 0
            paras = [b_fits[0][col_idx]["max"], b_fits[0][col_idx]["x_0"],
                     b_fits[0][col_idx]["y_0"], b_fits[0][col_idx]["w1"]]
            mod_z = gauss2D_cst_offs.f(XX, YY, paras)
            resi = mod_z - measurement
            max_resi = np.abs(resi).max()
            self.show_image2(im=resi, cmap="seismic", vmi2=-max_resi, vma2=max_resi)

    def show_image2(self, im=None, cmap=None, vmi2=None, vma2=None):
        print("show_image2 called")

        global axIM2

        if im is None:
            return
        if cmap is None:
            cmap = self.CoboCmap2.currentText()
        if vmi2 is None:
            if self.CbUseVfixed.isChecked():  # color dynamics is defined on tab 1 but applied also to tab 2
                vmi2 = self.SbVmin.value()
            else:
                vmi2 = im.min()
        if vma2 is None:
            if self.CbUseVfixed.isChecked():  # color dynamics is defined on tab 1 but applied also to tab 2
                vma2 = self.SbVmax.value()
            else:
                vma2 = im.max()

        # Clear the figure from earlier uses
        self.figureIM2.clear()
        # prepare the axis
        axIM2 = self.figureIM2.add_subplot(111)
        axIM2.axis("off")

        # mark masked pixels if wished
        im_modified = im.copy()  # modify a copy of im
        if self.cbMaskShow.isChecked():
            # calculate the values of the color switch limits and the
            # replacement values (for the masked pixels to become visible)
            MiniVal = np.min(im)
            MaxiVal = np.max(im)
            infernoLimits = [0.35, 0.82]
            infernoColVals = [0.66, 1]
            # change the values (colors) of the pixels in the list
            for count in range(len(masked)):
                try:
                    if (masked[count][1] - crop_uly >= 0) and (masked[count][0] - crop_ulx >= 0):  # no neg indices
                        IsVal = im[masked[count][1] - crop_uly, masked[count][0] - crop_ulx]
                        if infernoLimits[0] < IsVal / (MaxiVal - MiniVal) < infernoLimits[1]:  # light color
                            putVal = infernoColVals[1] * (MaxiVal - MiniVal) + MiniVal
                        else:
                            putVal = infernoColVals[0] * (MaxiVal - MiniVal) + MiniVal
                        im_modified[masked[count][1] - crop_uly, masked[count][0] - crop_ulx] = putVal
                except IndexError:
                    pass  # does not crash if pixels outside the present crop are part of the mask

        cax = axIM2.imshow(im_modified, cmap=cmap, vmin=vmi2, vmax=vma2)
        # common color maps: jet, gray, hot, hsv, inferno, gist_ncar
        self.figureIM2.colorbar(cax, orientation='horizontal', fraction=0.1)
        self.figureIM2.tight_layout()

        self.UpdateLinesIm2(im=im_modified)

        # refresh canvas
        self.canvasIM2.draw()


    def UpdateLinesIm2(self, im=None):
        print("UpdateLinesIm2 called")
        global axIM2

        # delete lines on axis (if there were any)  possibly for the cross-sections
        try:
            axIM2.lines = []
        except NameError:  # just a guess (that seems to work)
            pass

        if self.CbShowBMax.isChecked():
            axIM2.plot(b_max[0], b_max[1], 'o', color='w', markerfacecolor='none', linewidth=2)

        maxX = im.shape[1]-1
        maxY = im.shape[0]-1
        if not(np.any(np.isnan(roiS[0, :]))):
            roi_li = self.get_local_roi_line()  # xfrom, xto, yfrom, yto
            col = 'r'
            ulx = max([roi_li[0], -1])
            lrx = min([roi_li[1], maxX + 1])
            uly = max([roi_li[2], -1])
            lry = min([roi_li[3], maxY + 1])
            axIM2.plot([ulx, lrx, lrx, ulx, ulx], [uly, uly, lry, lry, uly],
                       color=col, linestyle='--', linewidth=1)
        return


    def make_result_dict(self):
        print('make_result_dict called')
        global results  # dict with the results of the current stack defined by good_idx

        # Prepare the variables for the basic calculations results
        # The keys will be used as headers of the Excel sheet and the cobos fpr plotting the histograms
        results = {}  # reset from earlier uses
        res_imno = np.ones(len(good_idx)) * np.nan
        res_aeff = np.ones(len(good_idx)) * np.nan
        res_slope = np.ones(len(good_idx)) * np.nan
        res_straightn = np.ones(len(good_idx)) * np.nan
        res_max = np.ones((len(good_idx), 3)) * np.nan  # takes x, y, z
        res_backg = np.ones(len(good_idx)) * np.nan
        res_energy = np.ones(len(good_idx)) * np.nan

        # Prepare the variables for the fit results
        if ((self.TbFit2.cellWidget(0, 0).currentText() == "Yes") or
                (self.TbFit2.cellWidget(0, 1).currentText() == "Yes") or
                (self.TbFit2.cellWidget(0, 2).currentText() == "Yes")):
            res_b_fit = np.array(len(good_idx) * [[None, None, None]])  # array with dicts as elements

        # build the values (mostly arrays) of the dict
        self.tabs.setCurrentIndex(1)
        main.statBar.showMessage("Analysing stack")
        main.progBar.setMinimum(1)  # if it's not done here, it has to be done at different places later on
        main.progBar.setMaximum(len(good_idx))
        QApplication.processEvents()
        ti_stack_0 = ti.time()
        for idx in range(len(good_idx)):
            self.SlNum2.setValue(idx + 1)  # triggers the calculations using the settings defined on tab2

            res_imno[idx] = good_idx[idx] + 1  # image number in initial stack (wcf-file)
            res_aeff[idx] = Aeffg
            res_slope[idx] = slope
            res_straightn[idx] = np.nan
            res_max[idx] = b_max  # takes x, y, z
            res_backg[idx] = backg
            res_energy[idx] = energy

            if self.TbFit2.cellWidget(0, 0).currentText() == "Yes":  # Round Gauss fit done
                res_b_fit[idx, 0] = b_fits[0][0]
                res_b_fit[idx, 0]["max"] *= b_max[2]  # remove normalization
            if self.TbFit2.cellWidget(0, 1).currentText() == "Yes":  # SOT fit done
                res_b_fit[idx, 1] = b_fits[0][1]
                res_b_fit[idx, 1]["max"] *= b_max[2]  # remove normalization
            if self.TbFit2.cellWidget(0, 2).currentText() == "Yes":  # Elliptic Gauss fit done
                res_b_fit[idx, 2] = b_fits[0][2]
                res_b_fit[idx, 2]["max"] *= b_max[2]  # remove normalization

            main.progBar.setValue(idx + 1)
            QApplication.processEvents()

        # Fill the results-dict using the values above and define the keys at the same time
        # The keys will be used as headers of the Excel sheet and the cobos fpr plotting the histograms
        results["Image number"] = res_imno
        results["Effective surface (px²)"] = res_aeff
        results["Aeff slope"] = res_slope
        results["Aeff straightness"] = res_straightn
        results["Max value (GL)"] = res_max[:, 2]
        results["Background (GL)"] = res_backg
        results["Energy (GL)"] = res_energy
        results["Max pos X (px)"] = res_max[:, 0]
        results["Max pos Y (px)"] = res_max[:, 1]
        # "status" "GOF" "w1" "max" "x_0" "y_0" "w2" "angle" (useful entries of fit parameter dicts)
        if self.TbFit2.cellWidget(0, 0).currentText() == "Yes":  # Round Gauss fit done
            results["RG beam radius w1 (px)"] = np.array([res_b_fit[row, 0]["w1"] for row in range(len(good_idx))])
            results["RG Max value (GL)"] = np.array([res_b_fit[row, 0]["max"] for row in range(len(good_idx))])
            results["RG Max pos X (px)"] = np.array([res_b_fit[row, 0]["x_0"] for row in range(len(good_idx))])
            results["RG Max pos Y (px)"] = np.array([res_b_fit[row, 0]["y_0"] for row in range(len(good_idx))])
            results["RG status"] = list([res_b_fit[row, 0]["status"] for row in range(len(good_idx))])
            results["RG GOF"] = np.array([res_b_fit[row, 0]["GOF"] for row in range(len(good_idx))])
        if self.TbFit2.cellWidget(0, 1).currentText() == "Yes":  # SOT fit done
            results["Gsot beam radius w1 (px)"] = np.array([res_b_fit[row, 1]["w1"] for row in range(len(good_idx))])
            results["Gsot Max value (GL)"] = np.array([res_b_fit[row, 1]["max"] for row in range(len(good_idx))])
            results["Gsot status"] = list([res_b_fit[row, 1]["status"] for row in range(len(good_idx))])
            results["Gsot GOF"] = np.array([res_b_fit[row, 1]["GOF"] for row in range(len(good_idx))])
        if self.TbFit2.cellWidget(0, 2).currentText() == "Yes":  # Elliptic Gauss fit done
            results["EllG beam radius w1 (px)"] = np.array([res_b_fit[row, 2]["w1"] for row in range(len(good_idx))])
            results["EllG beam radius w2 (px)"] = np.array([res_b_fit[row, 2]["w1"] for row in range(len(good_idx))])
            results["EllG long axis angle (°)"] = np.array([res_b_fit[row, 2]["angle"] for row in range(len(good_idx))])
            results["EllG Max value (GL)"] = np.array([res_b_fit[row, 2]["max"] for row in range(len(good_idx))])
            results["EllG Max pos X (px)"] = np.array([res_b_fit[row, 2]["x_0"] for row in range(len(good_idx))])
            results["EllG Max pos Y (px)"] = np.array([res_b_fit[row, 2]["y_0"] for row in range(len(good_idx))])
            results["EllG status"] = list([res_b_fit[row, 2]["status"] for row in range(len(good_idx))])
            results["EllG GOF"] = np.array([res_b_fit[row, 2]["GOF"] for row in range(len(good_idx))])
        # probably there is a better way to do this

        # populate combo boxes for display of histograms and x-y-scatter
        CobPl_li = [self.CobPl31, self.CobPl32, self.CobPl33, self.CobPl34]
        key_li = list(results.keys())
        print("key_li = ", key_li)
        for this_cobo_num, this_cobo in enumerate(CobPl_li):
            this_cobo.clear()
            this_cobo.addItems(list(results.keys()))
            this_cobo.setCurrentIndex(this_cobo_num + 1)

        self.show_res_statistics_table(ti_stack_0)

    def show_res_statistics_table(self, ti_stack_0):
        print("show_res_statistics_table called")

        global shown_li  # list of the shown columns (for which mean etc can be calculated
        global good_filt_idx

        # Populate summary table with all headers except those containing "status"
        # TODO: possibly make formatting dependent on data to display (not all with "{:.2f}")
        # TODO: make a list of column header that do not need to appear in the table
        # TODO: make decimal point justification
        self.tb_res3.clearContents()
        self.tb_res3.setColumnCount(0)

        rem_this = self.range_str_conv(self.EdFiFilt.text())
        rem_this = list(np.array(rem_this) - 1)
        # print("good_idx: ", good_idx)
        # print("rem_this: ", rem_this)
        good_filt_idx = [x for x in range(len(good_idx)) if (good_idx[x] not in rem_this)]
        # print("filtered good_idx:", np.array(good_idx)[good_filt_idx])
        # ok this works, but I did not succeed with numpy array manipulation

        key_li = list(results.keys())
        shown_li = []
        for col_idx in range(len(key_li)):  # the names can be reused but new instances need to be created
            # print("col name = ", key_li[col_idx])
            # print("list test: ", type(results[key_li[col_idx]]) is list)
            if (type(results[key_li[col_idx]]) is list) or (key_li[col_idx] == "Image number"):
                # print("jump")
                continue  # do not use text data which is stored in lists, nor the image numbers

            print("nan test", np.any(np.isnan(results[key_li[col_idx]])))
            if np.any(np.isnan(results[key_li[col_idx]])):
                # print("jump")
                continue  # do not use columns that contain nan

            shown_li.append(col_idx)  # keep track of the displayed columns
            # populate combo boxes of tab 4 (for z-evolution)
            self.CobPl1.addItem(key_li[col_idx])
            self.CobPl2.addItem(key_li[col_idx])

            # print(" good col name = ", key_li[col_idx])
            # print("old ColCount = ", self.tb_res3.columnCount())
            self.tb_res3.setColumnCount(self.tb_res3.columnCount() + 1)  # add a column
            # print("new ColCount = ", self.tb_res3.columnCount())
            header_item = QTableWidgetItem(key_li[col_idx])
            last_col_idx = self.tb_res3.columnCount() - 1
            self.tb_res3.setHorizontalHeaderItem(last_col_idx, header_item)

            cob_YesNo = QComboBox()
            cob_YesNo.setEditable(True)  # if not it is impossible to align on the center
            cob_YesNo.lineEdit().setAlignment(Qt.AlignCenter)
            cob_YesNo.lineEdit().setReadOnly(True)  # works, but the backg is white
            cob_YesNo.addItems(['Yes', 'No'])
            cob_YesNo.setCurrentText("Yes")
            self.tb_res3.setCellWidget(0, last_col_idx, cob_YesNo)

            the_mean = results[key_li[col_idx]][good_filt_idx].mean()
            LbResMean = QLabel('{:.2f}'.format(the_mean))
            LbResMean.setAlignment(Qt.AlignRight)
            self.tb_res3.setCellWidget(1, last_col_idx, LbResMean)

            abs_std_err_of_mean = (sem(results[key_li[col_idx]][good_filt_idx])
                                   * t.ppf((1 + confidence) / 2,
                                           df=(len(results[key_li[col_idx]][good_filt_idx]) - 1)))
            # PPF (Percent point function) is the inverse of the CDF. That is, PPF gives the value
            # of the variate for which the cumulative probability has the given
            # value: One edge + main part = (1-c)/2 + c = (1+c)/2; df = deg. of freedom
            LbResUnc = QLabel('{:.2f}'.format(abs_std_err_of_mean))
            LbResUnc.setAlignment(Qt.AlignRight)
            self.tb_res3.setCellWidget(3, last_col_idx, LbResUnc)

            rel_std_err_of_mean = abs_std_err_of_mean / the_mean
            LbResRelErr = QLabel('{:.1%}'.format(rel_std_err_of_mean))
            LbResRelErr.setAlignment(Qt.AlignRight)
            self.tb_res3.setCellWidget(2, last_col_idx, LbResRelErr)

            LbResMin = QLabel('{:.2f}'.format(results[key_li[col_idx]][good_filt_idx].min()))
            LbResMin.setAlignment(Qt.AlignRight)
            self.tb_res3.setCellWidget(4, last_col_idx, LbResMin)

            LbResMax = QLabel('{:.2f}'.format(results[key_li[col_idx]][good_filt_idx].max()))
            LbResMax.setAlignment(Qt.AlignRight)
            self.tb_res3.setCellWidget(5, last_col_idx, LbResMax)

            LbResMed = QLabel('{:.2f}'.format(np.median(results[key_li[col_idx]][good_filt_idx])))
            LbResMed.setAlignment(Qt.AlignRight)
            self.tb_res3.setCellWidget(6, last_col_idx, LbResMed)
        # end of table writing loop
        self.tb_res3.resizeColumnsToContents()
        self.tb_res3.resizeRowsToContents()
        self.tb_res3.setStyleSheet(""" 
        * {padding-right: 5px; 
           selection-background-color: lightgray} """)
        # Avoid that the right symbol touches the grid line: padding-right: 5px;
        # works but lifts everything a bit upwards

        self.make_comment_sheet_for_export()

        # main.statBar.showMessage("Finished analyzing", 3000)
        main.statBar.showMessage(
            "Finished analyzing: processing took {:.2f} seconds".format(ti.time()-ti_stack_0),
            5000)
        main.progBar.setValue(1)
        self.tabs.setCurrentIndex(2)
        self.plot3_1D()


    def plot3_1D(self):
        print("plot3_1D called")

        global axPlt_li
        global CobPl_li
        # print('plot3_1D called')
        # clear figure
        self.figurePlt3.clear()
        axPlt31 = self.figurePlt3.add_subplot(221)
        axPlt32 = self.figurePlt3.add_subplot(222)
        axPlt33 = self.figurePlt3.add_subplot(223)
        axPlt34 = self.figurePlt3.add_subplot(224)
        axPlt_li = [axPlt31, axPlt32, axPlt33, axPlt34]
        col_li = ['green', 'grey', 'red', 'blue']
        # label and x_axis label  will be combo box text
        CobPl_li = [self.CobPl31, self.CobPl32, self.CobPl33, self.CobPl34]

        for (this_ax, this_col, this_cobo) in zip(axPlt_li, col_li, CobPl_li):
            curr_text = this_cobo.currentText()
            if "POS" in curr_text.upper():  # make a depointing plot
                key_li = list(results.keys())
                idx = key_li.index(curr_text)  # could also use this_cobo.currentIndex()
                if " X" in curr_text.upper():
                    x_positions = results[curr_text]
                    y_positions = results[key_li[idx + 1]]
                else:
                    x_positions = results[key_li[idx - 1]]
                    y_positions = results[curr_text]

                this_ax.plot(x_positions, y_positions, '.',
                             color=this_col,
                             label=this_cobo.currentText())
                # this_ax.legend(loc=(x_positions.max()+1, y_positions.min()))  # Does not work with tight_layout
                this_ax.set_title("Position of maximum (depointing)")
                this_ax.set_xlabel("Horizontal position (px)")
                this_ax.set_ylabel("Vertical position (px)")
                this_ax.set_aspect("equal")
            else:  # Make a histogram
                if not type(results[this_cobo.currentText()]) is np.ndarray:
                    continue  # do not do what follows

                if np.any(np.isnan(results[this_cobo.currentText()])):
                    continue  # do not do what follows

                _, bins, _ = this_ax.hist(results[this_cobo.currentText()],
                                          bins="auto",  # bins="auto" , bins=15
                                          fill=True,
                                          linewidth=2,
                                          edgecolor=this_col,
                                          facecolor='none',
                                          label="All good images (tab1)")
                this_ax.hist(results[this_cobo.currentText()][good_filt_idx],
                             bins=bins,
                             color=this_col,
                             label="Remaining images (tab3 filter)")

                this_ax.legend()
                this_ax.set_xlabel(this_cobo.currentText())
                this_ax.set_ylabel("Counts")
                this_ax.set_aspect("auto")

        # refresh canvas
        self.figurePlt3.tight_layout()
        self.canvasPlt3.draw()

    def set_checkb_state4(self):
        if self.sender() is self.CobPl1:
            if self.CobPl1.currentText() == "Effective surface (px²)":
                self.Cb4Pl1.setEnabled(True)
            else:
                self.Cb4Pl1.setEnabled(False)
        if self.sender() is self.CobPl2:
            if self.CobPl2.currentText() == "Effective surface (px²)":
                self.Cb4Pl2.setEnabled(True)
            else:
                self.Cb4Pl2.setEnabled(False)

    def make_z_curves(self):

        # generate list of filenames for files of (i) same type as the initial file, (ii) same file size,
        # (iii) containing information on the z position.

        def extract_z_from_filename(fi):
            z = -20e20
            # extract the position floating point number from the string fi

            zz = re.search(r'[+-]?[\d]', fi).group(0)

            # print(type(zz)) -> class 'str'
            z = float(zz)
            return z

        fi_size = os.path.getsize(f_name)
        list_files = []
        list_z_positions = []
        for filename in os.listdir(dirname):
            thisBaseName = filename.split('/')[-1]  # usually does nothing if not in subdirectory
            thisFile_type = thisBaseName.split('.')[-1]
            if thisFile_type != file_type:  # wrong file type
                continue
            thisFileSize = os.path.getsize(filename)
            if thisFileSize != fi_size:  # wrong file size
                continue
            z_pos = extract_z_from_filename(thisBaseName)
            if z_pos == -20e20:  # or error in position extraction
                continue
            list_files.append(thisBaseName)
            list_z_positions.append(z_pos)

        #file_list = QtWidgets.QFileDialog.getOpenFileName(parent=self, directory=dirname)
        for count in range(len(list_files)):
            print("File: ", list_files[count], '; z-val : ', list_z_positions[count])


        """       
        modify the global f_name
        call load_stack
        call make_results_dict
        get wanted value like this:
            ourMean1 = results[self.CobPl1.currentText()][good_filt_idx].mean()
            print("ourMean1 = ", ourMean1)
            ourStdErr1 = sem(results[self.CobPl1.currentText()][good_filt_idx])
            print("ourStdErr1 = ", ourStdErr1)
        loop
        """

####################################################
if __name__ == "__main__":
    app = QApplication(sys.argv)  # std : create and launch Application
    main = MainWindow()  # self.show() is the last line of MainWindow.__init__()
    sys.exit(app.exec_())  # std : exits python when the app finishes
