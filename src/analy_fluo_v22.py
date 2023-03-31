"""
@author: Wagner
"""

import glob
import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget,
                             QGridLayout, QVBoxLayout, QHBoxLayout, QSlider,
                             QTabWidget, QButtonGroup, QRadioButton,
                             QPushButton, QCheckBox, QLabel, QFileDialog, QLineEdit,
                             QSpinBox, QDoubleSpinBox, QProgressBar, QComboBox, QTableWidget)
from PyQt5.QtCore import Qt  # for the focus from keyboard

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure  # better do not use pyplot (but it's possible)
import matplotlib.pyplot as plt
from scipy import ndimage as ndi
from scipy import stats as st

import numpy as np
import datetime as ti
import openpyxl as xl  # for writing excel files with multi line header
import csv  # for reading more easily the text file with the mask data

roi_3_width = 4  # pixels
hole_start_ringmean = -10
hole_start_maxpos = -10

class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)

        # Get screen size.
        scr_si = QApplication.desktop().size()
        scr_si_x = scr_si.width()
        scr_si_y = scr_si.height()

        # Define window size, position and title.
        self.left = 5
        self.top = 40
        self.width = scr_si_x - 50
        self.height = scr_si_y - self.top - 50
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.title = "analy_fluo_v22"
        self.setWindowTitle(self.title)

        # Instantiate the tab component and set as Central.
        self.tw = MyTableWidget(self)
        self.setCentralWidget(self.tw)

        self.statBar = self.statusBar()
        self.progBar = QProgressBar()
        self.statBar.addPermanentWidget(self.progBar)  # Layout is automatic
        self.statBar.showMessage('Ready')

        self.show()


####################################################
class MyTableWidget(QWidget):
    def __init__(self, parent):
        print('__init__ called')
        super().__init__(parent)

        # Initialize tab screen
        self.tabs = QTabWidget()
        # self.tabs.currentChanged.connect(self.plot_graphs) --> at the end of init as it calls
        # plot_graphs before the necessary objects exist
        self.tabs.resize(300, 200)
        # self.tabs.currentChanged.connect(self.onTabChange)

        # Add pages (tab1, tab2...) to tabWidget (tabs)
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, "Inspect stack and prepare analysis")

        self.tab2 = QWidget()
        self.tabs.addTab(self.tab2, "Analyse and export")

        def add_widgets_to_tab1():  # Create widgets of 1st tab : inspect
            print('add_widgets_to_tab1 called')
            # For loading the image file
            self.LbFolderName = QLabel('Click "Load File" to start')  # label (actif)
            self.BtLoadFile1 = QPushButton('1. select file from stack')
            self.BtLoadFile1.clicked.connect(self.Load_File1)

            # For characterizing the stack
            self.LbBaseName = QLabel('Basename: ')  # label (actif)
            self.LbFiStart = QLabel('Stack from no: ')  # label (actif)
            self.EdFiStart = QLineEdit(' ')  # label (actif)
            self.EdFiStart.returnPressed.connect(self.set_stack_limits)
            self.LbFiEnd = QLabel('to no: ')  # label (actif)
            self.EdFiEnd = QLineEdit(' ')  # label (actif)
            self.EdFiEnd.returnPressed.connect(self.set_stack_limits)
            self.LbImNo = QLabel("N.A.")  # label (actif)
            self.LbInfo = QLabel("For information")  # label (actif)

            # For selecting a file in the stack
            self.BtNumDwn = QPushButton('<-')
            self.BtNumDwn.clicked.connect(self.display_down)
            self.SlNum = QSlider(Qt.Horizontal, self)
            self.SlNum.valueChanged.connect(self.display)
            self.BtNumUp = QPushButton('->')
            self.BtNumUp.clicked.connect(self.display_up)

            # For plotting the image
            self.figureIM1 = Figure()
            self.canvasIM1 = FigureCanvas(self.figureIM1)
            self.toolbarIM1 = NavigationToolbar(self.canvasIM1, self)  # Canvas widget and a parent
            # Connect a function that gets the key that is pressed
            self.canvasIM1.mpl_connect('key_press_event',self.GetKeyIm1)
            # Configure the canvas widget to process keyboard events
            self.canvasIM1.setFocusPolicy(Qt.StrongFocus)  # listen to keyboard too!
            # Connect a function that gets the mouse position on every move
            self.canvasIM1.mpl_connect('motion_notify_event',self.GetMouseIm1)

            # For cropping
            self.CbCrop = QCheckBox("apply crop margins")
            self.CbCrop.stateChanged.connect(self.crop_or_uncrop)
            self.LbUpper = QLabel("pixels")
            self.SbUpper = QSpinBox()
            self.SbLeft = QSpinBox()
            self.SbRight = QSpinBox()
            self.SbLower = QSpinBox()
            self.SbUpper.setRange(0, 1000)
            self.SbLeft.setRange(0, 1000)
            self.SbRight.setRange(0, 1000)
            self.SbLower.setRange(0, 1000)
            self.SbUpper.setSingleStep(10)
            self.SbLeft.setSingleStep(10)
            self.SbRight.setSingleStep(10)
            self.SbLower.setSingleStep(10)
            self.SbUpper.setValue(400)
            self.SbLeft.setValue(300)
            self.SbRight.setValue(700)
            self.SbLower.setValue(350)
            self.SbUpper.valueChanged.connect(self.set_crop_margins)
            self.SbLeft.valueChanged.connect(self.set_crop_margins)
            self.SbRight.valueChanged.connect(self.set_crop_margins)
            self.SbLower.valueChanged.connect(self.set_crop_margins)

            # Radiobutton group for median filter
            rdBtLayoutMedFi = QHBoxLayout()  # layout for the radio button widget
            self.rdBtwidgetMedFi = QWidget(self)  # radio button widget :: Place this in the main Layout
            self.rdBtwidgetMedFi.setLayout(rdBtLayoutMedFi)
            self.rdbtgroupMedFi = QButtonGroup(self.rdBtwidgetMedFi)
            rdbtMedFiNo = QRadioButton("No Median filter")
            self.rdbtgroupMedFi.addButton(rdbtMedFiNo, 0)  # 0 is the id in the group
            rdBtLayoutMedFi.addWidget(rdbtMedFiNo)
            rdbtMedFiNo.setChecked(True)  # check one button on creation
            rdbtMedFiPrev = QRadioButton("Preview (1 image)")
            self.rdbtgroupMedFi.addButton(rdbtMedFiPrev, 1)
            rdBtLayoutMedFi.addWidget(rdbtMedFiPrev)
            rdbtMedFiApply = QRadioButton("Apply (stack)")
            self.rdbtgroupMedFi.addButton(rdbtMedFiApply, 2)
            rdBtLayoutMedFi.addWidget(rdbtMedFiApply)
            self.rdbtgroupMedFi.buttonClicked[int].connect(self.medFi_state_changed)
            self.LbMedFi = QLabel("pixels")
            self.SbMedFi = QSpinBox()
            self.SbMedFi.setRange(1, 10)
            self.SbMedFi.setValue(3)
            self.SbMedFi.valueChanged.connect(self.medFi_sb_changed)

            ### Set fixed color scale
            self.CbUseVfixed = QCheckBox('Freeze color scale')
            self.CbUseVfixed.stateChanged.connect(self.display)
            self.LbVmax = QLabel('Choose color scale ->')
            self.CoboCmap = QComboBox()
            self.CoboCmap.addItems(["jet", "gray", "hot", "hsv", "inferno", "gist_ncar"])
            # common color maps:     jet,   gray,   hot,   hsv,    inferno,  gist_ncar
            self.CoboCmap.currentIndexChanged.connect(self.display)
            self.LbVmin = QLabel('Set min, max:')
            self.SbVmin = QSpinBox()
            self.SbVmax = QSpinBox()
            self.SbVmin.valueChanged.connect(self.display)
            self.SbVmax.valueChanged.connect(self.display)

            # For ROI 1
            self.lbROI1 = QLabel("ROI 1 (background)")
            self.lbROI1ul = QLabel("tbd")
            self.lbROI1lr = QLabel("tbd")
            self.cbROI1redefine = QCheckBox('redefine')
            self.cbROI1redefine.clicked.connect(self.roi_def)

            # For ROI 2
            self.lbROI2 = QLabel("ROI 2 (max. flu.)")
            self.lbROI2ul = QLabel("tbd")
            self.lbROI2lr = QLabel("tbd")
            self.cbROI2redefine = QCheckBox('redefine')
            self.cbROI2redefine.clicked.connect(self.roi_def)

            # For mask definition, saving and loading
            self.lbMaskTit = QLabel("Modify mask: ")
            self.CoboMaskMode = QComboBox()
            self.CoboMaskMode.addItems(["Add pixels", "Remove pixels"])
            self.cbMaskActivateIt = QCheckBox('Do it now')
            self.cbMaskActivateIt.clicked.connect(self.roi_def)
            self.cbMaskShow = QCheckBox('Show mask')
            self.cbMaskShow.clicked.connect(self.display)
            self.cbMaskShow.setChecked(True)
            self.lbMaskInfo = QLabel("0 pixels masked")
            self.lb_mask_save = QLabel("Re-use mask (+c):")
            self.bt_mask_save = QPushButton("Save")
            self.bt_mask_save.clicked.connect(self.save_mask)
            self.bt_ask_load = QPushButton("2. Load")
            self.bt_ask_load.clicked.connect(self.load_mask)

        def add_widgets_to_tab2():  # Create widgets of 1st tab : inspect
            global mess_ring_mean, mess_glob_max
            print('add_widgets_to_tab2 called')

            self.Lb_stack_name = QLabel('Stack name: ')  # active label

            # For the task table
            self.tablines = 7
            self.LbTitle = QLabel('Define what to analyze')  # label
            self.TbTasks = QTableWidget(self.tablines, 6)  # adjust calls of fill_task_line
            self.TbTasks.setHorizontalHeaderLabels(['Select', 'Region', 'Out/In',
                                                    'Mask', 'Value', 'Plot', 'Export'])
            # self.TbTasks.setMinimumWidth(615)
            # self.TbTasks.setStyleSheet("border-image: url(fond_avec_plan.png) 0 0 0 0 stretch stretch;")
            def fill_task_line(line_no):
                """ Fills a line of the table with the widgets. line_no starts with 0 """
                # col 1-1: combo box for Region
                combo_temp = QComboBox()
                combo_temp.addItems(["Roi 1", "Roi 2", "Roi3 - Roi2", "Full frame"])
                self.TbTasks.setCellWidget(line_no, 1-1, combo_temp)
                # col 2-1: combo box for 'Out/In'
                combo_temp = QComboBox()
                combo_temp.addItems(["Outside", "Inside"])
                self.TbTasks.setCellWidget(line_no, 2-1, combo_temp)
                # col 3-1: combo box for 'Mask'
                combo_temp = QComboBox()
                combo_temp.addItems(["Ignore masked pixels", "Use all pixels"])
                self.TbTasks.setCellWidget(line_no, 3-1, combo_temp)
                # col 4-1: combo box for 'Value'
                combo_temp = QComboBox()
                combo_temp.addItems(["Mean Value", "Max Value", "Median Value", "Min Value", "Max Position"])
                self.TbTasks.setCellWidget(line_no, 4-1, combo_temp)
                # col 5-1: combo box for 'Plot'
                combo_temp = QComboBox()
                combo_temp.addItems(["Yes", "No"])
                self.TbTasks.setCellWidget(line_no, 5-1, combo_temp)
                # col 6-1: combo box for 'Export'
                combo_temp = QComboBox()
                combo_temp.addItems(["Yes", "No"])
                self.TbTasks.setCellWidget(line_no, 6-1, combo_temp)
            for line in range(self.tablines):
                fill_task_line(line)  # header displays 1
            self.TbTasks.setColumnWidth(1-1, 85)  # col no, pixels
            self.TbTasks.setColumnWidth(2-1, 75)  # col no, pixels
            self.TbTasks.setColumnWidth(3-1, 155)  # col no, pixels
            self.TbTasks.setColumnWidth(4-1, 110)  # col no, pixels
            self.TbTasks.setColumnWidth(5-1, 50)  # col no, pixels
            self.TbTasks.setColumnWidth(6-1, 55)  # col no, pixels
            self.TbTasks.setMaximumWidth(600)
            self.lb_info_2 = QLabel('tbd')  # label

            # For plotting the graphs (no calculus)
            self.figureIM2 = Figure()
            self.axIM2 = self.figureIM2.add_subplot(111)
            self.canvasIM2 = FigureCanvas(self.figureIM2)
            self.toolbarIM2 = NavigationToolbar(self.canvasIM2, self)  # Canvas widget and a parent
            self.bt_plot_now = QPushButton("1. Plot Now")
            self.bt_plot_now.clicked.connect(self.plot_graphs)
            self.bt_export_now = QPushButton("3. Export Now")
            self.bt_export_now.clicked.connect(self.export_graphs)

            # For finding the start of the fluorescence hole
            self.Lb_hole_title = QLabel("Parameters for automatic hole finding:")
            # input for method 1
            self.Lb_hole_meth_ring = QLabel("Ring mean vs. center mean ()")
            self.Lb_smoothing_intv = QLabel("Set smoothing interval (images)")
            self.Sb_smoothing_intv = QSpinBox()
            self.Sb_smoothing_intv.setRange(1, 25)
            self.Sb_smoothing_intv.setValue(5)  # set step to stay odd
            self.Sb_smoothing_intv.setSingleStep(2)
            # input for method 2
            self.Lb_hole_meth_max = QLabel("Global max vs. center max (t-test)")
            self.Lb_av_range = QLabel("Set averging interval (images)")
            self.Sb_av_range = QSpinBox()
            self.Sb_av_range.setRange(1, 31)
            self.Sb_av_range.setValue(5)  # set step to stay odd
            self.Sb_av_range.setSingleStep(2)
            self.Lb_diff_confi = QLabel("Confidence that max are diff.")
            self.Sb_diff_confi = QDoubleSpinBox()
            self.Sb_diff_confi.setRange(0, 1)
            self.Sb_diff_confi.setValue(0.99)  # set step to stay odd
            self.Sb_diff_confi.setSingleStep(0.01)
            # output
            self.bt_find_hole = QPushButton("2. Find hole")
            self.bt_find_hole.clicked.connect(self.find_hole)
            mess_ring_mean = "Did not try yet"
            self.Lb_hole_res_ring = QLabel(mess_ring_mean)
            mess_glob_max = "Did not try yet"
            self.Lb_hole_res_max = QLabel(mess_glob_max)

        add_widgets_to_tab1()
        add_widgets_to_tab2()

        def initialize_table():
            self.TbTasks.cellWidget(0, 1-1).setCurrentText("Full frame")
            self.TbTasks.cellWidget(0, 2-1).setCurrentText("Inside")
            self.TbTasks.cellWidget(0, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(0, 4-1).setCurrentText("Max Value")

            self.TbTasks.cellWidget(1, 1-1).setCurrentText("Roi 2")
            self.TbTasks.cellWidget(1, 2-1).setCurrentText("Inside")
            self.TbTasks.cellWidget(1, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(1, 4-1).setCurrentText("Max Value")

            self.TbTasks.cellWidget(2, 1-1).setCurrentText("Roi 2")
            self.TbTasks.cellWidget(2, 2-1).setCurrentText("Inside")
            self.TbTasks.cellWidget(2, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(2, 4-1).setCurrentText("Median Value")

            self.TbTasks.cellWidget(3, 1-1).setCurrentText("Roi 2")
            self.TbTasks.cellWidget(3, 2-1).setCurrentText("Inside")
            self.TbTasks.cellWidget(3, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(3, 4-1).setCurrentText("Mean Value")

            self.TbTasks.cellWidget(4, 1-1).setCurrentText("Roi3 - Roi2")
            self.TbTasks.cellWidget(4, 2-1).setCurrentText("Inside")
            self.TbTasks.cellWidget(4, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(4, 4-1).setCurrentText("Median Value")

            self.TbTasks.cellWidget(5, 1-1).setCurrentText("Roi3 - Roi2")
            self.TbTasks.cellWidget(5, 2-1).setCurrentText("Inside")
            self.TbTasks.cellWidget(5, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(5, 4-1).setCurrentText("Mean Value")

            self.TbTasks.cellWidget(6, 1-1).setCurrentText("Roi 1")
            self.TbTasks.cellWidget(6, 2-1).setCurrentText("Outside")
            self.TbTasks.cellWidget(6, 3-1).setCurrentText("Ignore masked pixels")
            self.TbTasks.cellWidget(6, 4-1).setCurrentText("Mean Value")

            """
            ["Roi 1", "Roi 2", "Roi3 - Roi2", "Full frame"]
            ["Outside", "Inside"]
            ["Ignore masked pixels", "Use all pixels"]
            ["Mean Value", "Max Value", "Median Value", "Min Value", "Max Position"]
            ["Yes", "No"]
            ["Yes", "No"]
            """
            # for row in range(4,6):
            #     for col in range(5,7):
            #         self.TbTasks.cellWidget(row, col).setCurrentText("No")

        initialize_table()

        def set_layout_of_tab1():
            print('set_layout_of_tab1 called')

            # setting the layout of the 1st tab
            grid1 = QGridLayout(self)
            grid1.setSpacing(5)  # defines the spacing between widgets
            # horizontally: end of label to edit, and vertically: edit to edit

            ### file loading
            grid1.addWidget(self.BtLoadFile1, 0, 0, 1, 2)  # (object, Zeile, Spalte, rowSpan, colSpan)
            grid1.addWidget(self.LbFolderName, 0, 2, 1, 7)

            ### Stack info : choose images here
            grid1.addWidget(self.LbBaseName, 1, 0, 1, 2)
            grid1.addWidget(self.LbFiStart, 1, 2)
            grid1.addWidget(self.EdFiStart, 1, 3)
            grid1.addWidget(self.LbFiEnd, 1, 4)
            grid1.addWidget(self.EdFiEnd, 1, 5)
            grid1.addWidget(self.LbInfo, 2, 0, 1, 6)

            ### Navigate in stack
            grid1.addWidget(self.BtNumDwn, 3, 0)
            grid1.addWidget(self.SlNum, 3, 1, 1, 4)
            grid1.addWidget(self.BtNumUp, 3, 5)

            ### Showing the present image
            grid1.addWidget(self.canvasIM1, 5, 0, 15, 6)
            grid1.addWidget(self.toolbarIM1, 21, 0, 1, 4)
            grid1.addWidget(self.LbImNo, 21, 5)

            ### Define crop margins (in pixels)
            grid1.addWidget(self.CbCrop, 2, 6)
            grid1.addWidget(self.SbUpper, 2, 7)
            grid1.addWidget(self.LbUpper, 2, 8)
            grid1.addWidget(self.SbLeft, 3, 6)
            grid1.addWidget(self.SbRight, 3, 8)
            grid1.addWidget(self.SbLower, 4, 7)

            ### Activate Median Filter
            grid1.addWidget(self.rdBtwidgetMedFi, 6, 6, 1, 4)
            grid1.addWidget(self.SbMedFi, 7, 7)
            grid1.addWidget(self.LbMedFi, 7, 8)

            ### Set fixed color scale
            grid1.addWidget(self.CbUseVfixed, 10, 6)
            grid1.addWidget(self.LbVmax, 10, 7)
            grid1.addWidget(self.CoboCmap, 10, 8)
            grid1.addWidget(self.LbVmin, 11, 6)
            grid1.addWidget(self.SbVmin, 11, 7)
            grid1.addWidget(self.SbVmax, 11, 8)

            # For ROI 1
            grid1.addWidget(self.lbROI1, 13, 6)
            grid1.addWidget(self.lbROI1ul, 13, 7)
            grid1.addWidget(self.lbROI1lr, 14, 8)
            grid1.addWidget(self.cbROI1redefine, 14, 6)

            # For ROI 2
            grid1.addWidget(self.lbROI2, 16, 6)
            grid1.addWidget(self.lbROI2ul, 16, 7)
            grid1.addWidget(self.lbROI2lr, 17, 8)
            grid1.addWidget(self.cbROI2redefine, 17, 6)

            # For mask definition and display, saving and loading
            grid1.addWidget(self.lbMaskTit, 19, 6)
            grid1.addWidget(self.CoboMaskMode, 19, 7)
            grid1.addWidget(self.cbMaskActivateIt, 19, 8)
            grid1.addWidget(self.cbMaskShow, 20, 6)
            grid1.addWidget(self.lbMaskInfo, 20, 7)
            grid1.addWidget(self.lb_mask_save, 21, 6)
            grid1.addWidget(self.bt_mask_save, 21, 7)
            grid1.addWidget(self.bt_ask_load, 21, 8)

            self.tab1.setLayout(grid1)
            # grid1.setRowStretch(2, 1)

        def set_layout_of_tab2():
            print('set_layout_of_tab2 called')

            # setting the layout of the 1st tab
            grid2 = QGridLayout(self)
            grid2.setSpacing(5)  # defines the spacing between widgets
            # horizontally: end of label to edit, and vertically: edit to edit

            # Table for the task definition
            grid2.addWidget(self.LbTitle, 0, 1)  # (object, Zeile, Spalte, rowSpan, colSpan)
            grid2.addWidget(self.TbTasks, 1, 0, 4, 6)

            # Display the image number of appearance of the fluo hole
            grid2.addWidget(self.Lb_hole_title, 6, 2, 1, 2)
            grid2.addWidget(self.Lb_hole_meth_ring, 7, 1, 1, 2)
            grid2.addWidget(self.Lb_smoothing_intv, 8, 0, 1, 2)
            grid2.addWidget(self.Sb_smoothing_intv, 8, 2)

            grid2.addWidget(self.Lb_hole_meth_max, 7, 4, 1, 2)
            grid2.addWidget(self.Lb_av_range, 8, 3, 1, 2)
            grid2.addWidget(self.Sb_av_range, 8, 5)
            grid2.addWidget(self.Lb_diff_confi, 9, 3, 1, 2)
            grid2.addWidget(self.Sb_diff_confi, 9, 5)

            grid2.addWidget(self.Lb_hole_res_ring, 11, 1, 1, 2)
            grid2.addWidget(self.Lb_hole_res_max, 11, 4, 1, 2)

            # Some buttons
            grid2.addWidget(self.bt_plot_now, 1, 7)
            grid2.addWidget(self.bt_find_hole, 10, 3)
            grid2.addWidget(self.bt_export_now, 1, 11)

            # Showing the graphs
            grid2.addWidget(self.canvasIM2, 3, 6, 8, 7)
            grid2.addWidget(self.toolbarIM2, 11, 6, 1, 7)

            self.tab2.setLayout(grid2)
            # grid2.setRowStretch(2, 1)

        set_layout_of_tab1()
        set_layout_of_tab2()

        # Add tabs to widget
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.tabs)

        self.firstLoad = True
        self.tabs.currentChanged.connect(self.plot_graphs)

    def display_down(self):
        print('display_down called')

        if self.SlNum.value() > fromNo:
            self.SlNum.setValue(self.SlNum.value() - 1)

    def display_up(self):
        # print('display_up called')
        if self.SlNum.value() < toNo:
            self.SlNum.setValue(self.SlNum.value() + 1)

    def save_mask(self):
        with open(dirname + '/' + baseName + '_MCinfo.txt', 'w') as fi_handle:
            fi_handle.write("Mask and cropping information for this image stack\n\n")
            fi_handle.write("Crop margins (pixels):\n")
            fi_handle.write("\t{:5d}\t\n".format(self.SbUpper.value()))  # for safety use 5 digits
            fi_handle.write("{:5d}\t\t{:5d}\n".format(self.SbLeft.value(), self.SbRight.value()))
            fi_handle.write("\t{:5d}\t\n".format(self.SbLower.value()))
            # masked is a list of tuples
            fi_handle.write("\nCoordinates of masked pixels (pixels):\n")
            for masked_pixel in masked:
                fi_handle.write("{:5d}\t{:5d}\n".format(masked_pixel[0], masked_pixel[1]))
        return

    def load_mask(self):
        global masked
        fname = QFileDialog.getOpenFileName(
            self,
            "Open crop and mask information file",
            dirname,
            "All files (*_MCinfo.txt)")

        if not fname[0]:
            return  # function continues only if a file was chosen
        # print('File name: ', fname[0])

        with open(fname[0], 'r') as fi_handle:
            # instead of csv.reader : line = fi_handle.readline() and then
            # nb = int(line.split('\t')[1]) works too (see immergut).
            reader = csv.reader(fi_handle, delimiter='\t')
            for count in range(3):  # jump 3 lines
                _ = next(reader)

            line = next(reader)  # read crop margin values
            self.SbUpper.setValue(int(line[1]))
            line = next(reader)
            self.SbLeft.setValue(int(line[0]))
            self.SbRight.setValue(int(line[2]))
            line = next(reader)
            self.SbLower.setValue(int(line[1]))

            for count in range(2):  # jump 2 lines
                _ = next(reader)

            masked = []  # empty masked
            for row in reader:  # fill masked
                masked.append((int(row[0]), int(row[1])))
        return

    def Load_File1(self):
        """Is called by
        self.BtLoadFile1.clicked.connect(self.Load_File1)
        """
        print('Load_File1 called')

        # Deal with the file names of the stack of images
        global im1, Vmi1, Vma1
        global dirname, baseName, fromNo, toNo
        global imFromNo, imToNo, fromNoOld, toNoOld

        # choose an image of the stack using a dialog
        fname = QFileDialog.getOpenFileName(
            self,
            "Open file",
            "C:\_DATA\RECHERCHE\Experimental data\Contamination IBS coatings\_LIC Thèse\Test EC2216",
            "All files (*.*) ;; image files (*.tif *.tiff *.png)")

        if not fname[0]:
            return  # function continues only if a file was chosen
        print('File name: ', fname[0])

        # display the filename in the label
        self.LbFolderName.setText(fname[0])
        # publish the filename for use in other functions and tabs
        baseName = fname[0].split('/')[-1]  # just the file name not the path
        dirname = '/'.join(fname[0].split('/')[:-1])  # just the path
        fromNo = int(baseName.split('_')[-1].split('.')[0])
        self.EdFiStart.setText(str(fromNo))
        baseName = '_'.join(baseName.split('_')[:-1])  # all but the last _-section that contains the number
        self.LbBaseName.setText('Basename: ' + baseName)

        fiNames = []
        for fname in glob.glob(dirname + '/' + baseName + '*.tif'):
            fiNames.append(fname)

        fromNo = sys.maxsize
        toNo = -10
        for fname in fiNames:
            fname = fname.split('/')[-1]  # just the file name not the path
            Nb = int(fname.split('_')[-1].split('.')[0])
            if Nb > toNo:
                toNo = Nb
            if Nb < fromNo:
                fromNo = Nb
        self.EdFiStart.setText(str(fromNo))
        self.EdFiEnd.setText(str(toNo))
        imFromNo = fromNo  # never change it's the min for the stack
        imToNo = toNo  # never change it's the max for the stack
        fromNoOld = imToNo + 1  # initialize that the full stack is loaded at the first execution of load_stack
        toNoOld = imFromNo - 1  # initialize that the full stack is loaded at the first execution of load_stack
        self.SlNum.setValue(fromNo)  # displays the first image of the stack
        self.set_stack_limits()  # this calls load_stack

    def load_stack(self, uncropped=False):
        # print('load_stack called')

        # Only reloads if enlarged range; if smaller range, it makes a 'vertical crop'
        global imStack
        global im1_width, im1_height # size of the original images
        global fromNoOld, toNoOld
        global roiS
        global masked, mask_during_modif

        print("load_stack called by ")

        print("sender =", self.sender())
        print("sender.__name = ", str(self.sender().accessibleName()))

        if not uncropped and (fromNo >= fromNoOld) and (toNo <= toNoOld):  # It's a "vertical crop"
            imStack = imStack[:, :, (fromNo - fromNoOld):(fromNo - fromNoOld) + (toNo - fromNo + 1)]
        else:
            main.statBar.showMessage('Reading images... please wait')
            im = plt.imread(dirname + '/' + baseName + '_' + str(fromNo) + '.tif')
            im1_width = im.shape[1]
            im1_height = im.shape[0]
            self.LbInfo.setText("For info: The image shape is (x, y) = (" + str(im1_width) + ", " +
                                str(im1_height) + ") and the image numbers range from " + str(imFromNo) +
                                " to " + str(imToNo) + ".")
            imStack = np.ndarray(im.shape + (toNo - fromNo + 1,))  # reserve memory for stack
            imStack[:, :, 0] = im
            vMaxMax = 0  # for setting the maximum value of the vertical scale spin boxes
            main.progBar.setMinimum(fromNo)
            main.progBar.setMaximum(toNo)
            for Nb in range(fromNo + 1, toNo + 1):
                # loading the images
                imStack[:, :, Nb - fromNo] = plt.imread(dirname + '/' + baseName + '_' + str(Nb) + '.tif')
                main.progBar.setValue(Nb)

                # for setting the maximum value of the vertical scale spin boxes
                vMaxNow = imStack[:, :, Nb - fromNo].max()
                if vMaxNow > vMaxMax:
                    vMaxMax = vMaxNow
                QApplication.processEvents()
            main.statBar.showMessage('Finished loading images', 2000)  # the message shows 2000 ms and disappears
            main.progBar.setValue(fromNo)

            # Execute this only one time, if not, a user setting will be lost on reloading the stack.
            if self.firstLoad:
                # setting the maximum value of the vertical scale spin boxes
                self.SbVmin.setRange(0, vMaxMax)
                self.SbVmax.setRange(0, vMaxMax)
                self.SbVmin.setSingleStep(vMaxMax // 100)
                self.SbVmax.setSingleStep(vMaxMax // 100)
                self.SbVmax.setValue(vMaxMax)
                self.firstLoad = False

                roiS = np.ones((4, 4))*np.nan  # xfrom, xto, yfrom, yto (include all indices)
                # has to be float if not, there is no "nan"
                # first index = RoiNo, second index = xfrom, xto, yfrom, yto
                masked = []
                mask_during_modif = False

            if self.CbCrop.isChecked():  # keep the cropping, if it was on
                imStack = imStack[self.SbUpper.value():(im1_height - 1 - self.SbLower.value()),
                          self.SbLeft.value():(im1_width - 1 - self.SbRight.value()),
                          :]
                roiS = np.zeros((4, 4))  # xfrom, xto, yfrom, yto (include all indices)
                roiS[0, :] = [0, imStack[:, :, 0].shape[1]-1, 0, imStack[:, :, 0].shape[0]-1]  # roi1
                roiS[1, :] = [0, imStack[:, :, 0].shape[1]-1, 0, imStack[:, :, 0].shape[0]-1]  # roi2
                roiS[3, :] = [0, imStack[:, :, 0].shape[1] - 1, 0, imStack[:, :, 0].shape[0] - 1]  # roi4

            else:  # if cropping was off, just display and reset the spinbox ranges
                self.SbUpper.setRange(0, im1_height - 1)
                self.SbLower.setRange(0, im1_height - 1)
                self.SbLeft.setRange(0, im1_width - 1)
                self.SbRight.setRange(0, im1_width - 1)

            self.showRoiS()
            self.display()
        fromNoOld = fromNo
        toNoOld = toNo

    def showRoiS(self):  # only ROI 1 and 2, ROI 4 is centered around ROI 2
        # roi1
        self.lbROI1ul.setText("x: {:.0f}; y: {:.0f}".format(roiS[0,0],roiS[0,2]))
        self.lbROI1lr.setText("x: {:.0f}; y: {:.0f}".format(roiS[0,1],roiS[0,3]))
        # roi2
        self.lbROI2ul.setText("x: {:.0f}; y: {:.0f}".format(roiS[1,0],roiS[1,2]))
        self.lbROI2lr.setText("x: {:.0f}; y: {:.0f}".format(roiS[1,1],roiS[1,3]))

    def set_stack_limits(self):
        """Is called by
        EdFiStart.returnPressed.connect(self.set_stack_limits)
        EdFiEnd.returnPressed.connect(self.set_stack_limits)
        """
        print('set_stack_limits called')
        global fromNo, toNo
        try:
            fromNo = int(self.EdFiStart.text())
            toNo = int(self.EdFiEnd.text())
        except ValueError:
            return
        if (fromNo < imFromNo) or (toNo > imToNo):
            return  # input exceeds stack range
        # only continue if both limits are readable (this is the case when called by Load_file1)
        self.SlNum.setMinimum(fromNo)
        self.SlNum.setMaximum(toNo)
        main.progBar.setMinimum(fromNo)  # if it's not done here, it has to be done at different places later on
        main.progBar.setMaximum(toNo)
        self.load_stack()

    def set_crop_margins(self):
        """Is called by
        self.SbLeft.valueChanged.connect(self.set_crop_margins)
        and the other crop margin spin boxes
        """
        print('set_crop_margins called')

        # This makes sure that inversions are impossible upper/lower left/right
        if self.sender() is self.SbUpper:
            self.SbLower.setRange(0, im1_height - 1 - self.sender().value())
        elif self.sender() is self.SbLower:
            self.SbUpper.setRange(0, im1_height - 1 - self.sender().value())
        elif self.sender() is self.SbLeft:
            self.SbRight.setRange(0, im1_width - 1 - self.sender().value())
        elif self.sender() is self.SbRight:
            self.SbLeft.setRange(0, im1_width - 1 - self.sender().value())
        self.display()

    def display(self):
        """Is called by
        self.SlNum.valueChanged.connect(self.display)
        and many others to display the present image of the stack
        """
        # print('display called')
        try:
            if self.rdbtgroupMedFi.checkedId() == 1:  # Median filter preview is active
                self.display_MedFi_preview()
            else:  # just show the stack
                number = self.SlNum.value()  # read present slider value
                self.LbImNo.setText(str(number))  # display it in the label below the image
                ImDisp1 = imStack[:, :, number - fromNo]  # make a view of the right image
                self.showIm1(ImDisp1)  # show the image
        except NameError:  # necessary because the method is executed on initialisation
            return

    def showIm1(self, im_to_show):
        # print('showIm1 called')
        global axIM1

        # TODO: Accelerate this method
        # Clear the figure from earlier uses
        self.figureIM1.clear()
        # prepare the axis
        axIM1 = self.figureIM1.add_subplot(111)
        axIM1.axis("off")
        if self.CbUseVfixed.isChecked():
            vmi1 = self.SbVmin.value()
            vma1 = self.SbVmax.value()
        else:
            vmi1 = im_to_show.min()
            vma1 = im_to_show.max()
            
        # mark masked pixels if wished
        im_modified = im_to_show.copy()  # modify a copy of im_to_show
        if self.cbMaskShow.isChecked():
            # calculate the values of the color switch limits and the
            # replacement values (for the masked pixels to become visible)
            MiniVal = np.min(im_to_show)
            MaxiVal = np.max(im_to_show)
            infernoLimits = [0.35, 0.82]
            infernoColVals = [0.66, 1]
            # change the values (colors) of the pixels in the list
            for count in range(len(masked)):
                IsVal = im_to_show[masked[count][1], masked[count][0]]
                if infernoLimits[0] < IsVal / (MaxiVal - MiniVal) < infernoLimits[1]:  # light color
                    putVal = infernoColVals[1] * (MaxiVal - MiniVal) + MiniVal
                else:
                    putVal = infernoColVals[0] * (MaxiVal - MiniVal) + MiniVal
                im_modified[masked[count][1], masked[count][0]] = putVal
        
        cax = axIM1.imshow(im_modified, cmap=self.CoboCmap.currentText(), vmin=vmi1, vmax=vma1)
        # common color maps: jet, gray, hot, hsv, inferno, gist_ncar
        self.figureIM1.colorbar(cax, orientation='vertical')
        self.figureIM1.tight_layout()

        self.UpdateLinesIm1()

        # refresh canvas
        self.canvasIM1.draw()

    def UpdateLinesIm1(self, xM =-10, yM = -10):
        global axIM1
        global roiS
        # print("UpdateLinesIm1 called")

        # delete lines on axis (if there were any)
        try:
            axIM1.lines = []
        except NameError:  # just a guess (that seems to work)
            pass

        # Draw white lines of crop limits if crop inactive
        if not self.CbCrop.isChecked():
            linew = 1
            col = 'w'
            ulx = self.SbLeft.value()
            lrx = im1_width - 1 - self.SbRight.value()
            uly = self.SbUpper.value()
            lry = im1_height - 1 - self.SbLower.value()
            axIM1.plot([ulx, lrx, lrx, ulx, ulx], [uly, uly, lry, lry, uly],
                       color=col, linestyle='--', linewidth=linew)

        maxX = imStack[:, :, 0].shape[1]-1
        maxY = imStack[:, :, 0].shape[0]-1

        if xM == -10 and yM == -10:  # was not called by mouse event
            for roiNum in [0, 1, 3]:  # possibly draw existing rois
                if np.any(np.isnan(roiS[roiNum, :])) or (roiS[roiNum, 1] > maxX) or (roiS[roiNum, 3] > maxY):
                    continue  # do not show this roi

                if roiNum == 0:
                    col = 'r'
                elif roiNum == 1:
                    col = 'b'
                elif roiNum == 3:
                    col = 'k'
                ulx = roiS[roiNum, 0]
                lrx = roiS[roiNum, 1]
                uly = roiS[roiNum, 2]
                lry = roiS[roiNum, 3]
                axIM1.plot([ulx, lrx, lrx, ulx, ulx], [uly, uly, lry, lry, uly],
                           color=col, linestyle='--', linewidth=1)
            return  # do not draw the lines depending on the mouse position

        # was called by mouse event: draw ROI
        if xM is not None and 0 <= xM <= maxX and yM is not None and 0 <= yM <= maxY:  # mouse in crop limits
            # print('** xM = ', xM, ', yM = ', yM)
            linew = 1
            if roiNo == 0:
                col = 'r'
            elif roiNo == 1:
                col = 'b'
            elif roiNo == 2:
                col = 'k'

            if np.isnan(roiS[roiNo,0]) :
                # add the cross
                ulx = xM
                uly = yM
                axIM1.plot([ulx, ulx], [0, maxY], color=col, linestyle='--', linewidth=linew)
                axIM1.plot([0, maxX], [uly, uly], color=col, linestyle='--', linewidth=linew)
            else:  # ul edge is already chosen
                # add the frame
                # use (ulx, uly) and (lrx, lry) 
                if xM >= roiS[roiNo, 0]:
                    ulx = roiS[roiNo, 0]
                    lrx = xM
                else:
                    ulx = xM
                    lrx = roiS[roiNo, 0]
                if yM >= roiS[roiNo, 2]:
                    uly = roiS[roiNo, 2]
                    lry = yM
                else:
                    uly = yM
                    lry = roiS[roiNo, 2]
                axIM1.plot([ulx, lrx, lrx, ulx, ulx],
                           [uly, uly, lry, lry, uly], color=col, linestyle='--', linewidth=1)
                # ca m'a l'air plus joli, mais c'est seulement légèrement plus rapide

    def modify_mask(self):
        """ Adds or removes the third line of 'roiS' to 'masked'
        """
        global masked  # list of tuples

        # Generate list of selected points
        points_new = []  # list of tuples
        ulx = int(np.round(roiS[roiNo, 0]))
        lrx = int(np.round(roiS[roiNo, 1]))
        uly = int(np.round(roiS[roiNo, 2]))
        lry = int(np.round(roiS[roiNo, 3]))
        # fill columnwise
        for xCoo in range(ulx, lrx + 1):  # from left to right
            # column loop
            for yCoo in range(uly, lry + 1):  # from upper to lower
                # line loop
                points_new.append((xCoo, yCoo))  # (x, y) coordinates

        # add or remove points_new to masked
        if self.CoboMaskMode.currentText() == "Add pixels":
            #             # add really new points to MaskList
            masked = list(set(masked + points_new))  # sorts at the same time
        elif self.CoboMaskMode.currentText() == "Remove pixels":
            # remove entries of newPoints that existed in MaskList
            masked = list(set(masked) - set(points_new))  # sorts at the same time

        print('len masked = ', len(masked))
        self.lbMaskInfo.setText("{:.0f} pixels masked".format(len(masked)))
        return

    def roi_def(self):  # called by checkbox 
        print("roi1_def called")
        global roiS, roiNo
        if self.cbROI1redefine.isChecked():
            roiNo = 0  # first line in roi array  
        elif self.cbROI2redefine.isChecked():
            roiNo = 1  # second line in roi array
            roiS[3, :] = np.ones(4) * np.nan  # link new ROI to ROI[1]
        elif self.cbMaskActivateIt.isChecked():
            roiNo = 2  # third line in roi array
        # if one cb was checked set the corresponding mine to nan
        roiS[roiNo, :] = np.ones(4)*np.nan
        self.showRoiS()

    def plot_graphs(self):
        global stack_vals
        global imNoS
        # delete lines on axis (if there were any)
        try:
            self.axIM2.lines = []
        except (NameError, AttributeError):  # just a guess (that seems to work)
            pass

        imNoS = np.array(range(fromNo, toNo + 1))  # x-axis
        stack_vals = np.zeros((len(imNoS), self.tablines))  # for the data
        for line in range(self.tablines):  # loop through lines of table
            # Create boolean mask 'ma' using 'Region', 'Out/In' and 'Mask' information
            if self.TbTasks.cellWidget(line, 0).currentText() == "Full frame":
                ma = np.ones(imStack[:, :, 0].shape, dtype=bool)
            else:
                ma = np.zeros(imStack[:, :, 0].shape, dtype=bool)  # use a roi
                # ma de base
                if (self.TbTasks.cellWidget(line, 0).currentText() == "Roi 1") and not np.any(np.isnan(roiS[0])):
                    roi_li = np.round(roiS[0]).astype(int)  # extract good roi line and convert to int: background
                elif (self.TbTasks.cellWidget(line, 0).currentText() == "Roi3 - Roi2") and not np.any(np.isnan(roiS[3])):
                    roi_li = np.round(roiS[3]).astype(int)  # enlarged max position
                elif not np.any(np.isnan(roiS[1])):
                    roi_li = np.round(roiS[1]).astype(int)  # max position
                else:
                    self.lb_info_2.setText("The ROI is not defined")
                    continue  # this line uses a roi that is not defined
                # count roi borders as inside
                ma[roi_li[2]:roi_li[3]+1, roi_li[0]:roi_li[1]+1] = np.ones(
                    (roi_li[3]-roi_li[2]+1, roi_li[1]-roi_li[0]+1), dtype=bool)
                if self.TbTasks.cellWidget(line, 0).currentText() == "Roi3 - Roi2":
                    # remove inner part (Roi2)
                    roi_li = np.round(roiS[1]).astype(int)  # max position
                    ma[roi_li[2]:roi_li[3] + 1, roi_li[0]:roi_li[1] + 1] = np.zeros(
                        (roi_li[3] - roi_li[2] + 1, roi_li[1] - roi_li[0] + 1), dtype=bool)
                # print("taille ma de base: ", np.sum(ma))
                # outside?
                if self.TbTasks.cellWidget(line, 1).currentText() == "Outside":
                    ma = ~ma  # or use  np.logical_not
                # print("taille ma apres out in: ", np.sum(ma))
            if self.TbTasks.cellWidget(line, 2).currentText() == "Ignore masked pixels":  # remove masked pixels from ma
                for masked_pixel in masked:  # would a try: except: continue slow down a lot? ? put for safety
                    ma[masked_pixel[1], masked_pixel[0]] = False
            # print("taille ma apres tout: ", np.sum(ma))

            # Evaluate values (later also positions) using 'Value':"Mean Value", "Max Value", "Min Value", "Max Position"
            if self.TbTasks.cellWidget(line, 3).currentText() == "Mean Value":  # put in loop if it does not work here
                func = np.mean
            elif self.TbTasks.cellWidget(line, 3).currentText() == "Max Value":
                func = np.max
            elif self.TbTasks.cellWidget(line, 3).currentText() == "Min Value":
                func = np.min
            elif self.TbTasks.cellWidget(line, 3).currentText() == "Median Value":
                func = np.median

            # stack_vals = np.zeros((len(imNoS),self.tablines))  # for the data
            for imNo in imNoS:
                stack_vals[imNo - fromNo, line] = func(imStack[:, :, imNo - fromNo][ma])  # each tab line makes a col

            # plot labelled with line 1, line 2, etc.... if wanted
            if self.TbTasks.cellWidget(line, 4).currentText() == "Yes":  # plot the data
                self.axIM2.plot(imNoS, stack_vals[:, line], label="Line {:d}".format(line+1))

        self.axIM2.legend()
        self.axIM2.set_title("Raw data ")  # add info on test number?
        self.axIM2.set_ylabel('Grey level (saturates at 4095)')
        self.axIM2.set_xlabel('Image number in stack')
        self.figureIM2.tight_layout()
        # refresh canvas
        self.canvasIM2.draw()

    def moving_average(self, x, y, box_pts):
        """moving average, use an odd nb of box_pts."""
        if box_pts > 1:
            if box_pts % 2 == 0:
                box_pts += 1  # make box_points odd
            y_smooth = np.cumsum(y, dtype=float)
            y_smooth[box_pts:] = y_smooth[box_pts:] - y_smooth[:-box_pts]
            return x[box_pts // 2:-(box_pts // 2)], y_smooth[box_pts - 1:] / box_pts
        else:
            return x, y

    def find_hole(self):
        global mess_ring_mean, mess_glob_max
        global hole_start_ringmean, hole_start_maxpos
        ### 1. compare the ring mean to the center mean
        # use data in stack_vals
        # col 6 (line 6 in table) = background noise mean
        # col 5 (line 5 in table) = ring mean
        # col 3 (line 3 in table) = center mean
        bg_vals = stack_vals[:, 6]
        ring_vals = stack_vals[:, 5]
        center_vals = stack_vals[:, 3]
        # first subtract background
        ring_vals -= bg_vals
        center_vals -= bg_vals
        # then smooth
        imNoS_sm, ring_vals_sm = self.moving_average(imNoS, ring_vals, self.Sb_smoothing_intv.value())
        imNoS_sm, center_vals_sm = self.moving_average(imNoS, center_vals, self.Sb_smoothing_intv.value())

        if ring_vals_sm[-1] < center_vals_sm[-1]:
            mess_ring_mean = "No hole found"
            self.Lb_hole_res_ring.setText(mess_ring_mean)
        else:
            idx = np.where(ring_vals_sm > center_vals_sm)[0][0]  # first index where condition is true
            hole_start_ringmean = imNoS_sm[idx]
            self.Lb_hole_res_ring.setText("Hole starts in image no {:d}".format(hole_start_ringmean))

        ### 2. compare compare the center max to global max
        # use data in stack_vals
        # col 1 (line 1 in table) = center max
        # col 0 (line 0 in table) = global max
        # do not use moving_average function as we need noise information to compare
        wing = (self.Sb_av_range.value() + 2)//2
        imNo = toNo - wing
        _, p2 = st.ttest_ind(stack_vals[imNo-wing:imNo+wing+1, 0],
                             stack_vals[imNo-wing:imNo+wing+1, 1], equal_var=True)
        # if p2 is small then the samples are different (with confidence 1 - p2_limit)
        # print("\nimNo ", imNo)
        # print("a ", stack_vals[imNo-wing:imNo+wing+1, 0])
        # print("b ", stack_vals[imNo - wing:imNo + wing + 1, 1])
        # print("p2 ", p2)

        confidence = self.Sb_diff_confi.value()  # confidence for the means to be different
        p2_limit = 1 - confidence
        if p2 > p2_limit:
            mess_glob_max = "No hole found"
            self.Lb_hole_res_max.setText(mess_ring_mean)
            return
        while (imNo > fromNo + wing) and (p2 < p2_limit):
            imNo -= 1
            _, p2 = st.ttest_ind(stack_vals[imNo - wing:imNo + wing + 1, 0],
                                 stack_vals[imNo - wing:imNo + wing + 1, 1], equal_var=True)
        # print("imNo2 ", imNo)

        if p2 >= p2_limit:
            hole_start_maxpos = imNo
            self.Lb_hole_res_max.setText("Hole starts in image no {:d}".format(hole_start_maxpos))
        else:
            mess_glob_max = "Hole from he start"
            self.Lb_hole_res_max.setText(mess_glob_max)
        return

    def export_graphs(self):
        # teno = np.nan
        # dirname.split("/")[-1]
        # simple version
        mywb = xl.Workbook()  # workbook
        ws1 = mywb.active  # worksheet
        ws1.title = dirname.split("/")[-1][-7:-1]

        # Header
        ws1.cell(row=1, column=1).value = 'Folder: ' + dirname.split("/")[-1]
        print(ws1.cell(row=1, column=1).value )

        ws1.cell(row=1, column=3).value = 'Analyzed on: '
        ws1.cell(row=1, column=4).value = str(ti.date.today())

        ws1.cell(row=1, column=9).value = 'Hole start found: '
        ws1.cell(row=2, column=9).value = 'Ring mean vs. center mean : '
        if hole_start_ringmean >= 0:
            ws1.cell(row=2, column=10).value = hole_start_ringmean
        else:
            ws1.cell(row=2, column=10).value = mess_ring_mean
        ws1.cell(row=3, column=9).value = 'global max vs. center max : '
        if hole_start_ringmean >= 0:
            ws1.cell(row=2, column=10).value = hole_start_maxpos
        else:
            ws1.cell(row=2, column=10).value = mess_glob_max

        ws1.cell(row=2, column=1).value = 'Peak fluence: '
        ws1.cell(row=2, column=3).value = 'J/cm2'

        ws1.cell(row=3, column=1).value = 'Plot line: '
        ws1.cell(row=4, column=1).value = 'ROI: '
        ws1.cell(row=5, column=1).value = 'Out/in: '
        ws1.cell(row=6, column=1).value = 'Masked?: '
        ws1.cell(row=7, column=1).value = 'Function: '
        ws1.cell(row=8, column=1).value = 'Image no \\'
        for imNo in imNoS:  # x-axis
            ws1.cell(row=imNo - fromNo + 9, column=1).value = imNo

        def make_leg_entry(line):
            entry = ""
            if self.TbTasks.cellWidget(line, 0).currentText() == "Full frame":
                entry += "full NA "
            else:
                if self.TbTasks.cellWidget(line, 0).currentText() == "Roi 1":
                    entry += "roi1"
                elif self.TbTasks.cellWidget(line, 0).currentText() == "Roi 2":
                    entry += "roi2"
                else:  # "Roi3 - Roi2"
                    entry += "roi3-2"
                # outside?
                if self.TbTasks.cellWidget(line, 1).currentText() == "Outside":
                    entry += " out"
                else:
                    entry += " in "

            # add masking info
            if self.TbTasks.cellWidget(line, 2).currentText() == "Ignore masked pixels":  # remove masked pixels from ma
                entry += " mas"
            else:
                entry += " all"

            # add function info
            if self.TbTasks.cellWidget(line, 3).currentText() == "Mean Value":  # put in loop if it does not work here
                entry += " mea"
            elif self.TbTasks.cellWidget(line, 3).currentText() == "Max Value":
                entry += " max"
            elif self.TbTasks.cellWidget(line, 3).currentText() == "Min Value":
                entry += " min"
            elif self.TbTasks.cellWidget(line, 3).currentText() == "Median Value":
                entry += " med"
            return entry

        for line in range(self.tablines):  # loop through lines of table
            if self.TbTasks.cellWidget(line, 5).currentText() == "No":  # do not export this data set
                continue

            ws1.cell(row=3, column=2+line).value = 'Line {:d}'.format(line+1)  # write line number
            for item_no in range(1, 5):  # write task table info of combo-boxes
                ws1.cell(row=3+item_no, column=2+line).value = self.TbTasks.cellWidget(line, item_no).currentText()
            ws1.cell(row=8, column=2+line).value = make_leg_entry(line)  # write legend entry (summary of above)

            for count in range(len(imNoS)):
                ws1.cell(row=9+count, column=2+line).value = stack_vals[count, line]

        # open a dialog for input of the filename
        try:
            fname = QFileDialog.getSaveFileName(self,
                                                'Save data to file',
                                                '../output/',
                                                "All files (*.*) ;; Excel files (*.xlsx *.xls)")
            # save the workbook
            if fname[0]:  # a file was chosen
                mywb.save(fname[0])
        except:
            pass
        return

    def GetMouseIm1(self, event):
        # Manages the drawing of the lines depending on the mouse position

        # useless events
        if event.inaxes is None:
            return
        # Cases where nothing should be done = inverse of cases of action
        if not(self.cbROI1redefine.isChecked()
               or self.cbROI2redefine.isChecked()
               or self.cbMaskActivateIt.isChecked()):
            return

        # useful events
        self.UpdateLinesIm1(int(np.round(event.xdata)),
                            int(np.round(event.ydata)))
        self.canvasIM1.draw()

    def GetKeyIm1(self, event):
        print("GetKeyIm1 called")
        # Manipulates the Mask (Pixels not to use for the fit evaluation)
        # h : upper left corner
        # n : lower right corner
        # b : single pixel in mask mode
        global roiS  # for keeping the memory between two calls:
        # roiS un champ 3x4
        # A line contains: x_min, x_max, y_min, y_max of the ROI if it is defined
        # Nan after ticking the cb
        # x_min, nan, y_min, nan during definition
        # the choice of the line is done using: if elif elif sur les cb.

        print('Im1: ', event.key)
        #print('Im1: ', event.ydata)

        # useless events
        if event.inaxes is None: return
        # Cases where nothing should be done = inverse of cases of action
        if not(self.cbROI1redefine.isChecked() or self.cbROI2redefine.isChecked()
               or self.cbMaskActivateIt.isChecked()): return

        letter = str(event.key).upper()
        if letter == 'H':
            # memorize first point
            roiS[roiNo, 0] = np.round(event.xdata)
            roiS[roiNo, 2] = np.round(event.ydata)
            # reset second point
            roiS[roiNo, 1] = np.nan
            roiS[roiNo, 3] = np.nan
            self.showRoiS()
            finished = False
        if letter == 'B':
            # memorize the single selected point
            roiS[roiNo, 0] = np.round(event.xdata)
            roiS[roiNo, 2] = np.round(event.ydata)
            # set second point identical to first one
            roiS[roiNo, 1] = np.round(event.xdata)
            roiS[roiNo, 3] = np.round(event.ydata)
            self.showRoiS()
            finished = True
        elif letter == 'N':
            # memorize second point
            roiS[roiNo, 1] = np.round(event.xdata)
            roiS[roiNo, 3] = np.round(event.ydata)

            if not(np.isnan(roiS[roiNo, 0]) or np.isnan(roiS[roiNo, 2])): # the rectangle gets finished
                # flip points if necessary
                if roiS[roiNo, 1] < roiS[roiNo, 0]:  # right x should be larger than left x
                    buffer = roiS[roiNo, 0]
                    roiS[roiNo, 0] = roiS[roiNo, 1]
                    roiS[roiNo, 1] = buffer
                if roiS[roiNo, 3] < roiS[roiNo, 2]:  # lower y should be larger than upper y
                    buffer = roiS[roiNo, 2]
                    roiS[roiNo, 2] = roiS[roiNo, 3]
                    roiS[roiNo, 3] = buffer

                self.showRoiS()
                finished = True
        # end letter 'N'

        if finished:
            # Tell other function to no longer update the rectangle drawing
            if roiNo == 0: # first line in roi array
                self.cbROI1redefine.setChecked(False)
            elif roiNo == 1:  # second line in roi array
                self.cbROI2redefine.setChecked(False)
                # Set roi[3] here
                # A line contains: x_min, x_max, y_min, y_max of the ROI if it is defined
                roiS[3, 0] = roiS[roiNo, 0] - roi_3_width
                roiS[3, 1] = roiS[roiNo, 1] + roi_3_width
                roiS[3, 2] = roiS[roiNo, 2] - roi_3_width
                roiS[3, 3] = roiS[roiNo, 3] + roi_3_width
            elif roiNo == 2:  # third line in roi array
                self.modify_mask()  # modifies the list of masked points
                self.cbMaskActivateIt.setChecked(False)

        print(roiS[roiNo, :])

    def crop_or_uncrop(self):
        """Is called by
        self.CbCrop.stateChanged.connect(self.crop_or_uncrop)
        """
        print('crop_or_uncrop called')

        global imStack
        if self.CbCrop.isChecked():  # take a view
            imStack = imStack[self.SbUpper.value():(im1_height - 1 - self.SbLower.value()),
                              self.SbLeft.value():(im1_width - 1 - self.SbRight.value()),
                              :]
            self.display()
        else:
            self.load_stack(uncropped=True)  # reload the full field images

    def medFi_state_changed(self):
        global oldStack, imStack
        newBtIdx = self.rdbtgroupMedFi.checkedId()
        if newBtIdx == 1:  # preview became active
            self.display_MedFi_preview()
        elif newBtIdx == 2:  # Apply filter to stack (but backup before)
            oldStack = imStack.copy()  # make a backup
            MedFiPara = self.SbMedFi.value()
            main.statBar.showMessage('Processing ...  please wait')
            for Nb in range(fromNo, toNo + 1):
                imStack[:, :, Nb - fromNo] = ndi.median_filter(imStack[:, :, Nb - fromNo], MedFiPara)
                main.progBar.setValue(Nb)
                QApplication.processEvents()
            main.statBar.showMessage('Finished', 2000)
            main.progBar.setValue(fromNo)
            # main.statBar.showMessage('Ready')
            self.display()
        elif newBtIdx == 0:  # revert to earlier stage if possible
            try:
                imStack = oldStack.copy()
                self.display()
            except NameError:  # necessary because the method is executed on initialisation
                return

    def medFi_sb_changed(self):
        if self.rdbtgroupMedFi.checkedId() == 1:
            self.display_MedFi_preview()

    def display_MedFi_preview(self):
        print('display_MedFi_preview called')
        try:
            number = self.SlNum.value()  # read present slider value
            self.LbImNo.setText(str(number))  # display it in the label below the image

            if self.SbMedFi.value() > 1:
                main.statBar.showMessage('Processing')
                ImDisp1 = ndi.median_filter(imStack[:, :, number - fromNo], self.SbMedFi.value())
                main.statBar.showMessage('Ready')
            else:
                ImDisp1 = imStack[:, :, number - fromNo]
            self.showIm1(ImDisp1)  # show the image
        except NameError:
            return


####################################################
if __name__ == '__main__':
    # app = QApplication(sys.argv) # std : create QApplication
    app = QApplication.instance()  # checks if QApplication already exists
    if not app:  # create QApplication if it doesnt exist
        app = QApplication(sys.argv)
    app.aboutToQuit.connect(app.deleteLater)

    main = MainWindow()
    # print(dir(main)) # for debugging
    main.show()

    # sys.exit(app.exec_()) # std : exits python when the app finishes
    app.exec_()  # do not exit Ipython when the app finishes

"""
        # prepare checking for saturation
        im1 = plt.imread(fname[0])
        print('The file data type is ', im1.dtype, ' and the shape is ',im1.shape)
        bits = ''.join([c for c in str(im1.dtype) if c in '1234567890.'])
        maxZ = 2**int(bits) - 1 # np.iinfo(im1.dtype).max() ne marche pas

        ### check for saturation but do not normalize (height values etc)
        print('Maximum pixel value of the image: ', im1.max() )
        print('Maximum pixel value of the input dtype: ', maxZ )
        if im1.max() >= maxZ :
            message = 'THIS IMAGE IS SATURATED!'
            print(message)
            # display message
            self.LbInfo12.setText(message)
        else:
            message = 'The image is not saturated!'
            print(message)
            # display message
            # self.LbInfo12.setText(message)
        print(' ')
"""
